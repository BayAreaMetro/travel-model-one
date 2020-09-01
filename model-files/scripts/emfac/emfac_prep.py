# this script is run after the Cube scripts for EMFAC


import pandas as pd
import numpy as np

import os.path
from os import path

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# -------------------------------------------------------------------
# Input/output file names and locations
# -------------------------------------------------------------------
run_id = os.getenv('myfolder')

emfac_input_xlsx = os.getenv('emfac_input_template')
emfac_input_xlsx_fullpath = "emfac_prep\\" + emfac_input_xlsx

output_excel_template = "emfac_prep\\ready4emfac_" + run_id + ".xlsx"
output_filename1 = "emfac_prep\\vmt_b4reshape.csv"
output_filename2 = "emfac_prep\\vmt.csv"
output_filename = "emfac_prep\\defaulthourlyfraction.csv"
output_filename3 = "emfac_prep\\modifiedhourlyfraction.csv"
output_filename5 = "emfac_prep\\defaultVMT.csv"
output_filename6 = "emfac_prep\\defaultVMT_calcs.csv"
output_filename7 = "emfac_prep\\modelled_vmt.csv"
output_filename8 = "emfac_prep\\temp.csv"
output_filename9 = "emfac_prep\\temp2_split.csv"
output_filename10 = "emfac_prep\\temp_solano.csv"
output_filename11 = "emfac_prep\\temp_sonoma.csv"
output_DefaultModelled = "emfac_prep\\temp_DefaultModelled.csv"
output_checkModelled = "emfac_prep\\temp_checkModelled.csv"
output_debug1 = "emfac_prep\\temp_veh_tech.csv"


# -------------------------------------------------------------------
# Read the output csv files from Cube (CreateSpeedBinsBetweenZones and CreateSpeedBinsWithinZones) and reshape them
# -------------------------------------------------------------------
print "\n\n====================================================="
print "Reading modelled VMT csv files from Cube scripts"
print "====================================================="

# read in between zones VMT and intrazonal VMT
BetweenZones_df = pd.read_csv(os.path.join(os.getcwd(), "emfac_prep", "CreateSpeedBinsBetweenZones_sums.csv"), index_col=False, sep=",")
WithinZones_df = pd.read_csv(os.path.join(os.getcwd(), "emfac_prep", "CreateSpeedBinsWithinZones_sums.csv"), index_col=False, sep=",")

# what is the format of these VMT files?
# rows    = countyName (9 counties + 1 external) x speed bin (13) + header (1) = 10 x 13 + 1 = 131
# (in the future, rows = AirBasinName (11 airbains + 1 external) x speed bin (18) + header (1) = 12 x 18 + 1 = 217)
# columns = 24 hours + 3 index columns (countyName, arbCounty, speedBin) = 27

# if it's the old format, rename countyName to subareaName
numLine_Between = len(BetweenZones_df.index)
print "\nFinished reading CreateSpeedBinsBetweenZones_sums.csv"
print "Number of rows in CreateSpeedBinsBetweenZones_sums.csv = " + str(numLine_Between)

numLine_Within = len(WithinZones_df.index)
print "\nFinished reading CreateSpeedBinsWithinZones_sums.csv"
print "Number of rows in CreateSpeedBinsWithinZones_sums.csv = " + str(numLine_Within)

if numLine_Between+numLine_Within == 260:
    BetweenZones_df.rename(columns={"countyName": "subareaName"}, inplace=True)
    WithinZones_df.rename(columns={"countyName": "subareaName"}, inplace=True)

print "\nStart reshaping the modelled VMT data"
# add the between zones VMT and intrazonal VMT together
VMT_df = BetweenZones_df.add(WithinZones_df, fill_value=0)

# fix the index columns (countyName, arbCounty, speedBin) - countyName is messed up but fixing it is not a priority
VMT_df[' arbCounty'] = VMT_df[' arbCounty']/2
VMT_df[' speedBin'] = VMT_df[' speedBin']/2

VMT_df['nameLen'] = (VMT_df['subareaName'].str.len() / 2).astype(int) # temporary column
# VMT_df['subareaName'] = VMT_df['subareaName'].str[:VMT_df['nameLen']]  # this way of slicing does not work
# VMT_df['subareaName'] = VMT_df['subareaName'].str.slice(start=-VMT_df['nameLen']) # this way of slicing does not work either
VMT_df['subareaName'] = VMT_df.apply(lambda r: r.subareaName[:r.nameLen], axis=1) # this way of slicing worked

# drop the temporary nameLen column
VMT_df.drop(['nameLen'], axis=1, inplace=True)

# drop external VMT
VMT_df = VMT_df.loc[VMT_df[' arbCounty'] != 9999]

VMT_df.to_csv(output_filename1, header=True, index=False)

# reshape the data frame so that:
# The rows are subarea (11) x hour (24)
# The columns are the 18 speed bins

# create dataframe for each subsarea and loop through tem
# store your dataframes into a dictionary

# if the cube outputs only have 13 speed bins (i.e. it is an old run without air basins info), use the 9 county list. Otherwise, use the 11 subarea list.
if numLine_Between+numLine_Within == 260:
    subareaList = ['Alameda', 'Contra Costa', 'Marin', 'Napa', 'San Francisco', 'San Mateo', 'Santa Clara', 'Solano', 'Sonoma']
else:
    subareaList = ['Alameda (SF)', 'Contra Costa (SF)', 'Marin (SF)', 'Napa (SF)', 'San Francisco (SF)', 'San Mateo (SF)', 'Santa Clara (SF)', 'Solano (SF)', 'Solano (SV)', 'Sonoma (NC)', 'Sonoma (SF)']

dict_of_df = {}
for subarea in subareaList:

    key_name = 'VMT_'+str(subarea)+'_df'

    dict_of_df[key_name] = VMT_df.loc[VMT_df['subareaName'] == subarea]
    dict_of_df[key_name] = dict_of_df[key_name].transpose()

    # take the data less the first 3 rows
    dict_of_df[key_name] = dict_of_df[key_name][3:]

    # give the columns the speedBin name

    # if the cube outputs only have 13 speed bins, add the extra speed bins
    if numLine_Between+numLine_Within == 260:
        dict_of_df[key_name]["70mph"] = 0
        dict_of_df[key_name]["75mph"] = 0
        dict_of_df[key_name]["80mph"] = 0
        dict_of_df[key_name]["85mph"] = 0
        dict_of_df[key_name]["90mph"] = 0

    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[0]: "5mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[1]: "10mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[2]: "15mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[3]: "20mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[4]: "25mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[5]: "30mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[6]: "35mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[7]: "40mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[8]: "45mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[9]: "50mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[10]: "55mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[11]: "60mph" }, inplace = True)
    dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[12]: "65mph" }, inplace = True)
    if numLine_Between+numLine_Within > 260:
        dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[13]: "70mph" }, inplace = True)
        dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[14]: "75mph" }, inplace = True)
        dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[15]: "80mph" }, inplace = True)
        dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[16]: "85mph" }, inplace = True)
        dict_of_df[key_name].rename(columns={ dict_of_df[key_name].columns[17]: "90mph" }, inplace = True)

    # after transposing, need to add a column with the subaraeName and hour
    dict_of_df[key_name]['subareaName'] = subarea

    dict_of_df[key_name]['Hour'] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24]

    # reorder columns
    dict_of_df[key_name] = dict_of_df[key_name][['subareaName','Hour', '5mph','10mph','15mph','20mph','25mph','30mph','35mph','40mph','45mph','50mph','55mph','60mph','65mph','70mph','75mph','80mph','85mph','90mph']]

    # add it to the vmt data DataFrame
    if subarea == subareaList[0]:
        VMT_reshape_df = dict_of_df[key_name]
    else:
        VMT_reshape_df = VMT_reshape_df.append(dict_of_df[key_name])

    # if the cube outputs only have 13 speed bins (i.e. it is an old run without air basins info)
    # repeat the data for Solano and Sonoma since they straddle two air basins
    if numLine_Between+numLine_Within == 260:
        if subarea == "Solano":
            dict_of_df[key_name]['subareaName'] = "Solano repeated"
            VMT_reshape_df = VMT_reshape_df.append(dict_of_df[key_name])
        if subarea == "Sonoma":
            dict_of_df[key_name]['subareaName'] = "Sonoma repeated"
            VMT_reshape_df = VMT_reshape_df.append(dict_of_df[key_name])

# if the cube outputs only have 13 speed bins (i.e. it is an old run without air basins info), add Air Basin number to the dataframe
if numLine_Between+numLine_Within == 260:
    VMT_reshape_df['AirBasinNum'] = 0

    CountyName_conditions = [
        (VMT_reshape_df['subareaName'] == "Alameda"),
        (VMT_reshape_df['subareaName'] == "Contra Costa"),
        (VMT_reshape_df['subareaName'] == "Marin"),
        (VMT_reshape_df['subareaName'] == "Napa"),
        (VMT_reshape_df['subareaName'] == "San Francisco"),
        (VMT_reshape_df['subareaName'] == "San Mateo"),
        (VMT_reshape_df['subareaName'] == "Santa Clara"),
        (VMT_reshape_df['subareaName'] == "Solano"),
        (VMT_reshape_df['subareaName'] == "Solano repeated"),
        (VMT_reshape_df['subareaName'] == "Sonoma"),
        (VMT_reshape_df['subareaName'] == "Sonoma repeated")]
    AirBasinNumber_choices = [39, 40, 41, 42, 43, 44, 45, 46, 33, 47, 22]
    VMT_reshape_df['AirBasinNum'] = np.select(CountyName_conditions, AirBasinNumber_choices, default='null')
    VMT_reshape_df['AirBasinNum'] = VMT_reshape_df['AirBasinNum'].astype(int) # make sure the data type is integer

print "\nFinished reshaping the modelled VMT data"

# -------------------------------------------------------------------
# Calculate the hourly fraction from CreateSpeedBinsBetweenZones and CreateSpeedBinsWithinZones
# -------------------------------------------------------------------
VMT_reshape_df['HourlyTotVMT'] = (VMT_reshape_df['5mph'] + VMT_reshape_df['10mph'] + VMT_reshape_df['15mph'] + VMT_reshape_df['20mph']
                                + VMT_reshape_df['25mph'] + VMT_reshape_df['30mph'] + VMT_reshape_df['35mph'] + VMT_reshape_df['40mph']
                                + VMT_reshape_df['45mph'] + VMT_reshape_df['50mph'] + VMT_reshape_df['55mph'] + VMT_reshape_df['60mph']
                                + VMT_reshape_df['65mph'] + VMT_reshape_df['70mph'] + VMT_reshape_df['75mph'] + VMT_reshape_df['80mph']
                                + VMT_reshape_df['85mph'] + VMT_reshape_df['90mph'])

VMT_reshape_df['HourlyFraction_5mph'] = VMT_reshape_df['5mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_10mph'] = VMT_reshape_df['10mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_15mph'] = VMT_reshape_df['15mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_20mph'] = VMT_reshape_df['20mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_25mph'] = VMT_reshape_df['25mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_30mph'] = VMT_reshape_df['30mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_35mph'] = VMT_reshape_df['35mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_40mph'] = VMT_reshape_df['40mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_45mph'] = VMT_reshape_df['45mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_50mph'] = VMT_reshape_df['50mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_55mph'] = VMT_reshape_df['55mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_60mph'] = VMT_reshape_df['60mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_65mph'] = VMT_reshape_df['65mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_70mph'] = VMT_reshape_df['70mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_75mph'] = VMT_reshape_df['75mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_80mph'] = VMT_reshape_df['80mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_85mph'] = VMT_reshape_df['85mph'] / VMT_reshape_df['HourlyTotVMT']
VMT_reshape_df['HourlyFraction_90mph'] = VMT_reshape_df['90mph'] / VMT_reshape_df['HourlyTotVMT']

VMT_reshape_df.to_csv(output_filename2, header=True, index=False)

# -------------------------------------------------------------------
# Reading from and writing to the EMFAC Custom Activity Template
# Part 1: hourly fractions
# -------------------------------------------------------------------
print "\n\n================================================================"
print "Reading from and writing to the EMFAC Custom Activity Template"
print "================================================================"

print "\nLoading the workbook "+emfac_input_xlsx_fullpath
workbook = load_workbook(filename=emfac_input_xlsx_fullpath)
print "\nWhat are the different tabs in this workbook?"
print workbook.sheetnames

# make "the hourly fraction tab" the active tab
sheet = workbook["Hourly_Fraction_Veh_Tech_Speed"]
print "\nActivated the tab:"
print sheet

# The key data items I want to extract from this tab are the Sub-Area column (column B) and the Hour column (column F)
# as these two columns will from the "index"
# but it seeems easier to just read all the data in this sheet, so this is what've done below.
print "\nReading in the data from <Hourly_Fraction_Veh_Tech_Speed>"
DefaultHourlyFraction = sheet.values

# For some reasons, Column A and C have the same header "Sub-Area" and are identical in contents
# Column C is renamed here temporarily to avoid confusion. It'll be renamed back near the end of the process, to ensure that the format of the custom activity template was replicated exactly.
# edit a specific cell
sheet["C1"] = "Sub-Area2"

# Set the first row as the headers for the DataFrame
cols = next(DefaultHourlyFraction)
DefaultHourlyFraction = list(DefaultHourlyFraction)

DefaultHourlyFraction_df = pd.DataFrame(DefaultHourlyFraction, columns=cols)

DefaultHourlyFraction_df.to_csv(output_filename, header=True, index=False)

print "\nFinish reading <Hourly_Fraction_Veh_Tech_Speed>"

# Some further notes about the "Hourly_Fraction_Veh_Tech_Speed" tab:

# there are 11 Sub-Areas (9 counties but Solano and Sonoma are in 2 air basins)
# the Sub-Areas are ordered as followed when the "custom activity template" was generated by EMFAC:
# Solano (SV)
# Alameda (SF)
# Contra Costa (SF)
# Marin (SF)
# Napa (SF)
# San Francisco (SF)
# San Mateo (SF)
# Santa Clara (SF)
# Solano (SF)
# Sonoma (SF)
# Sonoma (NC)

# The "Hourly_Fraction_Veh_Tech_Speed" is essentially a table of subarea (11) x hour (24) x speed bins (18).
# The rows are subarea (11) x hour (24) x Veh_Tech (51)
# The columns are the 18 speed bins, plus index columns (Sub-Area, GAI, Sub-Area, Cal_Year, Veh_Tech, Hour)

# We will merge on Sub-area (identified by GAI) and Hour
# i.e. the fractions are the assumed to be same for each subarea-hour combination
# The Travel Model doesn't have the concept of Veh_Tech, so:
#  - for 9 technologies, use the travel model's hourly fraction
#    LDA - Dsl
#    LDA - Gas
#    LDT1 - Dsl
#    LDT1 - Gas
#    LDT2 - Dsl
#    LDT2 - Gas
#    MCY - Gas
#    MDV - Dsl
#    MDV - Gas
#  - for the rest, keep the default hourly fraction
# Flavia to read more about why these 9 technologies

# add a "use TM hourly fractions" column to the default hourly fractons dataframe
DefaultHourlyFraction_df['useTMhourlyfractions'] = 99
DefaultHourlyFraction_df['useTMhourlyfractions'] = np.where(
    (DefaultHourlyFraction_df['Veh_Tech']=='LDA - Dsl') |
    (DefaultHourlyFraction_df['Veh_Tech']=='LDA - Gas') |
    (DefaultHourlyFraction_df['Veh_Tech']=='LDT1 - Dsl') |
    (DefaultHourlyFraction_df['Veh_Tech']=='LDT1 - Gas') |
    (DefaultHourlyFraction_df['Veh_Tech']=='LDT2 - Dsl') |
    (DefaultHourlyFraction_df['Veh_Tech']=='LDT2 - Gas') |
    (DefaultHourlyFraction_df['Veh_Tech']=='MCY - Gas') |
    (DefaultHourlyFraction_df['Veh_Tech']=='MDV - Dsl') |
    (DefaultHourlyFraction_df['Veh_Tech']=='MDV - Gas') ,
    1, 0)


# merge DataFrames
TM_HourlyFraction_df = pd.merge(DefaultHourlyFraction_df, VMT_reshape_df, left_on=['GAI','Hour'], right_on=['AirBasinNum','Hour'], how='left')

TM_HourlyFraction_df['5mph']  = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_5mph'], TM_HourlyFraction_df['5mph_x'])
TM_HourlyFraction_df['10mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_10mph'], TM_HourlyFraction_df['10mph_x'])
TM_HourlyFraction_df['15mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_15mph'], TM_HourlyFraction_df['15mph_x'])
TM_HourlyFraction_df['20mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_20mph'], TM_HourlyFraction_df['20mph_x'])
TM_HourlyFraction_df['25mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_25mph'], TM_HourlyFraction_df['25mph_x'])
TM_HourlyFraction_df['30mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_30mph'], TM_HourlyFraction_df['30mph_x'])
TM_HourlyFraction_df['35mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_35mph'], TM_HourlyFraction_df['35mph_x'])
TM_HourlyFraction_df['40mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_40mph'], TM_HourlyFraction_df['40mph_x'])
TM_HourlyFraction_df['45mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_45mph'], TM_HourlyFraction_df['45mph_x'])
TM_HourlyFraction_df['50mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_50mph'], TM_HourlyFraction_df['50mph_x'])
TM_HourlyFraction_df['55mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_55mph'], TM_HourlyFraction_df['55mph_x'])
TM_HourlyFraction_df['60mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_60mph'], TM_HourlyFraction_df['60mph_x'])
TM_HourlyFraction_df['65mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_65mph'], TM_HourlyFraction_df['65mph_x'])
TM_HourlyFraction_df['70mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_70mph'], TM_HourlyFraction_df['70mph_x'])
TM_HourlyFraction_df['75mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_75mph'], TM_HourlyFraction_df['75mph_x'])
TM_HourlyFraction_df['80mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_80mph'], TM_HourlyFraction_df['80mph_x'])
TM_HourlyFraction_df['85mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_85mph'], TM_HourlyFraction_df['85mph_x'])
TM_HourlyFraction_df['90mph'] = np.where(TM_HourlyFraction_df['useTMhourlyfractions'] == 1, TM_HourlyFraction_df['HourlyFraction_90mph'], TM_HourlyFraction_df['90mph_x'])

TM_HourlyFraction_df.to_csv(output_debug1, header=True, index=False)


# keep only the relevant columns
TM_HourlyFraction_df = TM_HourlyFraction_df[['Sub-Area', 'GAI', 'Sub-Area2', 'Cal_Year', 'Veh_Tech', 'Hour',
                                             '5mph', '10mph', '15mph', '20mph', '25mph', '30mph', '35mph', '40mph', '45mph',
                                             '50mph', '55mph', '60mph', '65mph', '70mph', '75mph', '80mph', '85mph', '90mph']]


# round the results to two decimal places
# commented out the rounding step because it seems to cause problems - emfac won't run when a row doesn't add up to 1
#TM_HourlyFraction_df['5mph'] = TM_HourlyFraction_df['5mph'].astype(float).round(4)
#TM_HourlyFraction_df['10mph'] = TM_HourlyFraction_df['10mph'].astype(float).round(4)
#TM_HourlyFraction_df['15mph'] = TM_HourlyFraction_df['15mph'].astype(float).round(4)
#TM_HourlyFraction_df['20mph'] = TM_HourlyFraction_df['20mph'].astype(float).round(4)
#TM_HourlyFraction_df['25mph'] = TM_HourlyFraction_df['25mph'].astype(float).round(4)
#TM_HourlyFraction_df['30mph'] = TM_HourlyFraction_df['30mph'].astype(float).round(4)
#TM_HourlyFraction_df['35mph'] = TM_HourlyFraction_df['35mph'].astype(float).round(4)
#TM_HourlyFraction_df['40mph'] = TM_HourlyFraction_df['40mph'].astype(float).round(4)
#TM_HourlyFraction_df['45mph'] = TM_HourlyFraction_df['45mph'].astype(float).round(4)
#TM_HourlyFraction_df['50mph'] = TM_HourlyFraction_df['50mph'].astype(float).round(4)
#TM_HourlyFraction_df['55mph'] = TM_HourlyFraction_df['55mph'].astype(float).round(4)
#TM_HourlyFraction_df['60mph'] = TM_HourlyFraction_df['60mph'].astype(float).round(4)
#TM_HourlyFraction_df['65mph'] = TM_HourlyFraction_df['65mph'].astype(float).round(4)
#TM_HourlyFraction_df['70mph'] = TM_HourlyFraction_df['70mph'].astype(float).round(4)
#TM_HourlyFraction_df['75mph'] = TM_HourlyFraction_df['75mph'].astype(float).round(4)
#TM_HourlyFraction_df['80mph'] = TM_HourlyFraction_df['80mph'].astype(float).round(4)
#TM_HourlyFraction_df['85mph'] = TM_HourlyFraction_df['85mph'].astype(float).round(4)
#TM_HourlyFraction_df['90mph'] = TM_HourlyFraction_df['90mph'].astype(float).round(4)

TM_HourlyFraction_df.to_csv(output_filename3, header=True, index=False)

#   CHECK WHY SOME OF THE FRACTIONS DO NOT ADD UP TO BE 100% - IS IS AFTER ROUNDING???

# write the data to excel
# rename the existing "Hourly_Fraction_Veh_Tech_Speed" and add a new sheet
print "\nStart writing Travel Model results to <Hourly_Fraction_Veh_Tech_Speed>"

sheet.title = 'default hourly fractions'

workbook.create_sheet('Hourly_Fraction_Veh_Tech_Speed')
# activate the new sheet
sheet = workbook['Hourly_Fraction_Veh_Tech_Speed']

for row in dataframe_to_rows(TM_HourlyFraction_df, index=False, header=True):
    sheet.append(row)

workbook.save(output_excel_template)

print "\nFinished writing to <Hourly_Fraction_Veh_Tech_Speed>"

# clean up
# if the cube outputs only have 13 speed bins (i.e. it is an old run without air basins info)
# drop the repeated rows for Solano and Sonoma in the vmt dataframe
if numLine_Between+numLine_Within == 260:
    VMT_reshape_df = VMT_reshape_df.loc[VMT_reshape_df['subareaName'] != "Solano repeated"]
    VMT_reshape_df = VMT_reshape_df.loc[VMT_reshape_df['subareaName'] != "Sonoma repeated"]

# -------------------------------------------------------------------
# Reading from and writing to the EMFAC Custom Activity Template
# Part 2: VMT
# -------------------------------------------------------------------
workbook2 = load_workbook(filename=output_excel_template)

# make "the hourly fraction tab" the active tab
sheet2 = workbook2["Daily_VMT_By_Veh_Tech"]
print "\nActivated the tab:"
print sheet2

print "\nReading in the data on the tab <Daily_VMT_By_Veh_Tech>"
DefaultVMT = sheet2.values

# For some reasons, Column A and C have the same header "Sub-Area" and are identical in contents
# Column C is renamed here temporarily to avoid confusion. It'll be renamed back near the end of the process, to ensure that the format of the custom activity template was replicated exactly.
# edit a specific cell
sheet2["C1"] = "Sub-Area2"

# Set the first row as the headers for the DataFrame
cols = next(DefaultVMT)
DefaultVMT = list(DefaultVMT)
DefaultVMT_df = pd.DataFrame(DefaultVMT, columns=cols)

DefaultVMT_df.to_csv(output_filename5, header=True, index=False)

# rename the sheet
sheet2.title = 'CARB default VMT'

# Calculate percent VMT by vehicle technology
# first calculate CARB default VMT by subarea
DefaultVMTbyGAI_df = DefaultVMT_df.groupby('GAI', as_index=False).sum()
# raneme the 'New Total VMT' column to avoid confusion
DefaultVMTbyGAI_df.rename(columns={"New Total VMT": "CARB VMT by subarea"}, inplace=True)
DefaultVMTbyGAI_df.to_csv(output_filename6, header=True, index=False)

# merge the CARB default VMT by subarea data back to the main DataFrame
DefaultVMT_df = pd.merge(DefaultVMT_df, DefaultVMTbyGAI_df, left_on=['GAI'], right_on=['GAI'], how='left')

# percentVMT_by_VehTech
DefaultVMT_df['percentVMT']= DefaultVMT_df['New Total VMT'] / DefaultVMT_df['CARB VMT by subarea']

# calculate TM VMT by subarea
ModelledVMTbyGAI_df = VMT_reshape_df[['subareaName', 'AirBasinNum', 'HourlyTotVMT']]
ModelledVMTbyGAI_df = ModelledVMTbyGAI_df.groupby(['AirBasinNum', 'subareaName'], as_index=False).sum()
ModelledVMTbyGAI_df.to_csv(output_filename7, header=True, index=False)

# if the modelled VMT has only 9 rows, that means the VMT in Solano and Sonoma need to be split into their air basins. See air basin numbers below.
# Solano (SF)	GAI=46
# Solano (SV)	GAI=33
# Sonoma (NC)	GAI=22
# Sonoma (SF)	GAI=47
numLine_ModelledVMTbyGAI = len(ModelledVMTbyGAI_df.index)
if numLine_ModelledVMTbyGAI == 9:

    # Solano default VMT
    DefaultVMT_Solano_df = DefaultVMTbyGAI_df.loc[(DefaultVMTbyGAI_df['GAI'] == 46) | (DefaultVMTbyGAI_df['GAI']== 33)]
    DefaultVMT_Solano_scalar = DefaultVMT_Solano_df['CARB VMT by subarea'].sum()
    print "Default VMT Solano " + str(DefaultVMT_Solano_scalar)
    DefaultVMT_Solano_df.to_csv(output_filename10, header=True, index=False)

    # Sonoma default VMT
    DefaultVMT_Sonoma_df = DefaultVMTbyGAI_df.loc[(DefaultVMTbyGAI_df['GAI'] == 47) | (DefaultVMTbyGAI_df['GAI']== 22)]
    DefaultVMT_Sonoma_scalar = DefaultVMT_Sonoma_df['CARB VMT by subarea'].sum()
    print "Default VMT Sonoma " + str(DefaultVMT_Sonoma_scalar)
    DefaultVMT_Sonoma_df.to_csv(output_filename11, header=True, index=False)

    # Solano modelled VMT
    ModelledVMT_Solano_df = ModelledVMTbyGAI_df.loc[ModelledVMTbyGAI_df['subareaName']=="Solano"]
    ModelledVMT_Solano_scalar = ModelledVMT_Solano_df['HourlyTotVMT'].sum()
    print "Modelled VMT Solano " + str(ModelledVMT_Solano_scalar)

    # Sonoma modelled VMT
    ModelledVMT_Sonoma_df = ModelledVMTbyGAI_df.loc[ModelledVMTbyGAI_df['subareaName']=="Sonoma"]
    ModelledVMT_Sonoma_scalar = ModelledVMT_Sonoma_df['HourlyTotVMT'].sum()
    print "Modelled VMT Sonoma " + str(ModelledVMT_Sonoma_scalar)

    # merge modelled VMT by subarea with default vmt by subarea
    SplitVMT_df = pd.merge(DefaultVMTbyGAI_df, ModelledVMTbyGAI_df, left_on=['GAI'], right_on=['AirBasinNum'], how='left')

    # Sonoma (NC)	GAI=22
    Sonoma_NC = SplitVMT_df.loc[SplitVMT_df.GAI == 22, 'CARB VMT by subarea']/DefaultVMT_Sonoma_scalar*ModelledVMT_Sonoma_scalar
    # Solano (SV)	GAI=33
    Solano_SV = SplitVMT_df.loc[SplitVMT_df.GAI == 33, 'CARB VMT by subarea']/DefaultVMT_Solano_scalar*ModelledVMT_Solano_scalar
    # Solano (SF)	GAI=46
    Solano_SF = SplitVMT_df.loc[SplitVMT_df.GAI == 46, 'CARB VMT by subarea']/DefaultVMT_Solano_scalar*ModelledVMT_Solano_scalar
    # Sonoma (SF)	GAI=47
    Sonoma_SF = SplitVMT_df.loc[SplitVMT_df.GAI == 47, 'CARB VMT by subarea']/DefaultVMT_Sonoma_scalar*ModelledVMT_Sonoma_scalar

    GAI_as_conditions = [
        (SplitVMT_df['GAI'] == 22),
        (SplitVMT_df['GAI'] == 33),
        (SplitVMT_df['GAI'] == 39),
        (SplitVMT_df['GAI'] == 40),
        (SplitVMT_df['GAI'] == 41),
        (SplitVMT_df['GAI'] == 42),
        (SplitVMT_df['GAI'] == 43),
        (SplitVMT_df['GAI'] == 44),
        (SplitVMT_df['GAI'] == 45),
        (SplitVMT_df['GAI'] == 46),
        (SplitVMT_df['GAI'] == 47)]
    VMT_choices = [
                   Sonoma_NC,
                   Solano_SV,
                   SplitVMT_df.loc[SplitVMT_df.GAI == 39, 'HourlyTotVMT'],
                   SplitVMT_df.loc[SplitVMT_df.GAI == 40, 'HourlyTotVMT'],
                   SplitVMT_df.loc[SplitVMT_df.GAI == 41, 'HourlyTotVMT'],
                   SplitVMT_df.loc[SplitVMT_df.GAI == 42, 'HourlyTotVMT'],
                   SplitVMT_df.loc[SplitVMT_df.GAI == 43, 'HourlyTotVMT'],
                   SplitVMT_df.loc[SplitVMT_df.GAI == 44, 'HourlyTotVMT'],
                   SplitVMT_df.loc[SplitVMT_df.GAI == 45, 'HourlyTotVMT'],
                   Solano_SF,
                   Sonoma_SF]

    SplitVMT_df['ModelledVMT_all11'] = np.select(GAI_as_conditions, VMT_choices, default='null')

    newSubareaName_choices = [
                   "Sonoma_NC",
                   "Solano_SV",
                   "Alameda_SF",
                   "Contra_Costa_SF",
                   "Marin_SF",
                   "Napa_SF",
                   "San_Francisco_SF",
                   "San_Mateo_SF",
                   "Santa_Clara_SF",
                   "Solano_SF",
                   "Sonoma_SF"]
    SplitVMT_df['newSubareaName'] = np.select(GAI_as_conditions, newSubareaName_choices, default='null')
    SplitVMT_df.to_csv(output_filename9, header=True, index=False)

    # modify the modelled VMT data frame
    # drop the old HourlyTotVMT column
    ModelledVMTbyGAI_df = SplitVMT_df[['GAI','newSubareaName','ModelledVMT_all11']]
    ModelledVMTbyGAI_df.to_csv(output_checkModelled, header=True, index=False)
    # rename the columns
    ModelledVMTbyGAI_df.rename(columns={"GAI": "AirBasinNum", "newSubareaName": "subareaName", "ModelledVMT_all11": "HourlyTotVMT"}, inplace=True)

# merge in TM VMT by subarea
DefaultVMT_df = pd.merge(DefaultVMT_df, ModelledVMTbyGAI_df, left_on=['GAI'], right_on=['AirBasinNum'], how='left')
DefaultVMT_df.to_csv(output_DefaultModelled, header=True, index=False)

print "\nSomehow DefaultVMT_df['HourlyTotVMT'].dtype is an " + str(DefaultVMT_df['HourlyTotVMT'].dtype)
print "When converted to float, it generates a warning. I'm ignoring this because the column in question is meant to be float."

# DefaultVMT_df['modelled_VMT'] = DefaultVMT_df['percentVMT'] * DefaultVMT_df['HourlyTotVMT'] # this didn't work
DefaultVMT_df['modelled_VMT'] = DefaultVMT_df['percentVMT'] * DefaultVMT_df['HourlyTotVMT'].astype(float)
# DefaultVMT_df['modelled_VMT'] = DefaultVMT_df['percentVMT'].astype(float) * DefaultVMT_df['HourlyTotVMT'] #this doesn't work

DefaultVMT_df.to_csv(output_filename8, header=True, index=False)

# keep the relevant columns for writing to the excel file
TM_VMT_By_Veh_Tech_df =  DefaultVMT_df[['Sub-Area', 'GAI', 'Sub-Area2', 'Cal_Year_x', 'Veh_Tech', 'modelled_VMT']]

# writing to excel
print "\nStart writing to <Daily_VMT_By_Veh_Tech>"
workbook2.create_sheet('Daily_VMT_By_Veh_Tech')
# activate the new sheet
sheet3 = workbook2['Daily_VMT_By_Veh_Tech']

for row in dataframe_to_rows(TM_VMT_By_Veh_Tech_df, index=False, header=True):
    sheet3.append(row)

# rename the columns by changing the cell on the sheet
sheet3["C1"] = "Sub-Area"
sheet3["D1"] = "Cal_Year"
sheet3["F1"] = "New Total VMT"

workbook2.save(output_excel_template)

print "\nFinished writing to <Daily_VMT_By_Veh_Tech>"


# -------------------------------------------------------------------
# Reading from and writing to the EMFAC Custom Activity Template
# Part 3: add a readme
# -------------------------------------------------------------------
# the emfac default values were replaced by Travel Model outputs via emfac_prep.py
# other info e.g. model run name
