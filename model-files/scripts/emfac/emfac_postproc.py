USAGE = """

  This script is to be run after EMFAC is completed (any number of times).
  Run from the model run directory on M.

  This script will find all the EMFAC output files and extract relevant information, writing a summary file
  with all information included to OUTPUT\\emfac\\emfac_summary.csv
   
  Columns are:
  - analysis_type
  - source
  - emfac_version
  - season
  - sb375 = bool
  - planning = bool
  - run_timestamp
  - Veh_Tech
  - VMT
  - Trips
  - CO2_RUNEX
  - CO2_IDLEX
  - CO2_STREX
  - CO2_TOTEX

"""

import os, pathlib, re, sys
import openpyxl
import pandas as pd

# -------------------------------------------------------------------
# Input/output file names and locations
# -------------------------------------------------------------------
CWD              = pathlib.Path.cwd()
MODEL_RUN_ID     = CWD.name  # run this from M_DIR or model run dir on modeling machine
CWD_OUTPUT_EMFAC = CWD / "OUTPUT/emfac"
OUTPUT_FILE      = CWD_OUTPUT_EMFAC / "emfac_summary.csv"
# columns to extract
EXTRACT_COLUMNS  = ['Veh_Tech', 'VMT', 'Trips', 'CO2_RUNEX', 'CO2_IDLEX', 'CO2_STREX', 'CO2_TOTEX']

# emfac output file ends with timestamp: _YYYYMMDDHHMMSS.xlsx  e.g. _20240410131422.xlsx
# by convention, it's E[emfacversion]_[model_run_id]_[season][_sb375?]_[YYYYMMDDHHMMSS].xlsx
EMFAC_OUTPUT_STR = f"E(?P<emfac_version>2014|2017|2021)_{MODEL_RUN_ID}_(?P<season>annual|summer|winter)(?P<planning>_planning)?(?P<sb375>_sb375)?_(?P<run_timestamp>\d{{14}}).xlsx"
EMFAC_OUTPUT_RE  = re.compile(EMFAC_OUTPUT_STR)

if __name__ == '__main__':

    print(f"MODEL_RUN_ID     = {MODEL_RUN_ID}")
    print(f"CWD_OUTPUT_EMFAC = {CWD_OUTPUT_EMFAC}")
    print(f"EMFAC_OUTPUT_STR = {EMFAC_OUTPUT_STR}")
    print("")

    all_emfac_run_data = [] # list of dicts
    emfac_xlsx_files = sorted(CWD_OUTPUT_EMFAC.glob("**/*.xlsx"))
    for emfac_xlsx_file in emfac_xlsx_files:
        
        relative_emfac_filepath = emfac_xlsx_file.relative_to(CWD_OUTPUT_EMFAC)
        print(f"Checking {relative_emfac_filepath}")

        match = EMFAC_OUTPUT_RE.match(emfac_xlsx_file.name)
        if match == None: 
            print("  -> not an EMFAC output file")
            continue

        # fetch metadata
        emfac_run_data = {
            'analysis_type':    'unknown',
            'source':           str(emfac_xlsx_file),
        }
        if relative_emfac_filepath.parts[0] in ['SB375','EIR','AQConformity']:
            emfac_run_data['analysis_type'] = relative_emfac_filepath.parts[0]
        emfac_run_data['emfac_version'] = match.group('emfac_version')
        emfac_run_data['season']        = match.group('season')
        emfac_run_data['sb375' ]        = True if match.group('sb375')=="_sb375" else False
        emfac_run_data['planning' ]     = True if match.group('planning')=="_planning" else False
        emfac_run_data['run_timestamp'] = match.group('run_timestamp')

        # Load workbook
        workbook = openpyxl.load_workbook(filename=emfac_xlsx_file)
        print(f"  workbook.sheetnames:{workbook.sheetnames}")

        if 'Total MTC' in workbook.sheetnames:
            print(f"  Reading 'Total MTC' worksheet")
            total_mtc_df = pd.read_excel(emfac_xlsx_file, sheet_name="Total MTC")

            # remove leading or trailing spaces
            total_mtc_df.Veh_Tech = total_mtc_df.Veh_Tech.str.strip()

            # select row with All Vehicles
            total_mtc_df = total_mtc_df.loc[ total_mtc_df.Veh_Tech == 'All Vehicles', :]
            assert(len(total_mtc_df)==1)
            all_veh_series = total_mtc_df.squeeze()
            # print(all_veh_series)
            for extract_column in EXTRACT_COLUMNS:
                emfac_run_data[extract_column] = all_veh_series[extract_column]

            all_emfac_run_data.append(emfac_run_data)
            continue
    
    if len(all_emfac_run_data) == 0:
        print("Did not find any EMFAC output workbooks to process")
        sys.exit(0)

    # write it
    all_emfac_run_data_df = pd.DataFrame(all_emfac_run_data)
    all_emfac_run_data_df.to_csv(OUTPUT_FILE, index=False)
    print(f"Wrote {len(all_emfac_run_data_df)} rows to {OUTPUT_FILE}")
