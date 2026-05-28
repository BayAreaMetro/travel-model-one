"""Shared infrastructure for CT-RAMP calibration submodel scripts.

This module provides two reusable building blocks:

:class:`CalibrationConfig`
    Loads a YAML configuration file, applies optional environment-variable
    overrides (e.g. ``TARGET_DIR``, ``CALIB_ITER``), performs ``{param}``
    string substitution across all values, and exposes typed accessors for
    individual settings, file paths, and the county lookup table.

:class:`CalibrationBase`
    Abstract base class that every submodel script subclasses.  It handles
    logging setup, output-directory creation, Excel workbook lifecycle, and
    orchestrates the ``process_data → validate_outputs → generate_outputs``
    pipeline via :meth:`CalibrationBase.run`.  Concrete subclasses implement
    the three abstract methods.

Two module-level helper functions are also provided:

:func:`add_county_info`
    Join TAZ-level county IDs and names onto any DataFrame via a TAZ column.

:func:`create_histogram_tlfd`
    Build a 1-mile-bin trip-length frequency distribution from a distance
    series, supporting both weighted (survey) and sample-scaled (model) inputs.

Typical usage::

    class MySubmodelCalibration(CalibrationBase):
        def process_data(self): ...
        def validate_outputs(self, results): ...
        def generate_outputs(self, results): ...

    calibration = MySubmodelCalibration(config_file="calibration_config_BATS.yaml")
    calibration.run()
"""
import pandas as pd
import numpy as np
import os
import logging
import sys
import yaml
import argparse
from pathlib import Path
try:
    import xlwings as xw
    USE_XLWINGS = True
    xw.App().visible = False
except ImportError:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    USE_XLWINGS = False
from abc import ABC, abstractmethod
from typing import Dict, Any, Optional, List

from calibration_data_models import CTRAMPCounty



class CalibrationConfig:
    """Loads, validates, and exposes calibration configuration from a YAML file.

    Configuration values support ``{param}`` substitution using keys defined
    in the ``general`` section.  Selected ``general`` keys can also be
    overridden at runtime by environment variables (e.g. set by a batch
    launcher), which is useful when running many calibration iterations from a
    single wrapper script.

    Attributes:
        config (dict): Fully substituted configuration dictionary.
        raw_config (dict): Raw parsed YAML before substitution (used as the
            basis for environment-variable overrides).
    """

    def __init__(self, config_file: str = None, submodel: str = None):
        """Load and prepare the configuration file.

        Args:
            config_file: Path to the YAML configuration file.  Defaults to
                ``calibration_config.yaml`` in the same directory as this
                module.
            submodel: Submodel identifier string (e.g. ``"01"``).  Currently
                accepted for API symmetry but not used directly here.

        Raises:
            FileNotFoundError: If the resolved config file path does not exist.
        """
        self.config = {}
        self.raw_config = {}
        
        # Load config file
        config_file = config_file or str(Path(__file__).parent / 'calibration_config.yaml')
        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                self.raw_config = yaml.safe_load(f)
                self._apply_env_overrides()
                self.config = self._substitute_parameters(self.raw_config)

        else:
            raise FileNotFoundError(f"Configuration file not found: {config_file}")

    def _apply_env_overrides(self):
        """Override selected ``general`` config keys from environment variables.

        Reads the following environment variables and, if set, writes their
        values into the ``general`` section of :attr:`raw_config` before
        parameter substitution runs:

        ==================  ===================
        Environment var     Config key
        ==================  ===================
        ``TARGET_DIR``      ``target_dir``
        ``CALIB_ITER``      ``calib_iter``
        ``ITER``            ``iter``
        ``MODEL_DIR``       ``model_dir``
        ``WORKBOOK_BASE_PATH`` ``workbook_base_path``
        ==================  ===================

        Path-like values (keys ending in ``_dir`` or ``_path``) are
        normalised to POSIX separators so that downstream ``{param}``
        substitution stays platform-consistent.
        """
        general = self.raw_config.setdefault('general', {})
        env_to_key = {
            'TARGET_DIR': 'target_dir',
            'CALIB_ITER': 'calib_iter',
            'ITER': 'iter',
            'MODEL_DIR': 'model_dir',
            'WORKBOOK_BASE_PATH': 'workbook_base_path',
        }

        for env_name, config_key in env_to_key.items():
            env_val = os.getenv(env_name)
            if env_val:
                if config_key.endswith('_dir') or config_key.endswith('_path'):
                    general[config_key] = Path(env_val).as_posix()
                else:
                    general[config_key] = env_val
    
    def _substitute_parameters(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Recursively expand ``{param}`` placeholders throughout the config.

        Placeholder keys are resolved against the ``general`` section of the
        config, so a value like ``"{target_dir}/Output_{calib_iter}"`` becomes
        the fully resolved path.  Dict and list nodes are traversed
        recursively; non-string leaf values are returned unchanged.

        Args:
            config: Raw (or partially processed) configuration dictionary.

        Returns:
            A new dictionary with all resolvable placeholders substituted.
        """
        # Get parameters from general section
        params = config.get('general', {})
        
        def substitute_value(value):
            """Recursively substitute parameters in values."""
            if isinstance(value, dict):
                return {k: substitute_value(v) for k, v in value.items()}
            elif isinstance(value, list):
                return [substitute_value(item) for item in value]
            elif isinstance(value, str) and '{' in value:
                try:
                    return value.format(**params)
                except KeyError as e:
                    self.logger.info(f"Warning: Missing parameter {e} in string: {value}")
                    return value
            else:
                return value
        
        return substitute_value(config)
    
    def get(self, section: str, key: str, fallback: str = None) -> str:
        """Return a substituted string value from the configuration.

        Args:
            section: Top-level YAML section name (e.g. ``"general"``,
                ``"data_sources"``).
            key: Key within that section.
            fallback: Value to return if the key is absent.  Defaults to
                ``None``.

        Returns:
            The substituted string value, or *fallback* if not found.
        """
        value = self.config.get(section, {}).get(key, fallback)
        return value
    
    def get_path(self, section: str, key: str, fallback: str = None, validate: bool = False) -> str:
        """Return a file path from the configuration, optionally checking existence.

        Args:
            section: Top-level YAML section name.
            key: Key within that section.
            fallback: Default value if the key is absent.
            validate: If ``True``, log a warning when the resolved path does
                not exist on disk.

        Returns:
            The resolved path string, or *fallback* if not found.
        """
        path = self.get(section, key, fallback)
        if validate and path and not os.path.exists(path):
            self.logger.warning(f"Warning: Path does not exist: {path}")
        return path
    
    def get_all_paths(self, section: str, validate: bool = False) -> Dict[str, str]:
        """Return every file-path value found in a configuration section.

        Heuristically identifies path-like values by the presence of a
        directory separator or a ``.csv`` / ``.xlsx`` extension.

        Args:
            section: Top-level YAML section name to scan.
            validate: If ``True``, log a warning for each path that does not
                exist on disk.

        Returns:
            Dict mapping config key → resolved path string for all
            path-like entries in the section.
        """
        section_config = self.config.get(section, {})
        paths = {}
        for key, value in section_config.items():
            if isinstance(value, str) and ('/' in value or '\\' in value or value.endswith('.csv') or value.endswith('.xlsx')):
                paths[key] = value
                if validate and not os.path.exists(value):
                    self.logger.warning(f"Warning: {key} path does not exist: {value}")
        return paths
    
    def getfloat(self, section: str, key: str, fallback: float = None) -> float:
        """Return a configuration value cast to ``float``.

        Args:
            section: Top-level YAML section name.
            key: Key within that section.
            fallback: Default value if the key is absent.

        Returns:
            The value as a Python ``float``, or *fallback* if not found.
        """
        value = self.config.get(section, {}).get(key, fallback)
        return float(value) if value is not None else fallback
    
    def getint(self, section: str, key: str, fallback: int = None) -> int:
        """Return a configuration value cast to ``int``.

        Args:
            section: Top-level YAML section name.
            key: Key within that section.
            fallback: Default value if the key is absent.

        Returns:
            The value as a Python ``int``, or *fallback* if not found.
        """
        value = self.config.get(section, {}).get(key, fallback)
        return int(value) if value is not None else fallback
    
    def get_county_lookup(self) -> dict:
        """Return a mapping of Bay Area county integer IDs to county name strings.

        Values are sourced from :class:`~calibration_data_models.CTRAMPCounty`
        so they stay in sync with the canonical codebook.

        Returns:
            Dict mapping county ID (int) → county label (str),
            e.g. ``{1: "San Francisco", 2: "San Mateo", ...}``.
        """
        # Use canonical county labels from shared CTRAMP codebook.
        return {county.id: county.label for county in CTRAMPCounty}
    
    def get_submodel_config(self, submodel: str) -> Optional[Dict[str, str]]:
        """Return the configuration block for a specific submodel, or None.

        Looks for a top-level YAML section named ``calibration_<submodel>``
        (e.g. ``calibration_01``).  Returns ``None`` when the section is
        absent so that callers can decide how to handle a missing submodel
        (e.g. skip gracefully when running a BATS config against a submodel
        that does not use survey data).

        Args:
            submodel: Two-digit submodel identifier string (e.g. ``"01"``).

        Returns:
            Dict of all key/value pairs under ``calibration_<submodel>``,
            or ``None`` if the section does not exist in the config file.
        """
        section_name = f"calibration_{submodel}"
        return self.config.get(section_name)


class CalibrationBase(ABC):
    """Abstract base class for CT-RAMP calibration submodel processors.

    Provides shared infrastructure so that every submodel script only needs to
    implement the three domain-specific methods:

    * :meth:`process_data` — load inputs and compute summary DataFrames.
    * :meth:`validate_outputs` — check schemas before anything is written.
    * :meth:`generate_outputs` — write CSV files and update the Excel workbook.

    :meth:`run` orchestrates these three steps and wraps them with workbook
    setup and save.

    Attributes:
        config (CalibrationConfig): Loaded configuration object.
        submodel (str): Submodel identifier (e.g. ``"01"``).
        submodel_config (dict): Config block for this submodel.
        calib_iter (str): Calibration iteration label (e.g. ``"00"``, ``"01"``).
        iter (int): Model iteration number derived from *calib_iter*
            (``3`` for ``"00"``, ``1`` otherwise).
        sampleshare (float): Fraction of the full model sample
            (``0.5`` for ``"00"``, ``0.2`` otherwise).
        bats_data (bool): ``True`` when operating on survey data rather than
            model output.
        target_dir (str): Root output directory from config.
        output_dir (str): Resolved output directory (created on init).
        county_lookup (dict): Mapping of county ID → county name.
        logger (logging.Logger): Configured logger for this submodel.
    """

    def __init__(self, submodel: str, config_file: str = None):
        """Initialise shared calibration infrastructure.

        Loads config, resolves and creates the output directory, configures
        logging, and populates the county lookup table.

        Args:
            submodel: Two-digit submodel identifier string (e.g. ``"01"``).
            config_file: Path to the YAML configuration file.  ``None``
                defaults to ``calibration_config.yaml`` in the script
                directory.
        """
        self.config = CalibrationConfig(config_file, submodel)
        self.submodel = submodel
        self.submodel_config = self.config.get_submodel_config(submodel)
        if self.submodel_config is None:
            config_file_used = config_file or str(Path(__file__).parent / 'calibration_config.yaml')
            print(f"Submodel {submodel} is not configured in {config_file_used} — nothing to do.")
            sys.exit(0)
        self.excel_app = None
        
        # Set up parameters
        self.calib_iter = self.config.get('general', 'calib_iter')
        if self.calib_iter == "00":
            self.iter = 3
            self.sampleshare = 0.5
        else:
            self.iter = 1
            self.sampleshare = 0.2

        # Check if using BATS data
        self.bats_data = self.submodel_config.get("bats_data", False)

        # Set up paths
        self.target_dir = self.config.get('general', 'target_dir')
        if self.bats_data:
            self.output_dir = f"{self.target_dir}/BATS_Summaries"
        else:
            self.output_dir = f"{self.target_dir}/Output_{self.calib_iter}/calibration"
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        
        # Set up logging
        self._setup_logging()

        # Load shared data resources
        self.county_lookup = self.config.get_county_lookup()
       
        # Print configuration
        self._print_config()
    
    def _setup_logging(self):
        """Configure a named logger with console and file handlers.

        The logger name is the concrete subclass name (e.g.
        ``"WorkSchoolLocationCalibration"``).  Both handlers use UTF-8
        encoding so that Unicode characters such as ``✓`` render correctly
        even on Windows terminals that default to cp1252.

        **Console handler** — streams to stdout with UTF-8 encoding.

        **File handler** — writes to ``<output_dir>/<name>.log``
        (or ``<name>_BATS.log`` for BATS survey runs), where *name* is the
        submodel name with spaces removed (e.g.
        ``UsualWorkAndSchoolLocation_BATS.log``).  BATS runs open the file
        with ``mode='w'`` (overwrite) since they are one-shot reference runs;
        model calibration runs use ``mode='a'`` (append) so successive
        iterations accumulate in a single log.
        """
        
        # Create logger
        self.logger = logging.getLogger(self.__class__.__name__)
        self.logger.setLevel(logging.INFO)

        # Prevent duplicate handlers if called multiple times
        if self.logger.hasHandlers():
            self.logger.handlers.clear()

        # Create formatters
        formatter = logging.Formatter(
            '%(asctime)s %(levelname)s %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # Stop messages from bubbling up to the root logger
        self.logger.propagate = False

        # Console handler (force UTF-8 so Unicode characters like ✓ don't fail on cp1252 terminals)
        console_handler = logging.StreamHandler(stream=open(sys.stdout.fileno(), mode='w', encoding='utf-8', closefd=False))
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        self.logger.addHandler(console_handler)

        # File handler — named after the submodel (spaces removed); BATS runs overwrite ('w')
        # since they are one-shot reference runs, while model iterations append ('a')
        # so successive calibration passes accumulate in a single log.
        submodel_name = self.submodel_config.get('name', self.submodel).replace(' ', '')
        log_prefix = f"{submodel_name}_BATS" if self.bats_data else submodel_name
        log_file = os.path.join(self.output_dir, f'{log_prefix}.log')
        log_mode = 'w' if self.bats_data else 'a'
        file_handler = logging.FileHandler(log_file, mode=log_mode, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)

        self.logger.info(f"Logging to: {log_file}")

    def _print_config(self):
        """Log a human-readable summary of the active configuration.

        Emits submodel name, target directory, iteration parameters, sample
        share, and the Excel workbook template name at INFO level.
        """
        sep = "=" * 80
        self.logger.info(f"\n{sep}\nCONFIGURATION SUMMARY\n{sep}")
        self.logger.info(f"Submodel: {self.submodel} - {self.submodel_config['name']}")
        self.logger.info(f"TARGET_DIR  = {self.target_dir}")
        self.logger.info(f"ITER        = {self.iter}")
        self.logger.info(f"SAMPLESHARE = {self.sampleshare}")
        self.logger.info(f"CALIB_ITER  = {self.calib_iter}")
        self.logger.info(f"EXCEL_WORKBOOK = {self.submodel_config.get('workbook_template')}")

    def setup_workbook(self):
        """Resolve Excel workbook paths and open the blank template.

        Looks up the ``workbook_template`` name from the submodel config and
        constructs three path variants:

        ``workbook_path``
            Final versioned workbook path under the calibration base directory.
        ``workbook_temp``
            Temporary save path (``<name>_temp.xlsx``) used to avoid
            overwriting the current iteration file mid-save.
        ``workbook_iter``
            Iteration-stamped path (``<name>_<calib_iter>.xlsx``).
        ``workbook_blank``
            Source blank template, resolved relative to this module's
            ``workbook_templates/`` subdirectory.

        The blank template is opened via xlwings (or skipped gracefully if
        xlwings is unavailable or the file cannot be found).  Sets
        ``self.calib_workbook`` and ``self.modeldata_sheet`` to ``None`` if
        no template is configured or the open fails.
        """
        sep = "=" * 80
        self.logger.info(f"\n{sep}\nWORKBOOK SETUP\n{sep}")
        workbook_template = self.submodel_config.get('workbook_template')
        
        # Skip if no template specified
        if not workbook_template:
            self.calib_workbook = None
            self.modeldata_sheet = None
            return
        
        submodel_name = self.submodel_config['name']
        
        workbook_base = self.config.get('general', 'workbook_base_path', 
                                       'M:/Development/Travel Model One/Calibration/Version 1.7')
        
        self.workbook_path = f"{workbook_base}/{self.submodel} {submodel_name}/{workbook_template}.xlsx"
        self.workbook_temp = self.workbook_path.replace('.xlsx', '_temp.xlsx')
        self.workbook_iter = self.workbook_path.replace('.xlsx', f'_{self.calib_iter}.xlsx')
        # load workbook_blank from the repo (local copy) instead of from "workbook_base" path
        self.workbook_blank = str(Path(__file__).parent / "workbook_templates" / f"{workbook_template}_blank.xlsx")
        
        workbook_dir = Path(self.workbook_path).parent
        self.logger.info(f"Workbook Output Directory: {workbook_dir}")

        try: 
            self.logger.info(f"Loading Excel workbook from {self.workbook_blank}...")
            if USE_XLWINGS:
                # Use a dedicated Excel instance so we do not attach to another open workbook.
                self.excel_app = xw.App(visible=False, add_book=False)
                self.excel_app.display_alerts = False
                self.excel_app.screen_updating = False
                self.calib_workbook = self.excel_app.books.open(
                    self.workbook_blank,
                    update_links=False,
                    ignore_read_only_recommended=True,
                )
            else:
                self.calib_workbook = load_workbook(self.workbook_blank)
            self.modeldata_sheet = self.calib_workbook.sheets['modeldata']
        except Exception as e:
            self.logger.warning(f"Skipping Excel workbook: {e}")
            self.calib_workbook = None
            self.modeldata_sheet = None
            if self.excel_app:
                try:
                    self.excel_app.quit()
                except Exception:
                    pass
                finally:
                    self.excel_app = None
    

    def write_dataframe_to_sheet(self, df: pd.DataFrame, start_row: int, start_col: int, sheet_name: str ="modeldata",
                                source_row: int = None, source_col: int = None, source_text: str = ""):
        """Write a DataFrame into an open Excel workbook sheet.

        Writes the column headers to *start_row* and the data values
        immediately below.  Optionally writes a source-file annotation to a
        separate cell (useful for traceability in the workbook).

        Row and column indices follow xlwings convention (1-based).

        Args:
            df: DataFrame to write.  The index is not written.
            start_row: 1-based row for the header row.
            start_col: 1-based column for the leftmost data column.
            sheet_name: Name of the target sheet.  Defaults to
                ``"modeldata"``.
            source_row: 1-based row for the source annotation cell.
                Skipped if ``None``.
            source_col: 1-based column for the source annotation cell.
                Skipped if ``None``.
            source_text: Text to write at *(source_row, source_col)*,
                typically the output CSV file path.

        Notes:
            Silently skips (with a warning log) if
            ``self.calib_workbook`` is ``None``.
        """
        if not self.calib_workbook:

            self.logger.warning(f"Warning: Calibration Workbook does not exist")
            return

        try:
            sheet = self.calib_workbook.sheets[sheet_name]
            
            # Write headers
            sheet.range((start_row, start_col)).value = df.columns.tolist()
            
            # Write data without index
            sheet.range((start_row + 1, start_col)).value = df.values

            if source_row and source_col and source_text:
                sheet.range((source_row, source_col)).value = source_text
                
        except Exception as e:
            self.logger.warning(f"Warning: Could not write to Excel sheet: {e}")

    @staticmethod
    def _read_xls_range(worksheet, column: int, start_row: int, end_row: int) -> list:
        """Read a 1-based row/column slice from an xlrd worksheet."""
        values = []
        for row in range(start_row, end_row + 1):
            values.append(worksheet.cell_value(row - 1, column - 1))
        return values

    @staticmethod
    def _read_xlsx_range(worksheet, column: int, start_row: int, end_row: int) -> list:
        """Read a 1-based row/column slice from an openpyxl worksheet."""
        values = []
        for row in range(start_row, end_row + 1):
            values.append(worksheet.cell(row=row, column=column).value)
        return values

    @staticmethod
    def _write_xlsx_range(worksheet, column: int, start_row: int, values: list) -> None:
        """Write values to a 1-based row/column slice in an openpyxl worksheet."""
        for offset, value in enumerate(values):
            worksheet.cell(row=start_row + offset, column=column).value = value
    
    def save_workbook(self):
        """Save the open calibration workbook.

        **BATS mode**: overwrites the blank template (:attr:`workbook_blank`)
        so that survey targets are baked into the source file used by future
        model runs.  No iteration snapshot is written.

        **Model mode**: saves to both :attr:`workbook_temp` (stable
        "current" filename) and :attr:`workbook_iter` (iteration-stamped
        snapshot, e.g. ``_01.xlsx``) so successive calibration passes can be
        compared without overwriting earlier results.

        If the workbook was never opened (e.g. no template configured) the
        method logs a note and returns without error.  Any save error is
        caught and logged as a warning.
        """
        sep = "=" * 80
        self.logger.info(f"\n{sep}\nSAVING WORKBOOK\n{sep}")
        if self.calib_workbook:
            try:
                if self.bats_data:
                    self.logger.info(f"Saving BATS targets to blank template: {self.workbook_blank}")
                    self.calib_workbook.save(self.workbook_blank)
                    self.logger.info(f"Wrote {self.workbook_blank}")
                else:
                    self.logger.info("Saving calibration workbook to:")
                    self.logger.info(f"  Temp Workbook:      {self.workbook_temp}")
                    self.logger.info(f"  Iteration Workbook: {self.workbook_iter}")
                    Path(self.workbook_temp).parent.mkdir(parents=True, exist_ok=True)
                    self.calib_workbook.save(self.workbook_temp)
                    self.calib_workbook.save(self.workbook_iter)
                    self.logger.info(f"Wrote {self.workbook_temp}")
                    self.logger.info(f"Wrote {self.workbook_iter}")
                self.calib_workbook.close()
            except Exception as e:
                self.logger.warning(f"Warning: Could not save Excel workbook: {e}")
                self.calib_workbook.close()
            finally:
                if self.excel_app:
                    try:
                        self.excel_app.quit()
                    except Exception:
                        pass
                    finally:
                        self.excel_app = None
        else:
            self.logger.info("Did not save workbook")
    
    @abstractmethod
    def process_data(self) -> Dict[str, pd.DataFrame]:
        """Load input files and produce summary DataFrames.

        Returns:
            Dict mapping result-name strings to DataFrames.  The exact keys
            are submodel-specific and must be consistent with what
            :meth:`validate_outputs` and :meth:`generate_outputs` expect.
        """
        pass

    @abstractmethod
    def validate_outputs(self, results: Dict[str, pd.DataFrame]):
        """Validate output DataFrames against their Pydantic schemas.

        Called by :meth:`run` after :meth:`process_data` and before
        :meth:`generate_outputs`.  Should raise ``ValidationError`` (or any
        exception) to abort the pipeline before any files are written.

        Args:
            results: The dict returned by :meth:`process_data`.
        """
        pass

    @abstractmethod
    def generate_outputs(self, results: Dict[str, pd.DataFrame]):
        """Write validated results to CSV files and update the Excel workbook.

        Args:
            results: The dict returned by :meth:`process_data` and validated
                by :meth:`validate_outputs`.
        """
        pass

    def populate_uec_values(self, src_workbook: str | Path | None = None,
                            dst_workbook: str | Path | None = None) -> None:
        """Populate destination calibration workbook cells (current iteration constants) from UEC source values.

        Subclasses opt in by defining two class attributes:

        * ``UEC_SOURCE_RANGES``: mapping key -> ``(sheet, column, start_row, end_row)``
          for the source ``.xls`` workbook.
        * ``CALIBRATION_DESTINATION_RANGES``: mapping key ->
          ``(sheet, column, start_row, end_row)`` for the destination ``.xlsx`` workbook.

        If either mapping is missing/empty, population is skipped.
        """
        if self.bats_data:
            self.logger.info("Skipping UEC value population for BATS mode.")
            return

        sep = "=" * 80
        self.logger.info(f"\n{sep}\nSAVING WORKBOOK\n{sep}")
        
        source_ranges = getattr(self, "UEC_SOURCE_RANGES", {}) or {}
        destination_ranges = getattr(self, "CALIBRATION_DESTINATION_RANGES", {}) or {}
        if not source_ranges or not destination_ranges:
            self.logger.info("No UEC range mappings configured, skipping population.")
            return

        src_path = Path(src_workbook) if src_workbook else Path(self.submodel_config.get("uec_src_file", ""))
        dst_path = Path(dst_workbook) if dst_workbook else Path(getattr(self, "workbook_iter", ""))

        # Skip population if uec_src_file is not configured.
        if not src_path.name:
            self.logger.info("UEC source file not configured, skipping population.")
            return

        if not src_path.exists():
            raise FileNotFoundError(f"UEC source workbook not found: {src_path}")
        if not dst_path.exists():
            raise FileNotFoundError(f"Calibration destination workbook not found: {dst_path}")

        try:
            from xlrd import open_workbook
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "UEC population requires both 'xlrd' (for .xls) and 'openpyxl' (for .xlsx)."
            ) from e

        sep = "=" * 80
        self.logger.info(f"\n{sep}\nUEC VALUE POPULATION\n{sep}")
        self.logger.info(f"Source workbook: {src_path}")
        self.logger.info(f"Destination workbook: {dst_path}")

        src_workbook_obj = open_workbook(src_path)
        dst_workbook_obj = load_workbook(dst_path)
        total_written = 0

        for name, (src_sheet_name, src_column, src_start_row, src_end_row) in source_ranges.items():
            if name not in destination_ranges:
                raise KeyError(f"Missing destination mapping for UEC key: {name}")

            dst_sheet_name, dst_column, dst_start_row, dst_end_row = destination_ranges[name]
            src_values = self._read_xls_range(
                src_workbook_obj.sheet_by_name(src_sheet_name),
                src_column,
                src_start_row,
                src_end_row,
            )
            expected_dst_len = dst_end_row - dst_start_row + 1
            if len(src_values) != expected_dst_len:
                raise ValueError(
                    f"Population length mismatch for {name}: "
                    f"source has {len(src_values)} values and destination range has {expected_dst_len} rows"
                )

            self._write_xlsx_range(
                dst_workbook_obj[dst_sheet_name],
                dst_column,
                dst_start_row,
                src_values,
            )
            total_written += len(src_values)

        self.logger.info(f"Populated {total_written} UEC cells into {dst_path}.")
        dst_workbook_obj.save(dst_path)
        self.logger.info(f"Saved populated workbook to {dst_path}")

        # Keep the temp workbook in sync when using the default destination.
        if dst_workbook is None and getattr(self, "workbook_temp", None) and Path(self.workbook_temp) != dst_path:
            dst_workbook_obj.save(self.workbook_temp)
            self.logger.info(f"Saved populated workbook to {self.workbook_temp}")

    def run(self):
        """Run the complete calibration pipeline for this submodel.

        Executes the following steps in order:

        1. :meth:`process_data` — load inputs and compute summaries.
        2. :meth:`validate_outputs` — schema-check every output DataFrame.
        3. :meth:`setup_workbook` — open the Excel template.
        4. :meth:`generate_outputs` — write CSVs and populate the workbook.
        5. :meth:`save_workbook` — save the workbook to disk.
        6. :meth:`populate_uec_values` — populate the workbook with UEC values.

        Any unhandled exception is logged before being re-raised so that the
        error appears in the log file as well as the console.
        """
        try:
            results = self.process_data()
            self.validate_outputs(results)
            self.setup_workbook()
            self.generate_outputs(results)
            self.save_workbook()
            self.populate_uec_values()
        except Exception as e:
            self.logger.info(f"Error during calibration processing: {e}")
            raise


def add_county_info(df: pd.DataFrame, taz_data: pd.DataFrame,
                   county_lookup: dict, taz_col: str = 'TAZ',
                   county_col_name: str = 'COUNTY',
                   county_name_col: str = 'county_name') -> pd.DataFrame:
    """Join county ID and name columns onto a DataFrame via a TAZ key.

    Performs a left merge so that rows with an unrecognised TAZ (e.g. TAZ 0
    meaning "no location") are retained with ``NaN`` county values rather
    than being dropped.

    Args:
        df: DataFrame that contains a TAZ column to join on.
        taz_data: TAZ attribute table with at least ``ZONE`` and ``COUNTY``
            columns (typically read from the model's TAZ CSV).
        county_lookup: Mapping of county integer ID → county name string,
            as returned by :meth:`CalibrationConfig.get_county_lookup`.
        taz_col: Name of the TAZ column in *df*.  Defaults to ``"TAZ"``.
        county_col_name: Desired name for the output county-ID column.
            Defaults to ``"COUNTY"``.
        county_name_col: Desired name for the output county-name column.
            Defaults to ``"county_name"``.

    Returns:
        *df* with two additional columns: *county_col_name* (integer county
        ID) and *county_name_col* (string county name).

    Example::

        wsloc = add_county_info(wsloc, taz_data, county_lookup,
                                taz_col='HomeTAZ',
                                county_col_name='HomeCOUNTY',
                                county_name_col='HomeCounty_name')
    """
    # Merge with TAZ county data
    taz_county = taz_data[['ZONE', 'COUNTY']].rename(columns={'ZONE': taz_col})
    df = df.merge(taz_county, on=taz_col, how='left')
    
    # Add county name
    df[county_name_col] = df['COUNTY'].map(county_lookup)
    
    # Rename COUNTY column if specified
    if county_col_name != 'COUNTY':
        df = df.rename(columns={'COUNTY': county_col_name})
    
    return df


def create_histogram_tlfd(data: pd.Series, bins: range = None, sampleshare: float = 1.0,
                         weights: pd.Series = None) -> pd.DataFrame:
    """Build a trip-length frequency distribution in 1-mile bins.

    Wraps ``numpy.histogram`` to produce a tidy DataFrame suitable for
    writing directly to a calibration CSV or Excel sheet.

    Two weighting modes are supported:

    * **Survey mode** (``weights`` provided): uses ``np.histogram`` weighted
      counts — each observation contributes its person weight to the bin.
    * **Model mode** (``weights=None``): uses raw counts divided by
      *sampleshare* to scale up to a full-population estimate.

    Args:
        data: Series of trip distances in miles (one value per person/trip).
        bins: Bin edges passed to ``np.histogram``.  The number of output
            rows equals ``len(bins) - 1``.  Defaults to ``range(151)``
            (bins 1–150 miles) for model runs; pass ``range(52)`` for
            BATS survey runs (bins 1–51 miles).
        sampleshare: Fraction of the full model population represented by
            the input data (e.g. ``0.5`` for a 50 % sample).  Only used
            when *weights* is ``None``.  Defaults to ``1.0``.
        weights: Per-observation person weights (BATS survey mode).  When
            provided, *sampleshare* is ignored.

    Returns:
        DataFrame with columns:

        ``distbin``
            Upper edge of each bin (i.e. the integer mile value, 1-based).
        ``count``
            Weighted or scaled count of trips in that bin.
    """
    if bins is None:
        bins = range(151)
    
    if weights is not None:
        # Use weighted histogram
        hist, bin_edges = np.histogram(data, bins=bins, weights=weights)
        return pd.DataFrame({
            'distbin': list(bins)[1:],
            'count': hist
        })
    else:
        # Use sampleshare scaling
        hist, bin_edges = np.histogram(data, bins=bins)
        return pd.DataFrame({
            'distbin': list(bins)[1:],
            'count': hist / sampleshare
        })

# def attach_distance_skim() -> pd.DataFrame:
    