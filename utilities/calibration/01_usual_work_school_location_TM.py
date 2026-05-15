"""Usual work and school location calibration submodel (submodel 01).

This script calibrates the usual work and school location (mandatory location)
submodel of CT-RAMP by comparing modeled destination TAZ distributions against
observed BATS survey data.  It produces three outputs:

- A home-to-work county-to-county flow matrix.
- Trip Length Frequency Distributions (TLFDs) by county for work, university,
  and K–12 school trips, binned in 1-mile increments.
- Average trip lengths by county and trip type.

When ``bats_data`` is ``true`` in the submodel config the script reads person-
weighted BATS 2023 survey records and writes results to the ``BATS_Summaries``
subdirectory (for use as calibration targets).  When ``bats_data`` is ``false``
it reads CT-RAMP model output and writes results to the iteration output
directory for comparison against those targets.

Usage::

    python 01_usual_work_school_location_TM.py [--config PATH]

Arguments:
    --config    Optional path to the YAML configuration file.  Defaults to
                ``calibration_config.yaml`` in the same directory as this script.
"""
import argparse
import math
import pandas as pd
import numpy as np
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from xlrd import open_workbook

# Import the calibration framework
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from calibration_framework import CalibrationBase, create_histogram_tlfd, add_county_info
from calibration_data_models import (
    CountyTripSummary,
    TripLengthFrequency,
    AverageTripLength,
    validate_dataframe
)

class WorkSchoolLocationCalibration(CalibrationBase):
    """Calibration processor for the usual work and school location submodel.

    Inherits shared infrastructure (config loading, logging, Excel workbook
    helpers) from :class:`CalibrationBase` and implements the three-stage
    pipeline defined by that base class:

    1. :meth:`process_data`   – loads inputs, attaches distances, and builds
       summary DataFrames.
    2. :meth:`validate_outputs` – checks that every output DataFrame conforms
       to its Pydantic schema before anything is written to disk.
    3. :meth:`generate_outputs` – writes CSV files and (optionally) updates
       the calibration Excel workbook.

    The behaviour of the script differs depending on the ``bats_data`` flag in
    the submodel configuration:

    * ``bats_data: true``  – survey-weighted BATS 2023 run; outputs land in
      ``<target_dir>/BATS_Summaries/`` and serve as calibration targets.
    * ``bats_data: false`` – model run; outputs land in
      ``<target_dir>/Output_<calib_iter>/calibration/`` for comparison.
    """

    # sheet, column, startRow, endRow
    # used for validating UEC values in calibration workbook against model input
    UEC_SOURCE_RANGES = {
        "work": ("Work", 7, 22, 26),
        "work_county": ("Work", 7, 38, 47),
        "university": ("University", 7, 12, 16),
        "highschool": ("HighSchool", 7, 12, 16),
        "gradeschool": ("GradeSchool", 7, 12, 16),
    }

    CALIBRATION_DESTINATION_RANGES = {
        "work": ("calibration", 3, 4, 8),
        "work_county": ("calibration", 9, 4, 13),
        "university": ("calibration", 15, 4, 8),
        "highschool": ("calibration", 31, 4, 8),
        "gradeschool": ("calibration", 32, 4, 8),
    }

    def __init__(self, config_file: str = None):
        """Initialise the calibration processor.

        Args:
            config_file: Path to the YAML configuration file.  ``None`` causes
                :class:`CalibrationBase` to look for ``calibration_config.yaml``
                in the same directory as this script.
        """
        super().__init__("01", config_file)
        self.bats_data = self.submodel_config.get("bats_data", False)

    @staticmethod
    def _read_xls_range(worksheet, column: int, start_row: int, end_row: int) -> list:
        """Read a 1-based row/column slice from an xlrd worksheet."""
        values = []
        for row in range(start_row, end_row + 1):
            values.append(worksheet.cell_value(row - 1, column - 1))
        return values

    @staticmethod
    def _read_xlsx_range(worksheet, column: int, start_row: int, end_row: int) -> list:
        """Read a 1-based row/column slice from an openpyxl worksheet."""
        values = []
        for row in range(start_row, end_row + 1):
            values.append(worksheet.cell(row=row, column=column).value)
        return values

    @staticmethod
    def _values_match(src_value, dst_value) -> bool:
        """Compare Excel values while tolerating numeric type differences."""
        if isinstance(src_value, str) or isinstance(dst_value, str):
            return str(src_value).strip() == str(dst_value).strip()

        if isinstance(src_value, bool) or isinstance(dst_value, bool):
            return src_value == dst_value

        if isinstance(src_value, (int, float)) and isinstance(dst_value, (int, float)):
            return math.isclose(float(src_value), float(dst_value), rel_tol=1e-9, abs_tol=1e-9)

        return src_value == dst_value

    def validate_uec_values(self, src_workbook: str | Path | None = None,
                            dst_workbook: str | Path | None = None) -> None:
        """Validate that mapped UEC source values match the saved calibration workbook."""
        if self.bats_data:
            self.logger.info("Skipping UEC value validation for BATS mode.")
            return

        src_path = Path(src_workbook) if src_workbook else Path(self.submodel_config["uec_src_file"])
        dst_path = Path(dst_workbook) if dst_workbook else Path(self.workbook_iter)

        if not src_path.exists():
            raise FileNotFoundError(f"UEC source workbook not found: {src_path}")
        if not dst_path.exists():
            raise FileNotFoundError(f"Calibration destination workbook not found: {dst_path}")

        sep = "=" * 80
        self.logger.info(f"\n{sep}\nUEC VALUE VALIDATION\n{sep}")
        self.logger.info(f"Source workbook: {src_path}")
        self.logger.info(f"Destination workbook: {dst_path}")

        src_workbook_obj = open_workbook(src_path)
        dst_workbook_obj = load_workbook(dst_path, data_only=True)
        mismatches = []

        for name, (src_sheet_name, src_column, src_start_row, src_end_row) in self.UEC_SOURCE_RANGES.items():
            dst_sheet_name, dst_column, dst_start_row, dst_end_row = self.CALIBRATION_DESTINATION_RANGES[name]
            src_values = self._read_xls_range(
                src_workbook_obj.sheet_by_name(src_sheet_name),
                src_column,
                src_start_row,
                src_end_row,
            )
            dst_values = self._read_xlsx_range(
                dst_workbook_obj[dst_sheet_name],
                dst_column,
                dst_start_row,
                dst_end_row,
            )

            if len(src_values) != len(dst_values):
                raise ValueError(
                    f"Validation length mismatch for {name}: "
                    f"source has {len(src_values)} values and destination has {len(dst_values)} values"
                )

            for offset, (src_value, dst_value) in enumerate(zip(src_values, dst_values)):
                if not self._values_match(src_value, dst_value):
                    mismatches.append(
                        f"{name}: {src_sheet_name}!R{src_start_row + offset}C{src_column}={src_value!r} "
                        f"!= {dst_sheet_name}!R{dst_start_row + offset}C{dst_column}={dst_value!r}"
                    )

        if mismatches:
            mismatch_preview = "\n".join(mismatches[:10])
            extra_count = len(mismatches) - 10
            if extra_count > 0:
                mismatch_preview += f"\n... and {extra_count} more mismatches"
            raise ValueError(f"UEC value validation failed:\n{mismatch_preview}")

        self.logger.info("Validated UEC source values against the calibration workbook.")
    
    def process_data(self) -> dict:
        """Load inputs, merge spatial attributes, and compute summary statistics.

        Steps:

        1. Read the CT-RAMP (or BATS) work/school location file and the TAZ
           attribute table.
        2. Attach home, work, and school county names via :func:`add_county_info`.
        3. Join skim distances for home→work and home→school pairs.
        4. Optionally merge BATS person weights (``bats_data`` mode only).
        5. Compute a home-county × work-county flow matrix.
        6. Build per-county and total TLFDs (1-mile bins) for work, university,
           and K–12 school trip types.
        7. Compute weighted (BATS) or unweighted/scaled (model) average trip
           lengths by county and trip type.

        Returns:
            A dict with the following keys:

            ``county_summary``
                Wide-format DataFrame of home→work county flows
                (rows = home county, columns = work county).
            ``trip_tlfd_work``
                TLFD DataFrame for work trips
                (distbin column + one column per county + ``Total``).
            ``trip_tlfd_univ``
                TLFD DataFrame for university trips (college or higher).
            ``trip_tlfd_school``
                TLFD DataFrame for K–12 school trips (grade or high school).
            ``avg_trip_lengths``
                Wide-format DataFrame of mean trip distances
                (rows = county, columns = ``work`` / ``univ`` / ``school``).
        """
        # Load input data
        sep = "=" * 80
        self.logger.info(f"\n{sep}\nPROCESS INPUT DATA\n{sep}")
        self.logger.info("Loading input data files:")
        self.logger.info(f'TAZ Data: {self.config.get('data_sources', 'taz_data')}')
        self.logger.info(f"Work School Location: {self.submodel_config['input_file']}")
        taz_data = pd.read_csv(self.config.get('data_sources', 'taz_data'))
        wsloc_results = pd.read_csv(self.submodel_config['input_file'])
        
        # Load distance skim
        dist_skim = pd.read_csv(self.config.get('data_sources', 'dist_skim'), header=0,
                               usecols = ['orig', 'dest', 'DIST'])
        
        # Add Home COUNTY
        self.logger.info("Merging Home County Data")
        wsloc_results = add_county_info(wsloc_results, taz_data, self.county_lookup,
                                       taz_col='HomeTAZ',
                                       county_col_name='HomeCOUNTY',
                                       county_name_col='HomeCounty_name')
        
        # Add Work and School COUNTY
        self.logger.info("Merging Work and School County Data")
        wsloc_results = add_county_info(wsloc_results, taz_data, self.county_lookup,
                                       taz_col='WorkLocation',
                                       county_col_name='WorkCOUNTY',
                                       county_name_col='WorkCounty_name')
        
        wsloc_results = add_county_info(wsloc_results, taz_data, self.county_lookup,
                                       taz_col='SchoolLocation',
                                       county_col_name='SchoolCOUNTY',
                                       county_name_col='SchoolCounty_name')
        
        # Attach distances from distance skim
        self.logger.info("Attaching work distances...")
        work_dist = dist_skim.rename(columns={'orig': 'HomeTAZ', 'dest': 'WorkLocation', 'DIST': 'WorkDist'})
        wsloc_results = wsloc_results.merge(work_dist, on=['HomeTAZ', 'WorkLocation'], how='left')
        
        self.logger.info("Attaching school distances...")
        school_dist = dist_skim.rename(columns={'orig': 'HomeTAZ', 'dest': 'SchoolLocation', 'DIST': 'SchoolDist'})
        wsloc_results = wsloc_results.merge(school_dist, on=['HomeTAZ', 'SchoolLocation'], how='left')
        
        
        # Save enhanced wsloc_results with distances
        if self.bats_data:
            # Join with PersonData to get weights BEFORE processing
            person_data = pd.read_csv(self.submodel_config['bats_person_data'])
            wsloc_results = wsloc_results.merge(person_data[['hh_id', 'person_id', 'person_weight', 'sampleRate']], left_on = ['HHID', 'PersonID'], right_on = ['hh_id', "person_id"])
            wsloc_results.fillna({'person_weight':0}, inplace = True)

            wsloc_with_dist_file = f"{self.target_dir}/MandatoryLocation_with_Distance.csv"
            
            # Process county summary using person weights
            wsloc_county = wsloc_results.groupby(['HomeCOUNTY', 'HomeCounty_name', 'WorkCOUNTY', 'WorkCounty_name'])['person_weight'].sum().reset_index(name='num_pers')
        else:
            wsloc_with_dist_file = f"{self.output_dir}/wsloc_results_with_distances.csv"
            
            # Process county summary using sampleshare
            
            wsloc_county = wsloc_results.groupby(['HomeCOUNTY','HomeCounty_name', 'WorkCOUNTY','WorkCounty_name']).size().reset_index(name='num_pers')
            wsloc_county['num_pers'] = wsloc_county['num_pers'] / self.sampleshare
        
        county_ids = sorted(self.county_lookup)
        wsloc_county_spread = wsloc_county.pivot(index='HomeCOUNTY', columns='WorkCOUNTY', values='num_pers')
        wsloc_county_spread = wsloc_county_spread.reindex(index=county_ids, columns=county_ids)
        wsloc_county_spread = wsloc_county_spread.rename(index=self.county_lookup, columns=self.county_lookup)
        wsloc_county_spread.index.name = 'HomeCounty_name'
        wsloc_county_spread = wsloc_county_spread.fillna(0).reset_index()

        wsloc_results.to_csv(wsloc_with_dist_file, index=False)
        self.logger.info(f"Saved wsloc results with distances to {wsloc_with_dist_file}")

        # Process trip length distributions and averages
        self.logger.info("Processing trip length distributions...")
        trip_types = ['work', 'univ', 'school']
        trip_tlfd_results = {}
        avg_trip_lengths = []
        
        for trip_type in trip_types:
            # Filter data based on trip type and use attached distances           
            if trip_type == 'work':
                filter_cols = ['HomeCounty_name', 'EmploymentCategory', 'WorkDist']
                if self.bats_data:
                    filter_cols.append('person_weight')
                trip_dists = wsloc_results[wsloc_results['WorkLocation'] > 0][filter_cols].copy()
                trip_dists = trip_dists.rename(columns={'WorkDist': 'DIST'})
            elif trip_type == 'univ':
                filter_cols = ['HomeCounty_name', 'StudentCategory', 'SchoolDist']
                if self.bats_data:
                    filter_cols.append('person_weight')
                trip_dists = wsloc_results[
                    (wsloc_results['SchoolLocation'] > 0) & 
                    (wsloc_results['StudentCategory'] == "College or higher")
                ][filter_cols].copy()
                trip_dists = trip_dists.rename(columns={'SchoolDist': 'DIST'})
            elif trip_type == 'school':
                filter_cols = ['HomeCounty_name', 'StudentCategory', 'SchoolDist']
                if self.bats_data:
                    filter_cols.append('person_weight')
                trip_dists = wsloc_results[
                    (wsloc_results['SchoolLocation'] > 0) & 
                    (wsloc_results['StudentCategory'] == "Grade or high school")
                ][filter_cols].copy()
                trip_dists = trip_dists.rename(columns={'SchoolDist': 'DIST'})
            
            # Calculate trip length frequency distribution
            if self.bats_data:
                trip_tlfd = pd.DataFrame({'distbin': range(1, 52)})
            else:
                trip_tlfd = pd.DataFrame({'distbin': range(1, 151)})
            
            # Process by county
            for county in self.county_lookup.values():
                county_trip_dists = trip_dists[trip_dists['HomeCounty_name'] == county]
                
                if len(county_trip_dists) > 0:
                    if self.bats_data:
                        hist_df = create_histogram_tlfd(county_trip_dists['DIST'], bins = range(52),
                                                       weights=county_trip_dists['person_weight'])
                        # Weighted average
                        weighted_mean = np.average(county_trip_dists['DIST'], 
                                                  weights=county_trip_dists['person_weight'])
                    else:
                        hist_df = create_histogram_tlfd(county_trip_dists['DIST'], 
                                                       sampleshare=self.sampleshare)
                        weighted_mean = county_trip_dists['DIST'].mean()
                    
                    # Remove county prefix for column name
                    col_name = county
                    hist_df = hist_df.rename(columns={'count': col_name})
                    trip_tlfd = trip_tlfd.merge(hist_df, on='distbin', how='left')
                    
                    # Add to average trip lengths
                    avg_trip_lengths.append({
                        'county': col_name,
                        'trip_type': trip_type,
                        'mean_trip_length': weighted_mean
                    })
            
            # Total across all counties
            if len(trip_dists) > 0:
                if self.bats_data:
                    hist_df = create_histogram_tlfd(trip_dists['DIST'], bins = range(52),
                                                   weights=trip_dists['person_weight'])
                    # Weighted average
                    total_weighted_mean = np.average(trip_dists['DIST'], 
                                                    weights=trip_dists['person_weight'])
                else:
                    hist_df = create_histogram_tlfd(trip_dists['DIST'], 
                                                   sampleshare=self.sampleshare)
                    total_weighted_mean = trip_dists['DIST'].mean()
                
                hist_df = hist_df.rename(columns={'count': 'Total'})
                trip_tlfd = trip_tlfd.merge(hist_df, on='distbin', how='left')
                
                avg_trip_lengths.append({
                    'county': 'Total',
                    'trip_type': trip_type,
                    'mean_trip_length': total_weighted_mean
                })
            
            # Reorder columns and fill NaN
            county_cols = sorted([col for col in trip_tlfd.columns if col not in ['distbin', 'Total']])
            col_order = ['distbin'] + county_cols + (['Total'] if 'Total' in trip_tlfd.columns else [])
            trip_tlfd = trip_tlfd[col_order].fillna(0)
            
            trip_tlfd_results[trip_type] = trip_tlfd


        # Process average trip lengths
        avg_trip_lengths_df = pd.DataFrame(avg_trip_lengths)
        avg_triplen_spread = avg_trip_lengths_df.pivot(index='county', columns='trip_type', values='mean_trip_length')
        avg_triplen_spread = avg_triplen_spread.reset_index()
        
        # Reorder columns
        desired_cols = ['county'] + [col for col in ['work', 'univ', 'school'] if col in avg_triplen_spread.columns]
        avg_triplen_spread = avg_triplen_spread[desired_cols]
        
        return {
            'county_summary': wsloc_county_spread,
            'trip_tlfd_work': trip_tlfd_results.get('work'),
            'trip_tlfd_univ': trip_tlfd_results.get('univ'),
            'trip_tlfd_school': trip_tlfd_results.get('school'),
            'avg_trip_lengths': avg_triplen_spread
        }
    
    def validate_outputs(self, results: dict):
        """Validate output DataFrames against their Pydantic schemas.

        Called by :meth:`CalibrationBase.run` after :meth:`process_data` and
        before :meth:`generate_outputs`.  Raises ``ValidationError`` if any
        DataFrame does not conform to its expected schema, halting the pipeline
        before any files are written.

        Validates:

        - ``county_summary``   → :class:`CountyTripSummary`
        - ``trip_tlfd_work``   → :class:`TripLengthFrequency`  (51 rows for BATS, 150 for model)
        - ``trip_tlfd_univ``   → :class:`TripLengthFrequency`
        - ``trip_tlfd_school`` → :class:`TripLengthFrequency`
        - ``avg_trip_lengths`` → :class:`AverageTripLength`

        Args:
            results: The dict returned by :meth:`process_data`.
        """
        sep = "=" * 80
        self.logger.info(f"\n{sep}\nOUTPUT VAlIDATION\n{sep}")

        # Validate county summary
        if results['county_summary'] is not None:
            validate_dataframe(results['county_summary'], CountyTripSummary)
            self.logger.info("✓ County Summary Validated")

        # Validate trip length frequency distribution
        expected_rows = 51 if self.bats_data else 150
        for trip_type in ['work', 'univ', 'school']:
            df = results[f'trip_tlfd_{trip_type}']
            if df is not None:
                validate_dataframe(df, TripLengthFrequency, expected_rows)
                self.logger.info(f"✓ {trip_type.capitalize()} TLFD validated")
        
        # Validate average trip lengths
        if results['avg_trip_lengths'] is not None:
            validate_dataframe(results['avg_trip_lengths'], AverageTripLength, )
            self.logger.info("✓ Average Trip Length Summary Validated")

    def generate_outputs(self, results: dict):
        """Write validated results to CSV files and update the calibration workbook.

        In **BATS mode** (``bats_data: true``) outputs are written to
        ``<output_dir>/`` and the corresponding sheets in the Excel workbook
        ("BATS 2023 TLFD", "BATS 2023 AvgTripLen") are updated:

        - ``workTLFD.csv``, ``univTLFD.csv``, ``schoolTLFD.csv``
        - ``AvgTripLen.csv``

        In **model mode** (``bats_data: false``) the county flow matrix, per-
        trip-type TLFDs, and average trip lengths are written to
        ``<output_dir>/`` and the ``modeldata`` sheet of the workbook is
        updated with column offsets that match the template layout.

        In both modes the workbook update is skipped gracefully if no workbook
        template was configured or if the file cannot be opened.

        Args:
            results: The dict returned by :meth:`process_data` and validated
                by :meth:`validate_outputs`.
        """
        sep = "=" * 80
        self.logger.info(f"\n{sep}\nGENERATE OUTPUTS\n{sep}")

        if (self.bats_data):
            trip_types = [('work', 2), ('univ', 15), ('school', 28)]
            for trip_type, col in trip_types:
                if results[f'trip_tlfd_{trip_type}'] is not None:
                    tlfd_file = f"{self.output_dir}/{trip_type}TLFD.csv"
                    results[f'trip_tlfd_{trip_type}'].to_csv(tlfd_file, index = False)
                    self.write_dataframe_to_sheet(results[f'trip_tlfd_{trip_type}'], start_row= 4,  start_col=col, sheet_name="BATS 2023 TLFD",
                                                 source_row=2, source_col=col, source_text=f"Source: {tlfd_file}")
                    
                    self.logger.info(f"Saving trip length frequency distributions for {trip_type} to {tlfd_file}")     
        
            # Average trip lengths
            avg_length_file = f"{self.output_dir}/AvgTripLen.csv"
            results['avg_trip_lengths'].to_csv(avg_length_file, index = False)
            self.write_dataframe_to_sheet(results['avg_trip_lengths'], start_row=3, start_col=1, sheet_name="BATS 2023 AvgTripLen",
                                        source_row=1, source_col=1, source_text=f"Source: {avg_length_file}")
            self.logger.info(f"Saving average trip lengths to {avg_length_file}")     
                
        else: 
            # County summary
            self.logger.info("Generating output files and Excel updates...")

            county_file = f"{self.output_dir}/{self.submodel}_usual_work_school_location_TM_county.csv"

            results['county_summary'].to_csv(county_file, index = False)
            
            self.write_dataframe_to_sheet(results['county_summary'], start_row=4, start_col=1,
                                        source_row=1, source_col=1, source_text=f"Source: {county_file}")

            self.logger.info(f"Saving county summary to {county_file}")            
            # Trip length frequency distributions
            trip_types = [('work', 1), ('univ', 14), ('school', 27)]
            for trip_type, col in trip_types:
                if results[f'trip_tlfd_{trip_type}'] is not None:

                    tlfd_file = f"{self.output_dir}/{self.submodel}_usual_work_school_location_TM_{trip_type}_TLFD.csv"
                    results[f'trip_tlfd_{trip_type}'].to_csv(tlfd_file, index = False)
                    self.write_dataframe_to_sheet(results[f'trip_tlfd_{trip_type}'], start_row=19, start_col=col,
                                                source_row=17, source_col=col, source_text=f"Source: {tlfd_file}")
                    self.logger.info(f"Saving trip length frequency distributions for {trip_type} to {tlfd_file}")            

            # Average trip lengths
            avg_length_file = f"{self.output_dir}/{self.submodel}_usual_work_school_location_TM_avgtriplen.csv"
            results['avg_trip_lengths'].to_csv(avg_length_file, index = False)
            self.write_dataframe_to_sheet(results['avg_trip_lengths'], start_row=4, start_col=14,
                                        source_row=3, source_col=14, source_text=f"Source: {avg_length_file}")
            self.logger.info(f"Saving average trip lengths to {avg_length_file}")            
            


def main():
    """Parse CLI arguments and run the work/school location calibration.

    Constructs a :class:`WorkSchoolLocationCalibration` instance, then calls
    its :meth:`~CalibrationBase.run` method which executes the
    ``process_data → validate_outputs → generate_outputs → validate UECs`` pipeline.
    """
    parser = argparse.ArgumentParser(description="Usual work and school location calibration")
    parser.add_argument("--config", default=None, help="Path to calibration_config.yaml (default: same directory as this script)")
    args = parser.parse_args()

    calibration = WorkSchoolLocationCalibration(config_file=args.config)
    calibration.logger.info("Starting usual work and school location calibration...")
    calibration.run()
    calibration.logger.info("Calibration complete.")


if __name__ == "__main__":
    main()


# %%
