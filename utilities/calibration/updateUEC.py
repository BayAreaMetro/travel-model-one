"""
Script to update UEC (Utility Expression Calculator) files from calibration workbooks.
Converts calibration data from Excel workbooks and updates UEC model files.
"""

import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook
import xlrd
from xlrd import open_workbook
from xlwt import easyxf
from xlutils.copy import copy as xl_copy
import shutil


# Configuration --------------------------------
CALIB_DIR = Path("M:/Development/Travel Model One/Calibration/Version 1.7")

UEC_DIR = Path("X:/travel-model-one-tm1.7_calibration/model-files/model")
BOX_DIR = Path("E:/Box/Modeling and Surveys/Development/Travel Model 1.7/Calibration/workbooks_TM1.7")


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Update UEC files from calibration workbooks'
    )
    parser.add_argument('submodel', type=str, 
                        choices = ['UsualWorkSchool','NonworkDestinationChoice', 'AutomobileOwnership', 
                                   'TourModeChoice', 'TripModeChoice', 'CoordinationDailyActivityPattern'  ],
                       help='Submodel name (e.g., DestinationChoice, AutomobileOwnership)')
    parser.add_argument('version', type=str,
                       help='Version string')
    
    args = parser.parse_args()
    return args.submodel, args.version


def get_config(submodel, version):
    """Get configuration based on submodel type."""
    
    if submodel == "UsualWorkSchool":
        # Note: includes UsualWorkAndSchoolLocation AND NonWorkDestinationChoice
        # so the workbook numbers need to be in sync
        uec_src_workbook = UEC_DIR / "TM1.6.1 version" / "DestinationChoice_TM1.6.1.xls"
        calib_workbook = CALIB_DIR / "01 Usual Work and School Location" / f"01_UsualWorkAndSchoolLocation_2023_{version}.xlsx"
        
        # sheet, column, startRow, endRow
        copy_src = {
            "work":        ("calibration",  4,  4,  8),
            "work_county": ("calibration", 10,  4, 13),
            "university":  ("calibration", 16,  4,  8),
            "highschool":  ("calibration", 33,  4,  8),
            "gradeschool": ("calibration", 34,  4,  8)
        }
        
        copy_dst = {
            "work":        [("Work",        7, 22, 26)],
            "work_county": [("Work",        7, 38, 47)],
            "university":  [("University",  7, 12, 16)],
            "highschool":  [("HighSchool",  7, 12, 16)],
            "gradeschool": [("GradeSchool", 7, 12, 16)]
        }
        
        config = [
            (calib_workbook, copy_src, copy_dst),
        ]

    elif submodel == 'NonworkDestinationChoice':
        uec_src_workbook = UEC_DIR / "TM1.0 version" / "DestinationChoice_TM1.xls"
        calib_workbook = CALIB_DIR / "09 Non-Work Destination Choice" / f"09_NonWorkDestinationChoice_{version}.xlsx"
        # sheet, column, startRow, endRow
        copy_src = {
            "escort1":   ("calibration",  5,  4,  8),
            "escort2":   ("calibration",  5,  4,  8),
            "shopping":  ("calibration", 17,  4,  8),
            "maint":     ("calibration", 29,  4,  8),
            "eatout":    ("calibration", 41,  4,  8),
            "social":    ("calibration", 53,  4,  8),
            "discr":     ("calibration", 65,  4,  8),
            "atwork":    ("calibration", 77,  4,  8)
        }
        
        copy_dst = {
            "escort1":   [("EscortKids",   7, 12, 16)],
            "escort2":   [("EscortNoKids", 7, 12, 16)],
            "shopping":  [("Shopping",     7, 12, 16)],
            "maint":     [("OthMaint",     7, 12, 16)],
            "eatout":    [("EatOut",       7, 12, 16)],
            "social":    [("Social",       7, 12, 16)],
            "discr":     [("OthDiscr",     7, 12, 16)],
            "atwork":    [("WorkBased",    7, 12, 16)]
        }

        config = [
            (calib_workbook, copy_src, copy_dst)
        ]
        
        
    elif submodel == "AutomobileOwnership":
        calib_workbook = CALIB_DIR / "02 Automobile Ownership" / f"02_AutoOwnership_2023_{version}.xlsx"
        uec_src_workbook = UEC_DIR / "TM1.6.1 version" / "AutoOwnership_TM1.6.1.xls"
        
        # sheet, column, startRow, endRow
        copy_src = {
            "1_car_a":    ("calibration", 4, 18, 19),
            "2_cars_a":   ("calibration", 5, 18, 19),
            "3_cars_a":   ("calibration", 6, 18, 19),
            "4+_cars_a":  ("calibration", 7, 18, 19),
            "1_car_sc":   ("calibration", 4, 21, 21),
            "2_cars_sc":  ("calibration", 5, 21, 21),
            "3_cars_sc":  ("calibration", 6, 21, 21),
            "4+_cars_sc": ("calibration", 7, 21, 21),
            "1_car_b":    ("calibration", 4, 24, 27),
            "2_cars_b":   ("calibration", 5, 24, 27),
            "3_cars_b":   ("calibration", 6, 24, 27),
            "4+_cars_b":  ("calibration", 7, 24, 27)
        }
        
        copy_dst = {
            "1_car_a":    [("Auto ownership",  8, 53, 54),
                          ("Auto ownership",  9, 53, 54)],
            "2_cars_a":   [("Auto ownership", 10, 53, 54),
                          ("Auto ownership", 11, 53, 54),
                          ("Auto ownership", 12, 53, 54)],
            "3_cars_a":   [("Auto ownership", 13, 53, 54),
                          ("Auto ownership", 14, 53, 54),
                          ("Auto ownership", 15, 53, 54),
                          ("Auto ownership", 16, 53, 54)],
            "4+_cars_a":  [("Auto ownership", 17, 53, 54)],
            
            "1_car_sc":   [("Auto ownership",  8, 59, 59),
                          ("Auto ownership",  9, 59, 59)],
            "2_cars_sc":  [("Auto ownership", 10, 59, 59),
                          ("Auto ownership", 11, 59, 59),
                          ("Auto ownership", 12, 59, 59)],
            "3_cars_sc":  [("Auto ownership", 13, 59, 59),
                          ("Auto ownership", 14, 59, 59),
                          ("Auto ownership", 15, 59, 59),
                          ("Auto ownership", 16, 59, 59)],
            "4+_cars_sc": [("Auto ownership", 17, 59, 59)],
            
            "1_car_b":    [("Auto ownership",  8, 55, 58),
                          ("Auto ownership",  9, 55, 58)],
            "2_cars_b":   [("Auto ownership", 10, 55, 58),
                          ("Auto ownership", 11, 55, 58),
                          ("Auto ownership", 12, 55, 58)],
            "3_cars_b":   [("Auto ownership", 13, 55, 58),
                          ("Auto ownership", 14, 55, 58),
                          ("Auto ownership", 15, 55, 58),
                          ("Auto ownership", 16, 55, 58)],
            "4+_cars_b":  [("Auto ownership", 17, 55, 58)]
        }
        
        config = [(calib_workbook, copy_src, copy_dst)]
        
    elif submodel == "TourModeChoice":
        calib_workbook = CALIB_DIR / "11 Tour Mode Choice" / f"11_TourModeChoice_{version}.xlsx"
        uec_src_workbook = UEC_DIR / "TM1.5.1 version" / "ModeChoice_TM1.5.1.xls"
        
        # sheet, column, startRow, endRow
        copy_src = {
            "work":          ("constants",  4,  3, 64),
            "university":    ("constants",  8,  3, 64),
            "school":        ("constants", 12,  3, 64),
            "escort":        ("constants", 16,  3, 64),
            "shopping":      ("constants", 20,  3, 64),
            "eatout":        ("constants", 24,  3, 64),
            "othmaint":      ("constants", 28,  3, 64),
            "social":        ("constants", 32,  3, 64),
            "othdiscr":      ("constants", 36,  3, 64),
            "workbased":     ("constants", 40,  3, 64),
            
            "cbd_work":      ("CBD_SF",     9,  3, 10),
            "cbd_university":("CBD_SF",     9, 11, 18),
            "cbd_school":    ("CBD_SF",     9, 19, 26),
            "cbd_escort":    ("CBD_SF",     9, 27, 34),
            "cbd_shopping":  ("CBD_SF",     9, 27, 34),
            "cbd_eatout":    ("CBD_SF",     9, 35, 42),
            "cbd_othmaint":  ("CBD_SF",     9, 27, 34),
            "cbd_social":    ("CBD_SF",     9, 35, 42),
            "cbd_othdiscr":  ("CBD_SF",     9, 35, 42),
            "cbd_workbased": ("CBD_SF",     9, 43, 50)
        }
        
        copy_dst = {
            "work":          [("Work",       5, 408, 469)],
            "university":    [("University", 5, 408, 469)],
            "school":        [("School",     5, 408, 469)],
            "escort":        [("Escort",     5, 408, 469)],
            "shopping":      [("Shopping",   5, 408, 469)],
            "eatout":        [("EatOut",     5, 408, 469)],
            "othmaint":      [("OthMaint",   5, 408, 469)],
            "social":        [("Social",     5, 408, 469)],
            "othdiscr":      [("OthDiscr",   5, 408, 469)],
            "workbased":     [("WorkBased",  5, 411, 472)],
            
            "cbd_work":      [("Work",       5, 470, 477)],
            "cbd_university":[("University", 5, 470, 477)],
            "cbd_school":    [("School",     5, 470, 477)],
            "cbd_escort":    [("Escort",     5, 470, 477)],
            "cbd_shopping":  [("Shopping",   5, 470, 477)],
            "cbd_eatout":    [("EatOut",     5, 470, 477)],
            "cbd_othmaint":  [("OthMaint",   5, 470, 477)],
            "cbd_social":    [("Social",     5, 470, 477)],
            "cbd_othdiscr":  [("OthDiscr",   5, 470, 477)],
            "cbd_workbased": [("WorkBased",  5, 473, 480)]
        }
        
        config = [(calib_workbook, copy_src, copy_dst)]
        
    elif submodel == "TripModeChoice":
        calib_workbook = CALIB_DIR / "15 Trip Mode Choice" / f"15_TripModeChoice_{version}.xlsx"
        uec_src_workbook = UEC_DIR / "TM1.5.1 version" / "TripModeChoice_TM1.5.1.xls"
        
        # sheet, column, startRow, endRow
        copy_src = {
            "work":       ("constants",  7,  3, 39),
            "university": ("constants", 15,  3, 39),
            "school":     ("constants", 23,  3, 39),
            "escort":     ("constants", 31,  3, 39),  # indiv maint, indiv
            "shopping":   ("constants", 31,  3, 68),  # indiv maint, joint
            "eatout":     ("constants", 39,  3, 68),  # indiv disc, joint
            "othmaint":   ("constants", 31,  3, 68),  # indiv maint, joint
            "social":     ("constants", 39,  3, 68),  # indiv disc, joint
            "othdiscr":   ("constants", 39,  3, 68),  # indiv disc
            "workbased":  ("constants", 47,  3, 39)   # at work
        }
        
        copy_dst = {
            "work":       [("Work",       5, 509, 545)],
            "university": [("University", 5, 512, 548)],
            "school":     [("School",     5, 512, 548)],
            "escort":     [("Escort",     5, 512, 548)],
            "shopping":   [("Shopping",   5, 512, 577)],
            "eatout":     [("EatOut",     5, 512, 577)],
            "othmaint":   [("OthMaint",   5, 512, 577)],
            "social":     [("Social",     5, 512, 577)],
            "othdiscr":   [("OthDiscr",   5, 512, 577)],
            "workbased":  [("WorkBased",  5, 511, 547)]
        }
        
        config = [(calib_workbook, copy_src, copy_dst)]
        
    else:
        raise ValueError(f"Don't understand SUBMODEL [{submodel}]")
    
    return uec_src_workbook, config


def read_column_data(worksheet, column, start_row, end_row):
    """Read data from a specific column range in a Calibration Worksheet."""
    data = []
    for row in range(start_row, end_row + 1):
        cell_value = worksheet.cell(row=row, column=column).value
        if cell_value is None:
            raise ValueError(f"Data contains None/NA at row {row}")
        data.append(cell_value)
    return data


def write_column_data(worksheet, data, column, start_row):
    """Write data to a specific column starting at start_row in UEC Workbook."""
    for i, value in enumerate(data):
        # xlwt uses 0-based (row, col) indexes.
        worksheet.write(start_row + i - 1, column - 1, value)


def main():
    """Main function to update UEC files."""
    submodel, version = parse_arguments()
    
    print(f"SUBMODEL = {submodel}")
    print(f"ITERATION  = {version}")
    
    # Get configuration
    uec_src_workbook, config = get_config(submodel, version)
    
    print(f"UEC_SRC_WORKBOOK = {uec_src_workbook}")
    
    # Determine destination workbook name
    uec_dst_workbook = Path(str(uec_src_workbook)
                            .replace("TM1.0 version\\", "")
                            .replace("TM1.5.1 version\\", "")
                            .replace("TM1.6.1 version\\", "")
                            .replace("_TM1.6.1", "")
                            .replace("_TM1.5.1", "")
                            .replace("_TM1", ""))

    
    print(f"UEC_DST_WORKBOOK = {uec_dst_workbook}")
    
    # Load xls workbook (read) then create a writable copy.
    uec_read_workbook = open_workbook(uec_src_workbook, formatting_info=True)
    uec_workbook = xl_copy(uec_read_workbook)
    sheet_index_by_name = {
        name: idx for idx, name in enumerate(uec_read_workbook.sheet_names())
    }
    
    # Process each configuration set
    for calib_workbook, copy_src, copy_dst in config:
        print("-----------------------------------------------")
        print(f"CALIB_WORKBOOK   = {calib_workbook}")
        
        # Load calibration workbook
        calib_workbook_obj = load_workbook(calib_workbook, data_only=True)
        
        # Process each copy operation
        for name in copy_src.keys():
            calib_sheetname, calib_column, calib_start_row, calib_end_row = copy_src[name]
            
            print(f"Reading {name} from sheet {calib_sheetname} @ "
                  f"({calib_start_row},{calib_column}) - ({calib_end_row},{calib_column})")
            
            # Read data from calibration workbook
            calib_sheet = calib_workbook_obj[calib_sheetname]
            data = read_column_data(calib_sheet, calib_column, calib_start_row, calib_end_row)
            print(data)
            
            # Write data to each destination in UEC workbook
            destinations = copy_dst[name]
            for copynum, (uec_sheetname, uec_column, uec_start_row, uec_end_row) in enumerate(destinations, 1):
                print(f"Copying set {copynum}")
                print(f"Copying to {uec_sheetname} @ "
                      f"({uec_start_row},{uec_column}) - ({uec_end_row},{uec_column})")
                
                read_sheet = uec_read_workbook.sheet_by_name(uec_sheetname)
                write_sheet = uec_workbook.get_sheet(sheet_index_by_name[uec_sheetname])
                
                # Get value after last cell (if any) to preserve it
                uec_suffix_row = uec_end_row + 1
                cell_after_last_val = read_sheet.cell_value(uec_suffix_row - 1, uec_column - 1)
                
                if cell_after_last_val in (None, ""):
                    print("Cell after last is null")
                else:
                    print(f"Cell after last: {cell_after_last_val}")
                
                # Write the data
                write_column_data(write_sheet, data, uec_column, uec_start_row)
                
                # Restore the cell after last if it had a value
                if cell_after_last_val not in (None, ""):
                    write_sheet.write(uec_suffix_row - 1, uec_column - 1, cell_after_last_val)
        
        # Copy calibration workbook to Box directory (without version suffix)
        calib_workbook_no_vers = calib_workbook.name.replace(f"_{version}.xlsx", ".xlsx")
        box_dest = BOX_DIR / calib_workbook_no_vers
        print(f"Copying {calib_workbook} to {box_dest}")
        shutil.copy2(calib_workbook, box_dest)
    
    # Save the updated UEC workbook
    uec_workbook.save(uec_dst_workbook)
    print(f"Wrote {uec_dst_workbook}")


if __name__ == "__main__":
    main()
