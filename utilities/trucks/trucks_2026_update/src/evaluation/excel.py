import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
from matplotlib.figure import Figure

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

PALETTE = ["#4E79A7", "#F28E2B", "#59A14F", "#E15759", "#B07AA1", "#76B7B2"]

TRUCK_TYPES = ["HV", "SM"]
TRUCK_LABELS = {"HV": "Heavy Trucks (HV)", "SM": "Very Small, Small & Medium Trucks (SM)"}

# Excel styling constants.
FONT_NAME = "Calibri"
FONT_SIZE = 11
HEADER_GRAY = "404040"
ALT_ROW_GRAY = "F2F2F2"
INT_FORMAT = "#,##0"
DECIMAL_FORMAT = "#,##0.00"
PCT_FORMAT = "+0.0%;-0.0%"
DIFF_POS_FILL = "FFCCCC"  # light red: predicted over-estimates observed
DIFF_NEG_FILL = "CCFFCC"  # light green: predicted under-estimates observed

def write_excel(
    cfg: dict,
    completed_scenarios: list[dict],
    scenario_color_map: dict[str, str],
    trip_gen_table: pd.DataFrame,
    vmt_table: pd.DataFrame,
    trip_lengths: pd.DataFrmae, 
    trip_distribution_figures: dict[tuple[str, str], Figure],
    scatter_figures: dict[tuple[str, str], Figure],
    vmt_figures: dict[str, Figure],
    output_dir: Path,
) -> None:
    """
    Write the comparison workbook with summary tables and embedded figures.

    Parameters
    ----------
    cfg : dict
        Full config dict (used for the Context sheet metadata).
    completed_scenarios : list of dict
        Scenario dicts with ``"name"`` and ``"path"`` keys.
    scenario_color_map : dict of str to str
        Scenario name to hex color.
    trip_gen_table : pd.DataFrame
        Output of :func:`build_trip_gen_table`.
    vmt_table : pd.DataFrame
        Output of :func:`build_vmt_table`.
    scatter_figures : dict of (str, str) to Figure
        Output of ``plot_scatter_all_scenarios``.
    vmt_figures : dict of str to Figure
        Output of ``plot_vmt_comparison``.
    output_dir : Path
        Destination directory; the workbook is written to
        ``truck_model_evaluation.xlsx`` inside it.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    scenario_names = [s["name"] for s in completed_scenarios]

    _write_context_sheet(wb, cfg, completed_scenarios)
    _write_table_sheet(
        wb, "Trip Generation", trip_gen_table, scenario_names, scenario_color_map,
        totals_label="Total productions",
    )
    _write_table_sheet(
        wb, "VMT by Type", vmt_table, scenario_names, scenario_color_map,
        totals_label="Total",
    )

    _write_table_sheet(
         wb, "Average Trip Length", trip_lengths, scenario_names, scenario_color_map,
        totals_label="Total",
    )

    _write_plot_sheet(
        wb=wb,
        sheet_name="Trip Distributions Plots",
        figure_map=trip_distribution_figures,
        row_labels=["Very Small", "Small", "Medium", "Large"],
        scenario_names=scenario_names,
        scenario_color_map=scenario_color_map,
        )

    _write_plot_sheet(
        wb=wb,
        sheet_name="Scatter Plots",
        figure_map=scatter_figures,
        row_labels=["HV", "SM"],
        scenario_names=scenario_names,
        scenario_color_map=scenario_color_map,
        )
    
    _write_vmt_comparison_sheet(
        wb,
        vmt_figures,
    )


    out_path = Path(output_dir) / "truck_model_evaluation.xlsx"
    wb.save(out_path)
    logger.info("Wrote Excel workbook: %s", out_path)


def _tint(hex_color: str, factor: float = 0.2) -> str:
    """Mix ``hex_color`` with white (``factor`` = share of color), return 6-hex."""
    hex_color = hex_color.lstrip("#")
    r, g, b = (int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    r = round(255 * (1 - factor) + r * factor)
    g = round(255 * (1 - factor) + g * factor)
    b = round(255 * (1 - factor) + b * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def _fill(hex_color: str) -> PatternFill:
    """Solid fill from a hex color string (with or without leading ``#``)."""
    return PatternFill(
        start_color=hex_color.lstrip("#"),
        end_color=hex_color.lstrip("#"),
        fill_type="solid",
    )


def _style_header(cell, fill_hex: str, font_color: str = "FFFFFF") -> None:
    """Apply the standard bold colored header style to a cell."""
    cell.fill = _fill(fill_hex)
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _data_font() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE)


def _autosize(ws: Worksheet) -> None:
    """Set each column's width to fit its longest cell value."""
    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            letter = cell.column_letter
            widths[letter] = max(widths.get(letter, 0), len(str(cell.value)))
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)


def _embed_figure(ws: Worksheet, fig: Figure, anchor: str, width_px: int, height_px: int) -> None:
    """Embed a matplotlib figure into ``ws`` at ``anchor`` as a sized PNG."""
    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=300, bbox_inches="tight")
    buffer.seek(0)
    img = XLImage(buffer)
    img.width = width_px
    img.height = height_px
    ws.add_image(img, anchor)


def _write_context_sheet(wb: Workbook, cfg: dict, completed_scenarios: list[dict]) -> None:
    """Write the metadata / scenario-listing Context sheet."""
    ws = wb.create_sheet("Context")
    ws.sheet_view.showGridLines = False

    header_fill = _fill(HEADER_GRAY)
    header_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")

    # Header row.
    for col, label in enumerate(["Field", "Value"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font

    rows = [
        ("Generated on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Config file", "configs/travel_model_scenarios.yaml"),
        ("Observed data", str(cfg.get("observed_data", ""))),
        ("Output folder", str(cfg.get("evaluation_output", ""))),
        ("Scenarios run", ""),
    ]
    for s in completed_scenarios:
        rows.append((s["name"], s["path"]))

    for i, (field, value) in enumerate(rows, start=2):
        fill = _fill(ALT_ROW_GRAY) if i % 2 == 0 else None
        for col, text in enumerate((field, value), start=1):
            cell = ws.cell(row=i, column=col, value=text)
            cell.font = _data_font()
            if fill is not None:
                cell.fill = fill

    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_table_sheet(
    wb: Workbook,
    sheet_name: str,
    table: pd.DataFrame,
    scenario_names: list[str],
    scenario_color_map: dict[str, str],
    totals_label: str,
) -> None:
    """Write a summary table sheet (Trip Generation / VMT by Type)."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    present = [name for name in scenario_names if name in table.columns]

    # --- Header row ---
    ws.cell(row=1, column=1, value="Truck Type").font = Font(
        name=FONT_NAME, size=FONT_SIZE, bold=True
    )
    col = 2
    scenario_cols: dict[str, int] = {}
    for name in present:
        cell = ws.cell(row=1, column=col, value=name)
        _style_header(cell, scenario_color_map.get(name, "#4E79A7"))
        scenario_cols[name] = col
        col += 1

    # --- % diff columns vs the first scenario ---
    diff_cols: dict[str, int] = {}
    if len(present) >= 2:
        base = present[0]
        for name in present[1:]:
            cell = ws.cell(row=1, column=col, value=f"% diff vs {base}")
            _style_header(cell, _tint(scenario_color_map.get(name, "#4E79A7")), font_color="000000")
            diff_cols[name] = col
            col += 1

    # --- Data rows ---
    for r, truck_type in enumerate(table.index, start=2):
        label_cell = ws.cell(row=r, column=1, value=str(truck_type))
        label_cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        for name in present:
            value = table.loc[truck_type, name]
            cell = ws.cell(row=r, column=scenario_cols[name], value=_num(value))
            cell.font = _data_font()
            cell.number_format = INT_FORMAT if sheet_name != "Average Trip Length" else DECIMAL_FORMAT
        if len(present) >= 2:
            base = present[0]
            base_val = table.loc[truck_type, base]
            for name in present[1:]:
                cell = ws.cell(row=r, column=diff_cols[name])
                cell.font = _data_font()
                cell.fill = _fill(_tint(scenario_color_map.get(name, "#4E79A7")))
                cell.number_format = PCT_FORMAT
                val = table.loc[truck_type, name]
                if pd.notna(base_val) and base_val != 0 and pd.notna(val):
                    cell.value = (val - base_val) / base_val

    # --- Totals row ---
    total_row = len(table.index) + 2
    total_cell = ws.cell(row=total_row, column=1, value=totals_label)
    total_cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    for name in present:
        total = table[name].sum(skipna=True)
        cell = ws.cell(row=total_row, column=scenario_cols[name], value=_num(total))
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        cell.number_format = INT_FORMAT if sheet_name != "Average Trip Length" else DECIMAL_FORMAT

    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_scatter_sheet(
    wb: Workbook,
    scenario_name: str,
    truck_type: str,
    fig: Figure,
    scenario_color_map: dict[str, str],
) -> None:
    """Write a scatter sheet: embedded figure, fit-statistics, and a QA data table."""
    sheet_name = _safe_sheet_name(f"Scatter {scenario_name} {truck_type}")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    color = scenario_color_map.get(scenario_name, "#4E79A7")
    ws.sheet_properties.tabColor = color.lstrip("#")

    _embed_figure(ws, fig, "A1", width_px=500, height_px=430)

    stats = getattr(fig, "scenario_stats", {}) or {}
    start_row = 33
    for col, label in enumerate(["Metric", "Value"], start=1):
        cell = ws.cell(row=start_row, column=col, value=label)
        _style_header(cell, color)

    rows = [
        ("Slope", _fmt_stat(stats.get("slope"))),
        ("Intercept", _fmt_stat(stats.get("intercept"))),
        ("R²", _fmt_stat(stats.get("r2"))),
        ("Count locations (n)", stats.get("n")),
    ]
    for i, (label, value) in enumerate(rows, start=start_row + 1):
        ws.cell(row=i, column=1, value=label).font = _data_font()
        ws.cell(row=i, column=2, value=value).font = _data_font()

    # Raw link-level QA table, below the stats (blank-row separator at row 41).
    data = getattr(fig, "scatter_data", None)
    if data is not None and len(data):
        _write_scatter_data_table(ws, data, color, header_row=42)

    _autosize(ws)


def _write_scatter_data_table(ws: Worksheet, data: pd.DataFrame, color: str, header_row: int) -> None:
    """Write the per-count-location QA table (observed/predicted/diff/pct_diff).

    Rows come pre-sorted by ``observed`` descending. The ``diff`` column is fill-
    coded per sign (light red over-estimate, light green under-estimate) and
    ``pct_diff`` is written as a signed string so the sign always shows. The
    table's first data row is frozen.
    """
    headers = ["link_id", "observed", "predicted", "diff", "pct_diff"]
    for col, label in enumerate(headers, start=1):
        _style_header(ws.cell(row=header_row, column=col, value=label), color)

    for offset, (_, record) in enumerate(data.iterrows(), start=1):
        r = header_row + offset
        observed = record["observed"]
        predicted = record["predicted"]
        diff = record["diff"]

        ws.cell(row=r, column=1, value=record["link_id"]).font = _data_font()

        for col, value in ((2, observed), (3, predicted)):
            cell = ws.cell(row=r, column=col, value=_num(value))
            cell.font = _data_font()
            cell.number_format = INT_FORMAT

        diff_cell = ws.cell(row=r, column=4, value=_num(diff))
        diff_cell.font = _data_font()
        diff_cell.number_format = INT_FORMAT
        if pd.notna(diff) and diff > 0:
            diff_cell.fill = _fill(DIFF_POS_FILL)
        elif pd.notna(diff) and diff < 0:
            diff_cell.fill = _fill(DIFF_NEG_FILL)

        pct_cell = ws.cell(row=r, column=5, value=_pct_diff_str(observed, predicted))
        pct_cell.font = _data_font()

    # Freeze the table header (everything above the first data row stays in view).
    ws.freeze_panes = f"A{header_row + 1}"


def _pct_diff_str(observed, predicted) -> str:
    """Signed percent-difference string, or ``"n/a"`` when observed is 0/missing."""
    if pd.isna(observed) or pd.isna(predicted) or observed == 0:
        return "n/a"
    return f"{(predicted - observed) / observed * 100:+.1f}%"


def _write_vmt_sheet(
    wb: Workbook,
    truck_type: str,
    fig: Figure,
    scenario_color_map: dict[str, str],
) -> None:
    """Write a per-truck-type VMT sheet: embedded figure plus the raw numbers.

    The numbers table is read from ``fig.vmt_values`` (``{category: VMT}``) so it
    matches the embedded chart exactly — ``Observed`` first, then each scenario.
    """
    ws = wb.create_sheet(f"VMT - {truck_type}")
    ws.sheet_view.showGridLines = False

    _embed_figure(ws, fig, "A1", width_px=430, height_px=360)

    start_row = 28
    for col, label in enumerate(["Category", "VMT"], start=1):
        cell = ws.cell(row=start_row, column=col, value=label)
        _style_header(cell, HEADER_GRAY)

    values: dict[str, float] = getattr(fig, "vmt_values", {}) or {}
    for i, (category, value) in enumerate(values.items(), start=start_row + 1):
        name_cell = ws.cell(row=i, column=1, value=category)
        name_cell.font = _data_font()
        if category != "Observed":
            name_cell.fill = _fill(_tint(scenario_color_map.get(category, "#4E79A7")))
        value_cell = ws.cell(row=i, column=2, value=_num(value))
        value_cell.font = _data_font()
        value_cell.number_format = INT_FORMAT


def _write_plot_sheet(
    wb: Workbook,
    sheet_name: str,
    figure_map: dict[tuple[str, str], Figure],
    row_labels: list[str],
    scenario_names: list[str],
    scenario_color_map: dict[str, str],
    image_width_px: int = 420,
    image_height_px: int = 320,
) -> None:
    """
    Write a comparison plot sheet where:

        Columns = scenarios
        Rows    = plot types

    Parameters
    ----------
    wb : Workbook
    sheet_name : str
        Name of worksheet.
    figure_map : dict[(scenario_name, plot_name), Figure]
        Mapping of scenarios and plot labels to figures.
    row_labels : list[str]
        Plot names displayed down the sheet.
    scenario_names : list[str]
        Scenario names displayed across the sheet.
    scenario_color_map : dict[str, str]
        Scenario colors used in column headers.
    image_width_px : int
    image_height_px : int
    """
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    ws.sheet_view.showGridLines = False

    # --------------------------------------------------
    # Header row
    # --------------------------------------------------

    ws.cell(row=1, column=1, value="Plot Type")

    for col_idx, scenario_name in enumerate(scenario_names, start=2):

        cell = ws.cell(
            row=1,
            column=col_idx,
            value=scenario_name,
        )

        color = scenario_color_map.get(
            scenario_name,
            "#4E79A7",
        )

        _style_header(cell, color)

        ws.column_dimensions[
            get_column_letter(col_idx)
        ].width = 60

    ws.column_dimensions["A"].width = 30

    # --------------------------------------------------
    # Layout calculations
    # --------------------------------------------------

    rows_per_plot = 20

    # --------------------------------------------------
    # Plot rows
    # --------------------------------------------------

    for plot_idx, plot_name in enumerate(row_labels):

        start_row = 2 + plot_idx * rows_per_plot

        label_cell = ws.cell(
            row=start_row,
            column=1,
            value=plot_name,
        )

        label_cell.font = Font(
            name=FONT_NAME,
            size=FONT_SIZE,
            bold=True,
            color="000000"
            )
        label_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        # Add each scenario figure
        for col_idx, scenario_name in enumerate(
            scenario_names,
            start=2,
        ):

            fig = figure_map.get(
                (scenario_name, plot_name)
            )

            if fig is None:
                continue

            anchor = (
                f"{get_column_letter(col_idx)}"
                f"{start_row}"
            )

            _embed_figure(
                ws,
                fig,
                anchor,
                width_px=image_width_px,
                height_px=image_height_px,
            )

        # Make the plot block tall enough
        for row in range(
            start_row,
            start_row + rows_per_plot,
        ):
            ws.row_dimensions[row].height = 18

    # --------------------------------------------------
    # Freeze panes
    # --------------------------------------------------

    ws.freeze_panes = "B2"

def _write_vmt_comparison_sheet(
    wb: Workbook,
    vmt_figures: dict[str, Figure],
) -> None:

    ws = wb.create_sheet("VMT Comparison")
    ws.sheet_view.showGridLines = False

    hv_fig = vmt_figures.get("HV")
    sm_fig = vmt_figures.get("SM")

    if hv_fig is not None:
        ws["A1"] = "Heavy Trucks (HV)"
        _embed_figure(
            ws,
            hv_fig,
            "A2",
            width_px=600,
            height_px=400,
        )

    if sm_fig is not None:
        ws["A28"] = "Very Small, Small & Medium Trucks (SM)"
        _embed_figure(
            ws,
            sm_fig,
            "A29",
            width_px=600,
            height_px=400,
        )


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def _num(value) -> float | int | None:
    """Return a numeric value for Excel, or ``None`` for NaN/missing."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    if pd.isna(value):
        return None
    return value


def _fmt_stat(value) -> float | None:
    """Round a fit statistic for display, or ``None`` if missing."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    return round(float(value), 4)


def _safe_sheet_name(name: str) -> str:
    """Truncate to Excel's 31-character sheet-name limit."""
    return name[:31]