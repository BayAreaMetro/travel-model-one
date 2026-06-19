"""Truck-Updates evaluation pipeline.
Summary comparison of multiple scenario runs.

* a Tableau-ready validation shapefile (preserved from the original
  ``run_evaluation.py``),
* per-scenario / per-truck-type observed-vs-predicted scatter PNGs,
* per-truck-type VMT comparison PNGs, and
* a single comparison Excel workbook that embeds the same figures.

The design is deliberately flat: one function per output, one file per plot
type under ``plots/``. Adding a plot means adding a file there and two lines in
:func:`run_evaluation`. Steps fail loudly but independently — one broken step
never blocks the others.
"""
from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.figure import Figure

import numpy as np
import pandas as pd
import geopandas as gpd

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.utils import setup_logging, save

logger = logging.getLogger(__name__)

# Single source of truth for the scenario color palette. Plot functions and the
# Excel writer receive the assigned colors as an argument — they never import
# this list directly.
PALETTE = ["#4E79A7", "#F28E2B", "#59A14F", "#E15759", "#B07AA1", "#76B7B2"]

TRUCK_TYPES = ["HV", "SM"]
TRUCK_LABELS = {"HV": "Heavy Trucks (HV)", "SM": "Very Small, Small & Medium Trucks (SM)"}

# Excel styling constants.
FONT_NAME = "Calibri"
FONT_SIZE = 11
HEADER_GRAY = "404040"
ALT_ROW_GRAY = "F2F2F2"
INT_FORMAT = "#,##0"
PCT_FORMAT = "+0.0%;-0.0%"


def run_evaluation(cfg: dict, completed_scenarios: list[dict]) -> None:
    """
    Run every evaluation output for the scenarios that completed successfully.

    Parameters
    ----------
    cfg : dict
        The full config loaded from ``travel_model_scenarios.yaml``. Must
        contain ``"observed_data"``, ``"network_crs"``, and
        ``"evaluation_output"``.
    completed_scenarios : list of dict
        Scenario dicts (each with ``"name"`` and ``"path"`` keys) for runs that
        succeeded. Scenarios that failed are already excluded by the caller.

    Notes
    -----
    Each sub-step is wrapped in its own try/except: a failure is logged with a
    traceback and the remaining steps still run. Nothing here re-raises.
    """
    setup_logging(log_dir="data/logs", log_name="evaluation")

    logger.info("=" * 60)
    logger.info("Starting truck model evaluation pipeline")
    logger.info("Scenarios to evaluate: %s", [s["name"] for s in completed_scenarios])
    logger.info("=" * 60)

    if not completed_scenarios:
        logger.warning("No completed scenarios to evaluate — nothing to do.")
        return

    output_dir = Path(cfg["evaluation_output"])
    output_dir.mkdir(parents=True, exist_ok=True)
    plots_dir = output_dir / "plots"
    plots_dir.mkdir(exist_ok=True)

    logger.info("Reading observed data")
    observed = read_observed(cfg)
    scenario_color_map = assign_colors(completed_scenarios)

    # --- Tables ---
    logger.info("Building trip generation table")
    trip_gen_table = _safe(build_trip_gen_table, completed_scenarios, default=pd.DataFrame())
    logger.info("Building VMT table (full network)")
    vmt_table = _safe(build_vmt_table, completed_scenarios, default=pd.DataFrame())

    # Observed-links-only VMT for an apples-to-apples chart comparison.
    observed_link_ids = set(observed["link_id"].astype(str))
    vmt_table_obs = _safe(
        build_vmt_table_observed_links_only,
        completed_scenarios,
        observed_link_ids,
        default=pd.DataFrame(),
    )

    # --- Plots (one function per file in plots/) ---
    from src.evaluation.plots.scatter_obs_vs_pred import plot_scatter_all_scenarios
    from src.evaluation.plots.vmt_comparison import plot_vmt_comparison

    logger.info("Building observed-vs-predicted scatter plots")
    scatter_figures = _safe(
        plot_scatter_all_scenarios,
        completed_scenarios,
        observed,
        scenario_color_map,
        default={},
    )
    logger.info("Building VMT comparison plots")
    vmt_figures = _safe(
        plot_vmt_comparison, vmt_table_obs, observed, scenario_color_map, default={}
    )

    # --- Excel workbook (embeds the same figures, before the figures are closed) ---
    try:
        write_excel(
            cfg=cfg,
            completed_scenarios=completed_scenarios,
            scenario_color_map=scenario_color_map,
            trip_gen_table=trip_gen_table,
            vmt_table=vmt_table,
            scatter_figures=scatter_figures,
            vmt_figures=vmt_figures,
            output_dir=output_dir,
        )
    except Exception:
        logger.exception("Failed to write Excel workbook")

    # --- PNG export (also closes all figures) ---
    try:
        save_pngs(scatter_figures, vmt_figures, plots_dir)
    except Exception:
        logger.exception("Failed to save PNG figures")

    # --- Tableau shapefile (preserved from the original run_evaluation.py) ---
    try:
        save_tableau_shapefile(completed_scenarios, observed, cfg)
    except Exception:
        logger.exception("Failed to write Tableau shapefile")

    logger.info("Evaluation pipeline finished. Outputs in %s", output_dir)


def _safe(func, *args, default=None):
    """Call ``func(*args)`` returning ``default`` (logging a traceback) on error."""
    try:
        return func(*args)
    except Exception:
        logger.exception("Step %s failed", getattr(func, "__name__", func))
        return default


def read_observed(cfg: dict) -> pd.DataFrame:
    """
    Read the observed truck-count dataset.

    Parameters
    ----------
    cfg : dict
        Config dict; ``cfg["observed_data"]`` is the path to the CSV.

    Returns
    -------
    pd.DataFrame
        The observed dataset. A ``"truck_type_norm"`` column is added,
        normalising ``"truck_type_2"`` so tolled categories collapse onto the
        two assignment categories (``"HVT"`` → ``"HV"``, ``"SMT"`` → ``"SM"``).
        The geometry column is left as-is (only needed by the shapefile output).

    Notes
    -----
    Returned as a plain ``DataFrame`` rather than a ``GeoDataFrame`` — geometry
    is only materialised when writing the Tableau shapefile.
    """
    df = pd.read_csv(cfg["observed_data"])
    if "truck_type_2" in df.columns:
        df["truck_type_norm"] = (
            df["truck_type_2"].astype(str).str.replace(r"T$", "", regex=True)
        )
    if "link_id" in df.columns:
        df["link_id"] = df["link_id"].astype(str)
    return df


def read_truck_tg(scenario_path: Path) -> dict[str, int] | None:
    """
    Read total truck-production trips by truck type from ``nonres/TruckTG.dat``.

    Parameters
    ----------
    scenario_path : Path
        Root of a scenario's output directory.

    Returns
    -------
    dict of str to int or None
        Total production trips keyed by truck type
        (``"Very Small"``, ``"Small"``, ``"Medium"``, ``"Large"``). Returns
        ``None`` if the file does not exist.

    Raises
    ------
    FileNotFoundError
        Propagated by the caller's handling — see Notes.

    Notes
    -----
    ``TruckTG.dat`` is a fixed-width file with no header and nine 12-character
    columns: ``zone`` followed by production/attraction pairs for the four truck
    sizes. Only productions are summed; the model balances attractions to
    productions, so they would be redundant.
    """
    path = Path(scenario_path) / "nonres" / "TruckTG.dat"
    if not path.exists():
        raise FileNotFoundError(f"TruckTG.dat not found at {path}")

    df = pd.read_fwf(
        path,
        widths=[12] * 9,
        header=None,
        names=[
            "zone", "verySmallP", "verySmallA", "smallP", "smallA",
            "mediumP", "mediumA", "largeP", "largeA",
        ],
    )

    return {
        "Very Small": int(df["verySmallP"].sum()),
        "Small": int(df["smallP"].sum()),
        "Medium": int(df["mediumP"].sum()),
        "Large": int(df["largeP"].sum()),
    }


def read_network(scenario_path: Path) -> gpd.GeoDataFrame:
    """
    Read a scenario's loaded highway network and derive truck volumes.

    Parameters
    ----------
    scenario_path : Path
        Root of a scenario's output directory. The network is read from
        ``hwy/iter1/avgload5period_links.shp``.

    Returns
    -------
    gpd.GeoDataFrame
        The network with at least ``link_id`` (``"A-B"``), ``vol_HV``,
        ``vol_SM``, ``DISTANCE``, and ``geometry``. ``vol_HV`` sums the
        non-tolled and tolled heavy-truck 24-hour volumes; ``vol_SM`` does the
        same for small/medium trucks.
    """
    gdf = gpd.read_file(Path(scenario_path) / "hwy" / "iter1" / "avgload5period_links.shp")
    gdf["link_id"] = gdf["A"].astype(str) + "-" + gdf["B"].astype(str)
    gdf["vol_HV"] = gdf["VOL24HR_HV"] + gdf["VOL24HR_HVT"]
    gdf["vol_SM"] = gdf["VOL24HR_SM"] + gdf["VOL24HR_SMT"]
    return gdf


def build_trip_gen_table(completed_scenarios: list[dict]) -> pd.DataFrame:
    """
    Build a trip generation summary table across all completed scenarios.

    Reads ``nonres/TruckTG.dat`` from each scenario's output directory and
    aggregates total production trips by truck type.

    Parameters
    ----------
    completed_scenarios : list of dict
        Scenario configuration dicts, each containing at least ``"name"`` and
        ``"path"`` keys. Scenarios that failed are already excluded.

    Returns
    -------
    pd.DataFrame
        DataFrame with truck types as the index
        (``["Very Small", "Small", "Medium", "Large"]``) and one column per
        scenario name containing total production trips. Columns are NaN for any
        scenario whose ``TruckTG.dat`` could not be read.

    Notes
    -----
    Attractions are not included. The model balances attractions to productions,
    so they are nearly identical and the table would be redundant.
    """
    index = ["Very Small", "Small", "Medium", "Large"]
    data: dict[str, dict] = {}
    for scenario in completed_scenarios:
        name = scenario["name"]
        try:
            totals = read_truck_tg(Path(scenario["path"]))
            data[name] = totals
        except Exception:
            logger.warning("Could not read TruckTG.dat for scenario %s — filling NaN", name)
            data[name] = {k: np.nan for k in index}

    return pd.DataFrame(data, index=index)


def build_vmt_table(completed_scenarios: list[dict]) -> pd.DataFrame:
    """
    Build a vehicle-miles-travelled summary table over the full network.

    Parameters
    ----------
    completed_scenarios : list of dict
        Scenario configuration dicts with ``"name"`` and ``"path"`` keys.

    Returns
    -------
    pd.DataFrame
        DataFrame indexed by truck type (``["HV", "SM"]``) with one column per
        scenario name. Each cell is total VMT (``volume * DISTANCE`` summed over
        all links). Columns are NaN for any scenario whose network could not be
        read.
    """
    return _vmt_table(completed_scenarios, observed_link_ids=None)


def build_vmt_table_observed_links_only(
    completed_scenarios: list[dict], observed_link_ids: set[str]
) -> pd.DataFrame:
    """
    Build a VMT table restricted to links that have an observed count.

    Parameters
    ----------
    completed_scenarios : list of dict
        Scenario configuration dicts with ``"name"`` and ``"path"`` keys.
    observed_link_ids : set of str
        The ``link_id`` values present in the observed dataset.

    Returns
    -------
    pd.DataFrame
        Same shape as :func:`build_vmt_table` (index ``["HV", "SM"]``, one column
        per scenario), but VMT is summed only over links whose ``link_id`` is in
        ``observed_link_ids``. Used for the apples-to-apples VMT comparison
        chart, where the simulated side must cover the same links as the
        observed side.
    """
    return _vmt_table(completed_scenarios, observed_link_ids=observed_link_ids)


def _vmt_table(
    completed_scenarios: list[dict], observed_link_ids: set[str] | None
) -> pd.DataFrame:
    """Shared VMT aggregation, optionally restricted to ``observed_link_ids``."""
    index = ["HV", "SM"]
    data: dict[str, dict] = {}
    for scenario in completed_scenarios:
        name = scenario["name"]
        try:
            gdf = read_network(Path(scenario["path"]))
            if observed_link_ids is not None:
                gdf = gdf[gdf["link_id"].isin(observed_link_ids)]
            data[name] = {
                "HV": float((gdf["vol_HV"] * gdf["DISTANCE"]).sum()),
                "SM": float((gdf["vol_SM"] * gdf["DISTANCE"]).sum()),
            }
        except Exception:
            logger.warning("Could not read network for scenario %s — filling NaN", name)
            data[name] = {k: np.nan for k in index}

    return pd.DataFrame(data, index=index)


def clean_output(df, link_col="link_id"):
    """
    Enrich a long-format dataframe by propagating geometry and count metadata.

    Propagates, across all rows sharing a ``link_id``:
      - geometry (and ``ROUTENUM`` / ``ROUTEDIR`` / ``DISTANCE``)
      - ``count_location_id``

    Parameters
    ----------
    df : pandas.DataFrame or GeoDataFrame
        Input dataframe containing both observed and simulated rows.
    link_col : str
        Column used to join (default: ``"link_id"``).

    Returns
    -------
    pandas.DataFrame or GeoDataFrame
        Enriched dataframe with geometry and ``count_location_id`` filled, the
        ``vmt`` column added, and columns renamed to the 10-character shapefile
        limit.
    """
    out = df.copy()

    geom_lookup = (
        out[[link_col, "geometry", "ROUTENUM", "ROUTEDIR", "DISTANCE"]]
        .dropna(subset=["geometry"])
        .drop_duplicates(subset=[link_col])
    )

    count_lookup = (
        out[[link_col, "count_location_id"]]
        .dropna(subset=["count_location_id"])
        .drop_duplicates(subset=[link_col])
    )

    out = out.merge(geom_lookup, on=link_col, how="left", suffixes=("", "_geom"))
    out = out.merge(count_lookup, on=link_col, how="left", suffixes=("", "_count"))

    if "geometry_geom" in out.columns:
        out["geometry"] = out["geometry"].combine_first(out["geometry_geom"])
    if "ROUTEDIR_geom" in out.columns:
        out["ROUTEDIR"] = out["ROUTEDIR"].combine_first(out["ROUTEDIR_geom"])
    if "ROUTENUM_geom" in out.columns:
        out["ROUTENUM"] = out["ROUTENUM"].combine_first(out["ROUTENUM_geom"])
    if "DISTANCE_geom" in out.columns:
        out["DISTANCE"] = out["DISTANCE"].combine_first(out["DISTANCE_geom"])
    if "count_location_id_count" in out.columns:
        out["count_location_id"] = out["count_location_id"].combine_first(
            out["count_location_id_count"]
        )

    drop_cols = [
        "geometry_geom",
        "ROUTEDIR_geom",
        "ROUTENUM_geom",
        "DISTANCE_geom",
        "count_location_id_count",
    ]
    cols_to_drop = [c for c in drop_cols if c in out.columns]
    out = out.drop(columns=cols_to_drop)

    out["vmt"] = out["volume"] * out["DISTANCE"]

    out_cols = [
        "count_location_id",
        "link_id", "tod",
        "ROUTENUM",
        "ROUTEDIR",
        "DISTANCE",
        "truck_type_1",
        "truck_type_2",
        "volume",
        "vmt",
        "type",
        "source",
        "geometry",
    ]
    out = out[out_cols]

    rename_dict = {
        "count_location_id": "cnt_loc_id",
        "truck_type_1": "trk_typ_1",
        "truck_type_2": "trk_typ_2",
        "ROUTENUM": "route",
        "ROUTEDIR": "direction",
        "DISTANCE": "distance",
    }
    return out.rename(columns=rename_dict)


def summarize_predicted_counts(model_cfg: dict, cfg: dict) -> gpd.GeoDataFrame:
    """
    Melt a scenario's loaded network into long-format simulated truck counts.

    Parameters
    ----------
    model_cfg : dict
        Scenario dict with ``"name"`` and ``"path"`` keys.
    cfg : dict
        The full evaluation config (unused here but kept for signature
        compatibility with the original implementation).

    Returns
    -------
    gpd.GeoDataFrame
        Long-format rows with one record per link, time period, and truck type,
        tagged with ``type="simulated"`` and ``source=<scenario name>``.
    """
    scenario_name = model_cfg["name"]
    scenario_path = Path(model_cfg["path"])

    loaded_network = gpd.read_file(scenario_path / "hwy/avgload5period/avgload5period_links.shp")

    loaded_network["link_id"] = (
        loaded_network["A"].astype(str) + "-" + loaded_network["B"].astype(str)
    )

    tods = ["EA", "AM", "MD", "PM", "EV"]
    truck_types = {"HV": ("HV", "HVT"), "SM": ("SM", "SMT")}

    cols = []
    for tod in tods:
        for truck_type, (notoll, toll) in truck_types.items():
            name = f"VOL_{tod}_{truck_type}"
            loaded_network[name] = loaded_network[
                [f"VOL{tod}_{notoll}", f"VOL{tod}_{toll}"]
            ].sum(axis=1)
            cols.append(name)

    df_long = loaded_network.melt(
        id_vars=["link_id", "ROUTENUM", "ROUTEDIR", "DISTANCE", "geometry"],
        value_vars=cols,
        var_name="var",
        value_name="volume",
    )

    df_long["tod"] = df_long["var"].str[4:6]
    df_long["truck_type_2"] = df_long["var"].str[-2:]
    df_long["type"] = "simulated"
    df_long["source"] = scenario_name
    return df_long


def save_tableau_shapefile(
    completed_scenarios: list[dict], observed: pd.DataFrame, cfg: dict
) -> None:
    """
    Write the Tableau-ready validation shapefile (observed + simulated counts).

    Parameters
    ----------
    completed_scenarios : list of dict
        Scenario dicts with ``"name"`` and ``"path"`` keys.
    observed : pd.DataFrame
        Observed truck counts (must carry the ``geometry`` column, as WKT or
        shapely geometries).
    cfg : dict
        Config dict; ``cfg["network_crs"]`` is the network's projected CRS and
        ``cfg["evaluation_output"]`` is the output folder.

    Notes
    -----
    Preserves the behaviour of the original ``run_evaluation.py``: simulated
    rows from every scenario are concatenated with the observed rows, enriched
    by :func:`clean_output`, set to ``network_crs``, then written as
    ``validation_table.shp`` reprojected to ``EPSG:4326``.
    """
    obs = observed.copy()
    # Materialise WKT geometry strings into shapely objects if needed.
    if "geometry" in obs.columns and obs["geometry"].dtype == object:
        try:
            obs["geometry"] = gpd.GeoSeries.from_wkt(obs["geometry"])
        except Exception:
            logger.warning("Could not parse observed geometry as WKT; using as-is")

    summaries = [obs]
    for scenario in completed_scenarios:
        logger.info("  → simulating counts for: %s", scenario["name"])
        summaries.append(summarize_predicted_counts(scenario, cfg))

    out = pd.concat(summaries, axis=0)
    out = clean_output(out)
    out = gpd.GeoDataFrame(out, geometry="geometry")
    out = out.set_crs(cfg["network_crs"])

    out_path = Path(cfg["evaluation_output"]) / "validation_table.shp"
    save(out, out_path, crs="EPSG:4326")
    logger.info("Wrote Tableau shapefile: %s", out_path)


# --------------------------------------------------------------------------- #
# Colors and PNG export
# --------------------------------------------------------------------------- #
def assign_colors(completed_scenarios: list[dict]) -> dict[str, str]:
    """
    Map each scenario name to a hex color, cycling through the palette.

    Parameters
    ----------
    completed_scenarios : list of dict
        Scenario dicts with a ``"name"`` key.

    Returns
    -------
    dict of str to str
        Mapping from scenario name to a hex color string. This is the single
        place the palette is applied; plot functions and the Excel writer take
        the mapping as an argument.
    """
    return {
        s["name"]: PALETTE[i % len(PALETTE)]
        for i, s in enumerate(completed_scenarios)
    }


def save_pngs(
    scatter_figures: dict[tuple[str, str], Figure],
    vmt_figures: dict[str, Figure],
    plots_dir: Path,
) -> None:
    """
    Save all figures to ``plots_dir`` as 300-dpi PNGs, then close them.

    Parameters
    ----------
    scatter_figures : dict of (str, str) to Figure
        Keyed by ``(scenario_name, truck_type)``.
    vmt_figures : dict of str to Figure
        Keyed by truck type.
    plots_dir : Path
        Destination directory (already created by the caller).
    """
    for (scenario_name, truck_type), fig in scatter_figures.items():
        fname = plots_dir / f"scatter_{scenario_name}_{truck_type}.png"
        fig.savefig(fname, dpi=300, bbox_inches="tight")
        logger.info("Saved PNG: %s", fname)

    for truck_type, fig in vmt_figures.items():
        fname = plots_dir / f"vmt_comparison_{truck_type}.png"
        fig.savefig(fname, dpi=300, bbox_inches="tight")
        logger.info("Saved PNG: %s", fname)

    for fig in list(scatter_figures.values()) + list(vmt_figures.values()):
        plt.close(fig)


# --------------------------------------------------------------------------- #
# Excel workbook
# --------------------------------------------------------------------------- #
def write_excel(
    cfg: dict,
    completed_scenarios: list[dict],
    scenario_color_map: dict[str, str],
    trip_gen_table: pd.DataFrame,
    vmt_table: pd.DataFrame,
    scatter_figures: dict[tuple[str, str], Figure],
    vmt_figures: dict[str, Figure],
    output_dir: Path,
) -> None:
    """
    Write the comparison workbook with summary tables and embedded figures.

    Parameters
    ----------
    cfg : dict
        Full config dict (used for the Context sheet metadata).
    completed_scenarios : list of dict
        Scenario dicts with ``"name"`` and ``"path"`` keys.
    scenario_color_map : dict of str to str
        Scenario name to hex color.
    trip_gen_table : pd.DataFrame
        Output of :func:`build_trip_gen_table`.
    vmt_table : pd.DataFrame
        Output of :func:`build_vmt_table`.
    scatter_figures : dict of (str, str) to Figure
        Output of ``plot_scatter_all_scenarios``.
    vmt_figures : dict of str to Figure
        Output of ``plot_vmt_comparison``.
    output_dir : Path
        Destination directory; the workbook is written to
        ``truck_model_evaluation.xlsx`` inside it.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    scenario_names = [s["name"] for s in completed_scenarios]

    _write_context_sheet(wb, cfg, completed_scenarios)
    _write_table_sheet(
        wb, "Trip Generation", trip_gen_table, scenario_names, scenario_color_map,
        totals_label="Total productions",
    )
    _write_table_sheet(
        wb, "VMT by Type", vmt_table, scenario_names, scenario_color_map,
        totals_label="Total",
    )

    for (scenario_name, truck_type), fig in scatter_figures.items():
        _write_scatter_sheet(wb, scenario_name, truck_type, fig, scenario_color_map)

    for truck_type in TRUCK_TYPES:
        fig = vmt_figures.get(truck_type)
        if fig is not None:
            _write_vmt_sheet(wb, truck_type, fig, scenario_color_map)

    out_path = Path(output_dir) / "truck_model_evaluation.xlsx"
    wb.save(out_path)
    logger.info("Wrote Excel workbook: %s", out_path)


def _tint(hex_color: str, factor: float = 0.2) -> str:
    """Mix ``hex_color`` with white (``factor`` = share of color), return 6-hex."""
    hex_color = hex_color.lstrip("#")
    r, g, b = (int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    r = round(255 * (1 - factor) + r * factor)
    g = round(255 * (1 - factor) + g * factor)
    b = round(255 * (1 - factor) + b * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def _fill(hex_color: str) -> PatternFill:
    """Solid fill from a hex color string (with or without leading ``#``)."""
    return PatternFill(
        start_color=hex_color.lstrip("#"),
        end_color=hex_color.lstrip("#"),
        fill_type="solid",
    )


def _style_header(cell, fill_hex: str, font_color: str = "FFFFFF") -> None:
    """Apply the standard bold colored header style to a cell."""
    cell.fill = _fill(fill_hex)
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _data_font() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE)


def _autosize(ws: Worksheet) -> None:
    """Set each column's width to fit its longest cell value."""
    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            letter = cell.column_letter
            widths[letter] = max(widths.get(letter, 0), len(str(cell.value)))
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)


def _embed_figure(ws: Worksheet, fig: Figure, anchor: str, width_px: int, height_px: int) -> None:
    """Embed a matplotlib figure into ``ws`` at ``anchor`` as a sized PNG."""
    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=300, bbox_inches="tight")
    buffer.seek(0)
    img = XLImage(buffer)
    img.width = width_px
    img.height = height_px
    ws.add_image(img, anchor)


def _write_context_sheet(wb: Workbook, cfg: dict, completed_scenarios: list[dict]) -> None:
    """Write the metadata / scenario-listing Context sheet."""
    ws = wb.create_sheet("Context")
    ws.sheet_view.showGridLines = False

    header_fill = _fill(HEADER_GRAY)
    header_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")

    # Header row.
    for col, label in enumerate(["Field", "Value"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font

    rows = [
        ("Generated on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Config file", "configs/travel_model_scenarios.yaml"),
        ("Observed data", str(cfg.get("observed_data", ""))),
        ("Output folder", str(cfg.get("evaluation_output", ""))),
        ("Scenarios run", ""),
    ]
    for s in completed_scenarios:
        rows.append((s["name"], s["path"]))

    for i, (field, value) in enumerate(rows, start=2):
        fill = _fill(ALT_ROW_GRAY) if i % 2 == 0 else None
        for col, text in enumerate((field, value), start=1):
            cell = ws.cell(row=i, column=col, value=text)
            cell.font = _data_font()
            if fill is not None:
                cell.fill = fill

    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_table_sheet(
    wb: Workbook,
    sheet_name: str,
    table: pd.DataFrame,
    scenario_names: list[str],
    scenario_color_map: dict[str, str],
    totals_label: str,
) -> None:
    """Write a summary table sheet (Trip Generation / VMT by Type)."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    present = [name for name in scenario_names if name in table.columns]

    # --- Header row ---
    ws.cell(row=1, column=1, value="Truck Type").font = Font(
        name=FONT_NAME, size=FONT_SIZE, bold=True
    )
    col = 2
    scenario_cols: dict[str, int] = {}
    for name in present:
        cell = ws.cell(row=1, column=col, value=name)
        _style_header(cell, scenario_color_map.get(name, "#4E79A7"))
        scenario_cols[name] = col
        col += 1

    # --- % diff columns vs the first scenario ---
    diff_cols: dict[str, int] = {}
    if len(present) >= 2:
        base = present[0]
        for name in present[1:]:
            cell = ws.cell(row=1, column=col, value=f"% diff vs {base}")
            _style_header(cell, _tint(scenario_color_map.get(name, "#4E79A7")), font_color="000000")
            diff_cols[name] = col
            col += 1

    # --- Data rows ---
    for r, truck_type in enumerate(table.index, start=2):
        label_cell = ws.cell(row=r, column=1, value=str(truck_type))
        label_cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        for name in present:
            value = table.loc[truck_type, name]
            cell = ws.cell(row=r, column=scenario_cols[name], value=_num(value))
            cell.font = _data_font()
            cell.number_format = INT_FORMAT
        if len(present) >= 2:
            base = present[0]
            base_val = table.loc[truck_type, base]
            for name in present[1:]:
                cell = ws.cell(row=r, column=diff_cols[name])
                cell.font = _data_font()
                cell.fill = _fill(_tint(scenario_color_map.get(name, "#4E79A7")))
                cell.number_format = PCT_FORMAT
                val = table.loc[truck_type, name]
                if pd.notna(base_val) and base_val != 0 and pd.notna(val):
                    cell.value = (val - base_val) / base_val

    # --- Totals row ---
    total_row = len(table.index) + 2
    total_cell = ws.cell(row=total_row, column=1, value=totals_label)
    total_cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    for name in present:
        total = table[name].sum(skipna=True)
        cell = ws.cell(row=total_row, column=scenario_cols[name], value=_num(total))
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        cell.number_format = INT_FORMAT

    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_scatter_sheet(
    wb: Workbook,
    scenario_name: str,
    truck_type: str,
    fig: Figure,
    scenario_color_map: dict[str, str],
) -> None:
    """Write a scatter sheet: embedded figure plus its fit-statistics table."""
    sheet_name = _safe_sheet_name(f"Scatter {scenario_name} {truck_type}")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    color = scenario_color_map.get(scenario_name, "#4E79A7")
    ws.sheet_properties.tabColor = color.lstrip("#")

    _embed_figure(ws, fig, "A1", width_px=500, height_px=430)

    stats = getattr(fig, "scenario_stats", {}) or {}
    start_row = 33
    for col, label in enumerate(["Metric", "Value"], start=1):
        cell = ws.cell(row=start_row, column=col, value=label)
        _style_header(cell, color)

    rows = [
        ("Slope", _fmt_stat(stats.get("slope"))),
        ("Intercept", _fmt_stat(stats.get("intercept"))),
        ("R²", _fmt_stat(stats.get("r2"))),
        ("Count locations (n)", stats.get("n")),
    ]
    for i, (label, value) in enumerate(rows, start=start_row + 1):
        ws.cell(row=i, column=1, value=label).font = _data_font()
        ws.cell(row=i, column=2, value=value).font = _data_font()


def _write_vmt_sheet(
    wb: Workbook,
    truck_type: str,
    fig: Figure,
    scenario_color_map: dict[str, str],
) -> None:
    """Write a per-truck-type VMT sheet: embedded figure plus the raw numbers.

    The numbers table is read from ``fig.vmt_values`` (``{category: VMT}``) so it
    matches the embedded chart exactly — ``Observed`` first, then each scenario.
    """
    ws = wb.create_sheet(f"VMT - {truck_type}")
    ws.sheet_view.showGridLines = False

    _embed_figure(ws, fig, "A1", width_px=430, height_px=360)

    start_row = 28
    for col, label in enumerate(["Category", "VMT"], start=1):
        cell = ws.cell(row=start_row, column=col, value=label)
        _style_header(cell, HEADER_GRAY)

    values: dict[str, float] = getattr(fig, "vmt_values", {}) or {}
    for i, (category, value) in enumerate(values.items(), start=start_row + 1):
        name_cell = ws.cell(row=i, column=1, value=category)
        name_cell.font = _data_font()
        if category != "Observed":
            name_cell.fill = _fill(_tint(scenario_color_map.get(category, "#4E79A7")))
        value_cell = ws.cell(row=i, column=2, value=_num(value))
        value_cell.font = _data_font()
        value_cell.number_format = INT_FORMAT


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def _num(value) -> float | int | None:
    """Return a numeric value for Excel, or ``None`` for NaN/missing."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    if pd.isna(value):
        return None
    return value


def _fmt_stat(value) -> float | None:
    """Round a fit statistic for display, or ``None`` if missing."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    return round(float(value), 4)


def _safe_sheet_name(name: str) -> str:
    """Truncate to Excel's 31-character sheet-name limit."""
    return name[:31]
