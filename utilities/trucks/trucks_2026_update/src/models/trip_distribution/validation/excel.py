"""Excel report renderer for the truck trip distribution model.

All openpyxl code lives here. No computation, no matplotlib.
Reads from a fully populated ``ReportData`` object and writes
``summary.xlsx`` to the output directory.

Workbook structure
------------------
Index       — Workbook map, one row per tab, colour-coded by run group
Summary     — One row per run: calibration parameters, ATL comparison,
               convergence flags.  OD columns added when any run has OD data.

Per-run tab group (one colour per run, cycling through 6 palette entries):
    {sn} · PA Stats   — MAE, RMSE, max error, % within tolerance
    {sn} · TLFD       — Observed vs modelled share per TLFD bin
    {sn} · Zones      — Per-zone P/A residuals with rank columns
    {sn} · {geo_col}  — One tab per ``geo_agg_col`` (when configured)
    {sn} · OD Stats   — OD summary statistics (when target_od_path is set)

The first three tabs are created for every run (including FAILED ones).
Geo and OD tabs are created only when data are available.
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .report import ReportData, RunReport


# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------

TAB_COLORS = [
    "BDD7EE",  # sky blue
    "C6EFCE",  # light green
    "FCE4D6",  # peach
    "E2EFDA",  # sage
    "FFEB9C",  # yellow
    "FFC7CE",  # rose
]

_HEADER_GRAY  = "D9D9D9"   # neutral header fill for Index and Summary
_FAILED_FILL  = "FFD5D5"   # light red — FAILED rows in Summary
_WHITE        = "FFFFFF"


# ---------------------------------------------------------------------------
# Low-level style helpers
# ---------------------------------------------------------------------------

def _fill(hex6: str) -> PatternFill:
    """Solid-fill from a 6-char hex colour string (no leading #)."""
    return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")


def _bold() -> Font:
    return Font(bold=True)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, n_cols: int, fill: PatternFill) -> None:
    """Bold + fill the first row of *ws* across *n_cols* columns."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = _bold()
        cell.alignment = _center()


def _style_row(ws, row_idx: int, n_cols: int, fill: PatternFill) -> None:
    """Apply *fill* to every cell in *row_idx* (1-based) across *n_cols*."""
    for c in range(1, n_cols + 1):
        ws.cell(row=row_idx, column=c).fill = fill


def _auto_width(ws) -> None:
    """Set each column width to fit its longest cell value (10–50 chars)."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value or ""))
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(10, min(50, max_len + 3))


# ---------------------------------------------------------------------------
# Cell-value sanitisation
# ---------------------------------------------------------------------------

def _safe(v: Any) -> Any:
    """Convert a value to a Python scalar that openpyxl can write.

    - None and all NA variants (NaN, pd.NA, pd.NaT) → None
    - numpy scalars → Python int / float / bool via ``.item()``
    - Infinite floats → None (Excel has no infinity representation)
    - Everything else returned as-is
    """
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, float) and not math.isfinite(v):
        return None
    if hasattr(v, "item"):          # numpy scalar
        try:
            return v.item()
        except Exception:
            return str(v)
    return v


# ---------------------------------------------------------------------------
# DataFrame writer
# ---------------------------------------------------------------------------

def _write_df(
    ws,
    df: pd.DataFrame | None,
    header_fill: PatternFill,
    freeze: bool = True,
) -> None:
    """Write *df* to *ws* with a styled header row.

    If *df* is None or empty, a single ``"No data available"`` cell is written
    instead.  *freeze* adds ``freeze_panes`` on A2 (useful for tall tables).
    """
    if df is None or df.empty:
        ws.cell(row=1, column=1, value="No data available")
        return

    # Header
    headers = list(df.columns)
    ws.append(headers)
    _style_header(ws, len(headers), header_fill)

    # Data rows
    for tup in df.itertuples(index=False):
        ws.append([_safe(v) for v in tup])

    if freeze:
        ws.freeze_panes = "A2"

    _auto_width(ws)


# ---------------------------------------------------------------------------
# ATL helper (used only in Summary)
# ---------------------------------------------------------------------------

def _compute_atl(
    tlfd_table: pd.DataFrame,
) -> tuple[float, float, float | None]:
    """Return (atl_observed, atl_modeled, atl_error_pct) from a TLFD table.

    All values are rounded to 2 decimal places.  ``atl_error_pct`` is None
    when the observed ATL is zero.
    """
    midpts = (
        (tlfd_table["bin_start"] + tlfd_table["bin_end"]) / 2.0
    ).to_numpy(dtype=np.float64)
    obs_s = tlfd_table["observed_share"].to_numpy(dtype=np.float64)
    mod_s = tlfd_table["modeled_share"].to_numpy(dtype=np.float64)

    atl_obs = float(round((midpts * obs_s).sum(), 2))
    atl_mod = float(round((midpts * mod_s).sum(), 2))
    atl_err: float | None = (
        float(round((atl_mod - atl_obs) / atl_obs * 100.0, 2))
        if atl_obs > 0.0
        else None
    )
    return atl_obs, atl_mod, atl_err


# ---------------------------------------------------------------------------
# Tab writers
# ---------------------------------------------------------------------------

def _write_index(
    ws,
    tab_registry: list[tuple[str, str, str, str]],
) -> None:
    """Populate the Index tab.

    *tab_registry* is a list of ``(tab_name, description, run_name, hex_color)``
    tuples in workbook tab order.  Rows for run-specific tabs are filled with
    the run's palette colour so the user can quickly locate each group.
    """
    headers = ["Tab Name", "Description", "Run"]
    ws.append(headers)
    _style_header(ws, len(headers), _fill(_HEADER_GRAY))

    for tab_name, description, run_name, color in tab_registry:
        ws.append([tab_name, description, run_name])
        if color:
            _style_row(ws, ws.max_row, len(headers), _fill(color))

    _auto_width(ws)


def _write_summary(ws, report: ReportData) -> None:
    """Populate the Summary tab with one row per run."""
    always_cols: list[str] = [
        "status", "run_name", "short_name",
        "b_initial", "c_initial",
        "b_final", "c_final",
        "atl_observed", "atl_modeled", "atl_error_pct",
        "converged", "n_iters", "final_loss",
        "error_message",
    ]
    od_cols: list[str] = (
        ["r2_log", "slope_log", "mean_abs_residual_pct"]
        if report.has_od
        else []
    )
    all_cols = always_cols + od_cols
    n_cols = len(all_cols)

    ws.append(all_cols)
    _style_header(ws, n_cols, _fill(_HEADER_GRAY))

    for run in report.runs:
        row: list[Any]

        if run.status == "OK" and run.tlfd_table is not None:
            atl_obs, atl_mod, atl_err = _compute_atl(run.tlfd_table)
            row = [
                run.status, run.run_name, run.short_name,
                run.b_initial, run.c_initial,
                run.b_final, run.c_final,
                atl_obs, atl_mod, atl_err,
                run.converged, run.n_iters,
                _safe(run.final_loss),
                None,                         # error_message blank for OK
            ]
            if report.has_od:
                if run.od_stats is not None:
                    mask = run.od_stats["metric"] == "mean absolute residual %"
                    vals = run.od_stats.loc[mask, "value"].values
                    mean_resid = float(round(vals[0], 2)) if len(vals) else None
                    row += [
                        _safe(run.r2_log),
                        _safe(run.slope_log),
                        mean_resid,
                    ]
                else:
                    row += [None, None, None]

        else:
            # FAILED — spec: all columns after short_name blank, except
            # error_message which is populated.
            row = [run.status, run.run_name, run.short_name]
            row += [None] * (n_cols - 3)
            row[all_cols.index("error_message")] = run.error_message

        ws.append(row)

        if run.status == "FAILED":
            _style_row(ws, ws.max_row, n_cols, _fill(_FAILED_FILL))

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def render_excel(report: ReportData, output_dir: Path) -> None:
    """Write the validation report to ``output_dir/summary.xlsx``.

    Creates the workbook, populates all tabs in the correct order, and
    saves to disk.  Reading from ``report`` is the only I/O; no matplotlib,
    no computation beyond display-level arithmetic (ATL).

    Tab creation order
    ------------------
    1. Index  (filled after all other tabs are registered)
    2. Summary
    3. For each run (in config order):
       a. {sn} · PA Stats   — always
       b. {sn} · TLFD       — always
       c. {sn} · Zones      — always
       d. {sn} · {geo_col}  — one per geo column, only when pa_geo is present
       e. {sn} · OD Stats   — only when od_stats is present

    Parameters
    ----------
    report : ReportData
        Fully populated report data object from ``build_report_data``.
    output_dir : Path
        Root output directory.  ``summary.xlsx`` is written directly here.

    Raises
    ------
    OSError
        If ``output_dir`` does not exist or is not writable.
    """
    wb = Workbook()
    wb.remove(wb.active)                    # drop the auto-created blank sheet

    # Index and Summary are created first so they appear as the leftmost tabs.
    # They will be populated AFTER the run tabs so the tab_registry is complete.
    ws_index   = wb.create_sheet("Index")
    ws_summary = wb.create_sheet("Summary")

    # tab_registry: (tab_name, description, run_name, hex_color)
    # Color is empty string for Index/Summary (neutral, no coloured fill in Index)
    tab_registry: list[tuple[str, str, str, str]] = [
        ("Index",   "Workbook map — all tabs with short descriptions", "—", ""),
        ("Summary", "One row per run — calibration results, ATL, convergence flags", "—", ""),
    ]

    for run_idx, run in enumerate(report.runs):
        color     = TAB_COLORS[run_idx % len(TAB_COLORS)]
        tab_color = "FF" + color     # opaque ARGB — Excel ignores alpha=00
        sn        = run.short_name
        hfill     = _fill(color)

        # ── PA Stats ──────────────────────────────────────────────────────
        tab = f"{sn} · PA Stats"
        ws  = wb.create_sheet(tab)
        ws.sheet_properties.tabColor = tab_color
        tab_registry.append((
            tab,
            "Balance statistics: MAE, RMSE, max error, % zones within tolerance",
            run.run_name, color,
        ))
        _write_df(ws, run.pa_stats, hfill)

        # ── TLFD ──────────────────────────────────────────────────────────
        tab = f"{sn} · TLFD"
        ws  = wb.create_sheet(tab)
        ws.sheet_properties.tabColor = tab_color
        tab_registry.append((
            tab,
            "Trip length frequency distribution — observed vs modelled shares per bin",
            run.run_name, color,
        ))
        _write_df(ws, run.tlfd_table, hfill)

        # ── Zones ─────────────────────────────────────────────────────────
        tab = f"{sn} · Zones"
        ws  = wb.create_sheet(tab)
        ws.sheet_properties.tabColor = tab_color
        tab_registry.append((
            tab,
            "Per-zone P/A residuals — sorted by attraction error (rank 1 = largest error)",
            run.run_name, color,
        ))
        _write_df(ws, run.pa_zones, hfill)

        # ── Geo tabs (one per geo_agg_col, only when available) ───────────
        for geo_col, gdf in run.pa_geo.items():
            tab = f"{sn} · {geo_col}"
            ws  = wb.create_sheet(tab)
            ws.sheet_properties.tabColor = tab_color
            tab_registry.append((
                tab,
                f"P/A aggregated by {geo_col} — target vs modelled with % error",
                run.run_name, color,
            ))
            _write_df(ws, gdf, hfill)

        # ── OD Stats (only when OD data are present) ──────────────────────
        if run.od_stats is not None:
            tab = f"{sn} · OD Stats"
            ws  = wb.create_sheet(tab)
            ws.sheet_properties.tabColor = tab_color
            tab_registry.append((
                tab,
                "OD comparison summary — log-log R², slope, residual percentages",
                run.run_name, color,
            ))
            _write_df(ws, run.od_stats, hfill, freeze=False)

    # Populate the pre-created Index and Summary tabs now that tab_registry
    # is fully assembled.
    _write_index(ws_index, tab_registry)
    _write_summary(ws_summary, report)

    out_path = output_dir / "summary.xlsx"
    wb.save(out_path)
