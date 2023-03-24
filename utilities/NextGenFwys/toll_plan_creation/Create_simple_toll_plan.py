# This script converts dynamic toll results to a simple toll plan
# The main output is a new tolls.csv
# 
# -------------------------------------------------------------------------------------
# how to use this script?
# since users may want to vary some of the user settings
# the following steps are recommended
#
# 1. save a local copy in a working directory
# e.g. L:\Application\Model_One\NextGenFwys\INPUT_DEVELOPMENT\Static_toll_plans\Test
# 
# 2. review the user settings section and edit as necessary
#
# 3. run the command "python Create_simple_toll_plan.py" from the working directory 
#
# -------------------------------------------------------------------------------------


import os
import pandas as pd
import geopandas as gpd
import numpy as np

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# for copying files
import shutil

#-------------------
# user settings
#-------------------
# may add a config file to store these configs, but not a high priority
# toll rates in 2023$
high_toll_cents      = 50
medhigh_toll_cents   = 40
medium_toll_cents    = 30
medlow_toll_cents    = 20
low_toll_cents       = 10

# cutoff points in 2023$
high_cutoff    = 43
medhigh_cutoff = 33 
medium_cutoff  = 23
medlow_cutoff  = 13
low_cutoff     = 3

# Inflation assumptions 
# Follow this table here: https://github.com/BayAreaMetro/modeling-website/wiki/InflationAssumptions. A
# According to this table + an assumption of 1.03 for 2023, the conversion factor from 2023$ to 2000$ would be 1.81*1.03 = 1.86. 
CoversionFactorTo2000cents = 1.86

# our planning colleagues have done some work to group the links into 8 major groups and 19 minor groups, based on their geography and congestion levels
# please indicate whether you'd like to use the major groups or the minor groups in the averaging process
# by commenting out either "major" or "minor" in the below
# grouping_option = "major"
grouping_option = "minor"

# specify "true" below if there will be midday tolls
tolls_in_midday = "true"

# specify HOV discounts
# a DiscountFactor of 0.5 means half price; a DiscountFactor of 0 means free; and a DiscountFactor of 1 means no discount.
# in NGF, there is a carpool discounts of 50% for HOV3
DiscountFactor_HOV2 = 1.0
DiscountFactor_HOV3 = 0.5

# specify the relationship between arterial tolls and freeway tolls
# For example, if ArterialFactor = 0.2, this script will set arterial tolls to be 20% of freeway tolls
ArterialFactor = 0.2

# specify working directories and run ids
project_dir                  ="L:/Application/Model_One/NextGenFwys/" 
modelrun_with_DynamicTolling = "2035_TM152_NGF_NP07_Path1b_01_TollCalibrated01"
modelrun_with_NoProject      = "2035_TM152_NGF_NP07_TollCalibrated02"
output_dir                   = "INPUT_DEVELOPMENT/Static_toll_plans/Static_toll_P1b_V09round"


# ------------------
# Inputs
#-------------------

# input 1
# the toll plan of the dynamic toll run
tollcsv_prev_df = pd.read_csv(os.path.join(project_dir,"Scenarios", modelrun_with_DynamicTolling, "INPUT", "hwy", "tolls.csv")) 

# input 2
# the loaded network from a dynamic tolling run, in which the all-lane tolling system is represented as many short segments (100+)
loadednetwork_100psegs_shp_gdf = gpd.read_file(os.path.join(project_dir, "Scenarios", modelrun_with_DynamicTolling, "OUTPUT", "shapefile", "network_links.shp")) 

# input 3
# the file with the tollclass groupings
workbook = load_workbook(filename="X:/travel-model-one-master/utilities/NextGenFwys/TOLLCLASS_Designations.xlsx")
# make "By "Inputs_for_tollcalib" the active tab
sheet = workbook["Inputs_for_tollcalib"]
# Read the results in the "Inputs_for_tollcalib" tab
TollclassGrouping = sheet.values
# Set the first row as the headers for the DataFrame
cols = next(TollclassGrouping)
TollclassGrouping = list(TollclassGrouping)

TollclassGrouping_df = pd.DataFrame(TollclassGrouping, columns=cols)
# TollclassGrouping_df = pd.read_csv('X:/travel-model-one-master/utilities/NextGenFwys/TollclassGrouping.csv', index_col=False, sep=",")

#-----------------------------------------------------
# merge the tollclass grouping to the loaded network 
#-----------------------------------------------------
if grouping_option == "major":
    TollclassGrouping_df["Tollclass Group"]=TollclassGrouping_df["Grouping major"]
if grouping_option == "minor":
    TollclassGrouping_df["Tollclass Group"]=TollclassGrouping_df["Grouping minor"]

loadednetwork_100psegs_shp_gdf = pd.merge(loadednetwork_100psegs_shp_gdf,
                             TollclassGrouping_df,
                             how='left',
                             left_on=['TOLLCLASS'], 
                             right_on = ['tollclass'],
                             indicator=True)


# keep only the links that are part of the all-lane tolling system i.e. tollclass > 700,000
loadednetwork_100psegs_shp_gdf = loadednetwork_100psegs_shp_gdf.loc[loadednetwork_100psegs_shp_gdf['TOLLCLASS'] > 700000]




# ------------------
# Convert average toll rates to high, medium, low 
#-------------------
TimePeriods = ["EA", "AM", "MD", "PM", "EV"]
for tp in TimePeriods:
    # print("\n", tp)
    if tp=="EA":
        loadednetwork_100psegs_shp_gdf['USEtp']     = loadednetwork_100psegs_shp_gdf['USEEA']
        loadednetwork_100psegs_shp_gdf['TOLLtp_DA'] = loadednetwork_100psegs_shp_gdf['TOLLEA_DA']
    if tp=="AM":
        loadednetwork_100psegs_shp_gdf['USEtp']     = loadednetwork_100psegs_shp_gdf['USEAM']
        loadednetwork_100psegs_shp_gdf['TOLLtp_DA'] = loadednetwork_100psegs_shp_gdf['TOLLAM_DA']  
    if tp=="MD":
        loadednetwork_100psegs_shp_gdf['USEtp']     = loadednetwork_100psegs_shp_gdf['USEMD']
        loadednetwork_100psegs_shp_gdf['TOLLtp_DA'] = loadednetwork_100psegs_shp_gdf['TOLLMD_DA']  
    if tp=="PM":
        loadednetwork_100psegs_shp_gdf['USEtp']     = loadednetwork_100psegs_shp_gdf['USEPM']
        loadednetwork_100psegs_shp_gdf['TOLLtp_DA'] = loadednetwork_100psegs_shp_gdf['TOLLPM_DA']         
    if tp=="EV":
        loadednetwork_100psegs_shp_gdf['USEtp']     = loadednetwork_100psegs_shp_gdf['USEEV']
        loadednetwork_100psegs_shp_gdf['TOLLtp_DA'] = loadednetwork_100psegs_shp_gdf['TOLLEV_DA']  

    # keep only the variables that are needed
    streamlined_shp_df = loadednetwork_100psegs_shp_gdf[['Grouping major','Tollclass Group','USEtp','DISTANCE', 'TOLLtp_DA']] 

    # calculate average toll rate
    SimpleToll_tp_df = streamlined_shp_df.groupby(['Tollclass Group', 'USEtp'], as_index=False).sum()
    SimpleToll_tp_df['avg_toll_rate'] = SimpleToll_tp_df['TOLLtp_DA']/SimpleToll_tp_df['DISTANCE']

    # apply a simple rule to convert the results to high, medium, low toll

    # toll rates in 2000 prices and in $
    high_toll_dollars      = round(high_toll_cents      / CoversionFactorTo2000cents / 100, 4)
    medhigh_toll_dollars   = round(medhigh_toll_cents   / CoversionFactorTo2000cents / 100, 4)
    medium_toll_dollars    = round(medium_toll_cents    / CoversionFactorTo2000cents / 100, 4)
    medlow_toll_dollars    = round(medlow_toll_cents    / CoversionFactorTo2000cents / 100, 4)
    low_toll_dollars       = round(low_toll_cents       / CoversionFactorTo2000cents / 100, 4)

    # cutoff points in 2000 cents
    high_cutoff_2000cents    = high_cutoff    / CoversionFactorTo2000cents
    medhigh_cutoff_2000cents = medhigh_cutoff / CoversionFactorTo2000cents
    medium_cutoff_2000cents  = medium_cutoff  / CoversionFactorTo2000cents
    medlow_cutoff_2000cents  = medlow_cutoff  / CoversionFactorTo2000cents
    low_cutoff_2000cents     = low_cutoff     / CoversionFactorTo2000cents

    TollLevel_conditions = [
        (SimpleToll_tp_df['avg_toll_rate'] >  high_cutoff_2000cents),
        (SimpleToll_tp_df['avg_toll_rate'] >= medhigh_cutoff_2000cents),
        (SimpleToll_tp_df['avg_toll_rate'] >= medium_cutoff_2000cents),
        (SimpleToll_tp_df['avg_toll_rate'] >= medlow_cutoff_2000cents),
        (SimpleToll_tp_df['avg_toll_rate'] >= low_cutoff_2000cents),
        (SimpleToll_tp_df['avg_toll_rate'] >= 0)]
    TollLevel_choices = [high_toll_dollars, medhigh_toll_dollars, medium_toll_dollars, medlow_toll_dollars, low_toll_dollars, 0.0001] 
    # note 0.0001 is applied instead of 0 
    # because settolls.job has an input verification step which rejects the tolls.csv if it finds a tolled link with peak period DA tolls equal to 0 
    # as ((TOLLAM_DA == 0) && (TOLLPM_DA == 0)) could means that a TOLLCLASS and USE combination is missing from the tolls.csv
    # https://github.com/BayAreaMetro/travel-model-one/blob/master/model-files/scripts/preprocess/SetTolls.JOB#L215-L221
    # settolls.job includes codes that sets anything toll value less than 1 cent to 0
    # https://github.com/BayAreaMetro/travel-model-one/blob/master/model-files/scripts/preprocess/SetTolls.JOB#L224-L269


    SimpleToll_tp_df['simplified_toll'] = np.select(TollLevel_conditions, TollLevel_choices, default='null')

    # store the average toll rate and simplied toll rate for each time period
    if tp=="EA":
        SimpleToll_tp_df['avg_toll_rate_EA']       = SimpleToll_tp_df['avg_toll_rate']
        SimpleToll_tp_df['simplified_toll_EA']     = SimpleToll_tp_df['simplified_toll']
        SimpleToll_EA_df = SimpleToll_tp_df
        # drop variables that are not needed for subsequent processing
        SimpleToll_EA_df.drop(columns=['DISTANCE','TOLLtp_DA', 'avg_toll_rate','simplified_toll'], inplace=True)
        # temp output 
        #output_filename1 = os.path.join(project_dir, output_dir,"average_n_simplified_toll_rate_EA.csv")
        #SimpleToll_EA_df.to_csv(output_filename1, header=True, index=False)
    if tp=="AM":
        SimpleToll_tp_df['avg_toll_rate_AM']       = SimpleToll_tp_df['avg_toll_rate']
        SimpleToll_tp_df['simplified_toll_AM']     = SimpleToll_tp_df['simplified_toll']
        SimpleToll_AM_df = SimpleToll_tp_df
        # drop variables that are not needed for subsequent processing
        SimpleToll_AM_df.drop(columns=['DISTANCE','TOLLtp_DA', 'avg_toll_rate','simplified_toll'], inplace=True)
        # temp output 
        #output_filename1 = os.path.join(project_dir, output_dir,"average_n_simplified_toll_rate_AM.csv")
        #SimpleToll_AM_df.to_csv(output_filename1, header=True, index=False)
    if tp=="MD":
        SimpleToll_tp_df['avg_toll_rate_MD']       = SimpleToll_tp_df['avg_toll_rate']
        SimpleToll_tp_df['simplified_toll_MD']     = SimpleToll_tp_df['simplified_toll']
        SimpleToll_MD_df = SimpleToll_tp_df
        # drop variables that are not needed for subsequent processing
        SimpleToll_MD_df.drop(columns=['DISTANCE','TOLLtp_DA', 'avg_toll_rate','simplified_toll'], inplace=True)
    if tp=="PM":
        SimpleToll_tp_df['avg_toll_rate_PM']       = SimpleToll_tp_df['avg_toll_rate']
        SimpleToll_tp_df['simplified_toll_PM']     = SimpleToll_tp_df['simplified_toll']
        SimpleToll_PM_df = SimpleToll_tp_df
        # drop variables that are not needed for subsequent processing
        SimpleToll_PM_df.drop(columns=['DISTANCE','TOLLtp_DA', 'avg_toll_rate','simplified_toll'], inplace=True)     
    if tp=="EV":
        SimpleToll_tp_df['avg_toll_rate_EV']       = SimpleToll_tp_df['avg_toll_rate']
        SimpleToll_tp_df['simplified_toll_EV']     = SimpleToll_tp_df['simplified_toll']
        SimpleToll_EV_df = SimpleToll_tp_df
        # drop variables that are not needed for subsequent processing
        SimpleToll_EV_df.drop(columns=['DISTANCE','TOLLtp_DA', 'avg_toll_rate','simplified_toll'], inplace=True)

    # get a combination of the 5 data frames
    if tp=="EA":
        SimpleToll_5tp_df =  SimpleToll_EA_df
    if tp=="AM":
        SimpleToll_5tp_df = pd.merge(SimpleToll_5tp_df,
                              SimpleToll_AM_df,
                              how='outer',
                              left_on=['Tollclass Group', 'USEtp'], 
                              right_on = ['Tollclass Group', 'USEtp'],
                              indicator=False)
    if tp=="MD":
        SimpleToll_5tp_df = pd.merge(SimpleToll_5tp_df,
                              SimpleToll_MD_df,
                              how='outer',
                              left_on=['Tollclass Group', 'USEtp'], 
                              right_on = ['Tollclass Group', 'USEtp'],
                              indicator=False)
    if tp=="PM":
        SimpleToll_5tp_df = pd.merge(SimpleToll_5tp_df,
                              SimpleToll_PM_df,
                              how='outer',
                              left_on=['Tollclass Group', 'USEtp'], 
                              right_on = ['Tollclass Group', 'USEtp'],
                              indicator=False)
    if tp=="EV":
        SimpleToll_5tp_df = pd.merge(SimpleToll_5tp_df,
                              SimpleToll_EV_df,
                              how='outer',
                              left_on=['Tollclass Group', 'USEtp'], 
                              right_on = ['Tollclass Group', 'USEtp'],
                              indicator=False)

        # temp output 
        output_filename1 = os.path.join(project_dir, output_dir,"average_n_simplified_toll_rate_5tp.csv")
        SimpleToll_5tp_df.to_csv(output_filename1, header=True, index=False)

        # ---
        # generate an output that can be used by Generate_tollscsv_from_TollclassGrouping.py
        # ---
        # keep only if use=1 (in the NGF results we have seen so far, we never have to toll the HOV lanes parallel to the all-lane tolling system; toll levels are set based on the conditions on the use=1 lanes)
        SimpleToll_MinorGroupingX3TP_df = SimpleToll_5tp_df.loc[SimpleToll_5tp_df['USEtp']==1]
        # drop the row with Tollclass Group = #N/A
        SimpleToll_MinorGroupingX3TP_df = SimpleToll_MinorGroupingX3TP_df[SimpleToll_MinorGroupingX3TP_df['Tollclass Group'].str.contains('#N/A')==False]
        # add back the "Grouping major" column
        MajorMinorGroupList_df = TollclassGrouping_df[['Grouping major', 'Grouping minor','Tollclass Group']]
        MajorMinorGroupList_df = MajorMinorGroupList_df.groupby(['Grouping major', 'Grouping minor'], as_index=False).first()
        MajorMinorGroupList_df = MajorMinorGroupList_df[MajorMinorGroupList_df['Grouping minor'].str.contains('#N/A')==False]
        SimpleToll_MinorGroupingX3TP_df= pd.merge(SimpleToll_MinorGroupingX3TP_df,
                                   MajorMinorGroupList_df,
                                   how='left',
                                   left_on=['Tollclass Group'], 
                                   right_on = ['Tollclass Group'],
                                   indicator=False)
        # keep only the relevant columns
        SimpleToll_MinorGroupingX3TP_df = SimpleToll_MinorGroupingX3TP_df[['Grouping major','Tollclass Group','simplified_toll_AM','simplified_toll_MD', 'simplified_toll_PM']]
        # rename the columns for Generate_tollscsv_from_TollclassGrouping.py
        # SimpleToll_MinorGroupingX3TP_df.rename(columns = {'Tollclass Group'   :'Grouping minor'}, inplace = True) 
        SimpleToll_MinorGroupingX3TP_df.rename(columns = {'simplified_toll_AM':'tollam_da'     }, inplace = True)
        SimpleToll_MinorGroupingX3TP_df.rename(columns = {'simplified_toll_MD':'tollmd_da'     }, inplace = True)
        SimpleToll_MinorGroupingX3TP_df.rename(columns = {'simplified_toll_PM':'tollpm_da'     }, inplace = True)

        output_excelfilename = os.path.join(project_dir, output_dir, "SimpleToll_MinorGroupingX3TP.xlsx")
        wb = openpyxl.Workbook()
        wb.save(output_excelfilename)
        SimpleToll_MinorGroupingX3TP_df.to_excel(output_excelfilename, sheet_name = 'Simple', index=False)         

        TollInputs_df = SimpleToll_MinorGroupingX3TP_df
        
#-------------------
# merge files
#-------------------
# merge the tollclass grouping to the tollcsv_prev_df
new_tollscsv_df = pd.merge(tollcsv_prev_df,
                       TollclassGrouping_df,
                       how='left',
                       left_on=['tollclass'], 
                       right_on = ['tollclass'],
                       indicator=False)

# merge in the new toll values 
new_tollscsv_df = pd.merge(new_tollscsv_df,
                       TollInputs_df,
                       how='left',
                       left_on=['Tollclass Group'], 
                       right_on = ['Tollclass Group'],
                       suffixes=('', '_new'),
                       indicator=False)

#-------------------
# replace the values
#-------------------
# only replace the values for AM, MD and PM


# arterials

#am    
new_tollscsv_df["tollam_da"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_da"])
new_tollscsv_df["tollam_s2"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV2, new_tollscsv_df["tollam_s2"])
new_tollscsv_df["tollam_s3"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV3, new_tollscsv_df["tollam_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the same toll as da
new_tollscsv_df["tollam_vsm"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_vsm"])
new_tollscsv_df["tollam_sml"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_sml"])
new_tollscsv_df["tollam_med"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_med"])
new_tollscsv_df["tollam_lrg"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_lrg"])
    
#pm
new_tollscsv_df["tollpm_da"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_da"])
new_tollscsv_df["tollpm_s2"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV2, new_tollscsv_df["tollpm_s2"])
new_tollscsv_df["tollpm_s3"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV3, new_tollscsv_df["tollpm_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the spme toll as da
new_tollscsv_df["tollpm_vsm"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_vsm"])
new_tollscsv_df["tollpm_sml"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_sml"])
new_tollscsv_df["tollpm_med"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_med"])
new_tollscsv_df["tollpm_lrg"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_lrg"])


#md
if tolls_in_midday == "true": 
    # use simplified tolls 
    new_tollscsv_df["tollmd_da"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_da"])
    new_tollscsv_df["tollmd_s2"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV2, new_tollscsv_df["tollmd_s2"])
    new_tollscsv_df["tollmd_s3"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV3, new_tollscsv_df["tollmd_s3"])
    # trucks (vsm, sml, med and lrg) are assumed to pay the smde toll as da
    new_tollscsv_df["tollmd_vsm"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_vsm"])
    new_tollscsv_df["tollmd_sml"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_sml"])
    new_tollscsv_df["tollmd_med"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_med"])
    new_tollscsv_df["tollmd_lrg"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_lrg"])
else: 
    # set midday tolls to zero 
    new_tollscsv_df["tollmd_da"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), 0, new_tollscsv_df["tollpm_da"])
    new_tollscsv_df["tollmd_s2"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), 0, new_tollscsv_df["tollpm_s2"])
    new_tollscsv_df["tollmd_s3"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), 0, new_tollscsv_df["tollpm_s3"])
    new_tollscsv_df["tollmd_vsm"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), 0, new_tollscsv_df["tollpm_vsm"])
    new_tollscsv_df["tollmd_sml"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), 0, new_tollscsv_df["tollpm_sml"])
    new_tollscsv_df["tollmd_med"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), 0, new_tollscsv_df["tollpm_med"])
    new_tollscsv_df["tollmd_lrg"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), 0, new_tollscsv_df["tollpm_lrg"])

# freeways
#am
new_tollscsv_df["tollam_da"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_da"])
new_tollscsv_df["tollam_s2"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollam_da_new']).astype(float)*DiscountFactor_HOV2, new_tollscsv_df["tollam_s2"])
new_tollscsv_df["tollam_s3"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollam_da_new']).astype(float)*DiscountFactor_HOV3, new_tollscsv_df["tollam_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the same toll as da
new_tollscsv_df["tollam_vsm"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_vsm"])
new_tollscsv_df["tollam_sml"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_sml"])
new_tollscsv_df["tollam_med"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_med"])
new_tollscsv_df["tollam_lrg"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_lrg"])
    
#pm
new_tollscsv_df["tollpm_da"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_da"])
new_tollscsv_df["tollpm_s2"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollpm_da_new']).astype(float)*DiscountFactor_HOV2, new_tollscsv_df["tollpm_s2"])
new_tollscsv_df["tollpm_s3"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollpm_da_new']).astype(float)*DiscountFactor_HOV3, new_tollscsv_df["tollpm_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the spme toll as da
new_tollscsv_df["tollpm_vsm"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_vsm"])
new_tollscsv_df["tollpm_sml"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_sml"])
new_tollscsv_df["tollpm_med"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_med"])
new_tollscsv_df["tollpm_lrg"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_lrg"])


#md
if tolls_in_midday == "true": 
    # use simplified tolls
    new_tollscsv_df["tollmd_da"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_da"])
    new_tollscsv_df["tollmd_s2"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollmd_da_new']).astype(float)*DiscountFactor_HOV2, new_tollscsv_df["tollmd_s2"])
    new_tollscsv_df["tollmd_s3"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollmd_da_new']).astype(float)*DiscountFactor_HOV3, new_tollscsv_df["tollmd_s3"])
    # trucks (vsm, sml, med and lrg) are assumed to pay the smde toll as da
    new_tollscsv_df["tollmd_vsm"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_vsm"])
    new_tollscsv_df["tollmd_sml"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_sml"])
    new_tollscsv_df["tollmd_med"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_med"])
    new_tollscsv_df["tollmd_lrg"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_lrg"])
else: 
    # set midday tolls to zero 
    new_tollscsv_df["tollmd_da"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, 0, new_tollscsv_df["tollpm_da"])
    new_tollscsv_df["tollmd_s2"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, 0, new_tollscsv_df["tollpm_s2"])
    new_tollscsv_df["tollmd_s3"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, 0, new_tollscsv_df["tollpm_s3"])
    new_tollscsv_df["tollmd_vsm"] = np.where(new_tollscsv_df["tollclass"] >= 900000, 0, new_tollscsv_df["tollpm_vsm"])
    new_tollscsv_df["tollmd_sml"] = np.where(new_tollscsv_df["tollclass"] >= 900000, 0, new_tollscsv_df["tollpm_sml"])
    new_tollscsv_df["tollmd_med"] = np.where(new_tollscsv_df["tollclass"] >= 900000, 0, new_tollscsv_df["tollpm_med"])
    new_tollscsv_df["tollmd_lrg"] = np.where(new_tollscsv_df["tollclass"] >= 900000, 0, new_tollscsv_df["tollpm_lrg"])


# rounding the values
new_tollscsv_df["tollam_da"] 	=	new_tollscsv_df["tollam_da"].astype(float).round(4)
new_tollscsv_df["tollam_s2"]  	=	new_tollscsv_df["tollam_s2"].astype(float).round(4)
new_tollscsv_df["tollam_s3"]  	=	new_tollscsv_df["tollam_s3"].astype(float).round(4)
new_tollscsv_df["tollam_vsm"] 	=	new_tollscsv_df["tollam_vsm"].astype(float).round(4)
new_tollscsv_df["tollam_sml"] 	=	new_tollscsv_df["tollam_sml"].astype(float).round(4)
new_tollscsv_df["tollam_med"] 	=	new_tollscsv_df["tollam_med"].astype(float).round(4)
new_tollscsv_df["tollam_lrg"] 	=	new_tollscsv_df["tollam_lrg"].astype(float).round(4)
		
new_tollscsv_df["tollpm_da"]  	=	new_tollscsv_df["tollpm_da"].astype(float).round(4)
new_tollscsv_df["tollpm_s2"]  	=	new_tollscsv_df["tollpm_s2"].astype(float).round(4)
new_tollscsv_df["tollpm_s3"]  	=	new_tollscsv_df["tollpm_s3"].astype(float).round(4)
new_tollscsv_df["tollpm_vsm"] 	=	new_tollscsv_df["tollpm_vsm"].astype(float).round(4)
new_tollscsv_df["tollpm_sml"] 	=	new_tollscsv_df["tollpm_sml"].astype(float).round(4)
new_tollscsv_df["tollpm_med"] 	=	new_tollscsv_df["tollpm_med"].astype(float).round(4)
new_tollscsv_df["tollpm_lrg"] 	=	new_tollscsv_df["tollpm_lrg"].astype(float).round(4)
		
new_tollscsv_df["tollmd_da"]  	=	new_tollscsv_df["tollmd_da"].astype(float).round(4)
new_tollscsv_df["tollmd_s2"]  	=	new_tollscsv_df["tollmd_s2"].astype(float).round(4)
new_tollscsv_df["tollmd_s3"]  	=	new_tollscsv_df["tollmd_s3"].astype(float).round(4)
new_tollscsv_df["tollmd_vsm"] 	=	new_tollscsv_df["tollmd_vsm"].astype(float).round(4)
new_tollscsv_df["tollmd_sml"] 	=	new_tollscsv_df["tollmd_sml"].astype(float).round(4)
new_tollscsv_df["tollmd_med"] 	=	new_tollscsv_df["tollmd_med"].astype(float).round(4)
new_tollscsv_df["tollmd_lrg"]	=	new_tollscsv_df["tollmd_lrg"].astype(float).round(4)
 
# keep only variables that are part of the tolls.csv
new_tollscsv_df.rename(columns = {'facility_name_x':'facility_name'}, inplace = True)
new_tollscsv_df = new_tollscsv_df[['facility_name', 'fac_index', 'tollclass', 'tollseg', 'tolltype', 'use', 'tollea_da', 'tollam_da', 'tollmd_da', 'tollpm_da', 'tollev_da', 'tollea_s2', 'tollam_s2', 'tollmd_s2', 'tollpm_s2', 'tollev_s2', 'tollea_s3', 'tollam_s3', 'tollmd_s3', 'tollpm_s3', 'tollev_s3', 'tollea_vsm', 'tollam_vsm', 'tollmd_vsm', 'tollpm_vsm', 'tollev_vsm', 'tollea_sml', 'tollam_sml', 'tollmd_sml', 'tollpm_sml', 'tollev_sml', 'tollea_med', 'tollam_med', 'tollmd_med', 'tollpm_med', 'tollev_med', 'tollea_lrg', 'tollam_lrg', 'tollmd_lrg', 'tollpm_lrg', 'tollev_lrg', 'toll_flat']]

# main output 
output_filename4 = os.path.join(project_dir, output_dir, "tolls_simplified.csv")
new_tollscsv_df.to_csv(output_filename4, header=True, index=False)

# -------------------------------------------
#
# copy other files needed for Tableau mapping
#
# -------------------------------------------

# Tableau workbook
source = "X:/travel-model-one-master/utilities/NextGenFwys/toll_plan_creation/Map_simplified_tolls.twb"
destination = os.path.join(project_dir, output_dir, "Map_simplified_tolls.twb")
shutil.copy(source, destination)

# Make a local copy of the tollclass designation file
source = "X:/travel-model-one-master/utilities/NextGenFwys/TOLLCLASS_Designations.xlsx"
destination = os.path.join(project_dir, output_dir, "TOLLCLASS_Designations.xlsx")
shutil.copy(source, destination)

# shapefiles from the dynamic toll run
source = os.path.join(project_dir,"Scenarios", modelrun_with_DynamicTolling, "OUTPUT", "Shapefile", "network_links.shp") 
destination = os.path.join(project_dir, output_dir, "network_links_100plusSeg.shp")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_DynamicTolling, "OUTPUT", "Shapefile", "network_links.cpg") 
destination = os.path.join(project_dir, output_dir, "network_links_100plusSeg.cpg")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_DynamicTolling, "OUTPUT", "Shapefile", "network_links.dbf") 
destination = os.path.join(project_dir, output_dir, "network_links_100plusSeg.dbf")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_DynamicTolling, "OUTPUT", "Shapefile", "network_links.prj") 
destination = os.path.join(project_dir, output_dir, "network_links_100plusSeg.prj")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_DynamicTolling, "OUTPUT", "Shapefile", "network_links.shp.xml") 
destination = os.path.join(project_dir, output_dir, "network_links_100plusSeg.shp.xml")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_DynamicTolling, "OUTPUT", "Shapefile", "network_links.shx") 
destination = os.path.join(project_dir, output_dir, "network_links_100plusSeg.shx")
shutil.copy(source, destination)

# shapefiles from the NoProject run
source = os.path.join(project_dir,"Scenarios", modelrun_with_NoProject, "OUTPUT", "Shapefile", "network_links.shp") 
destination = os.path.join(project_dir, output_dir, "network_links_NoProject.shp")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_NoProject, "OUTPUT", "Shapefile", "network_links.cpg") 
destination = os.path.join(project_dir, output_dir, "network_links_NoProject.cpg")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_NoProject, "OUTPUT", "Shapefile", "network_links.dbf") 
destination = os.path.join(project_dir, output_dir, "network_links_NoProject.dbf")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_NoProject, "OUTPUT", "Shapefile", "network_links.prj") 
destination = os.path.join(project_dir, output_dir, "network_links_NoProject.prj")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_NoProject, "OUTPUT", "Shapefile", "network_links.shp.xml") 
destination = os.path.join(project_dir, output_dir, "network_links_NoProject.shp.xml")
shutil.copy(source, destination)

source = os.path.join(project_dir,"Scenarios", modelrun_with_NoProject, "OUTPUT", "Shapefile", "network_links.shx") 
destination = os.path.join(project_dir, output_dir, "network_links_NoProject.shx")
shutil.copy(source, destination)
