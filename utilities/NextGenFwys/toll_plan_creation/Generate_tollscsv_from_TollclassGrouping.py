#-------------------
#
# this script takes a matrix of 18x3 toll inputs and creates a new tolls.csv
#
#-------------------
import os
import pandas as pd
import numpy as np

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


#-------------------
# user settings
#-------------------

# specify toll input file and toll plan option
toll_input_file="L:/Application/Model_One/NextGenFwys/INPUT_DEVELOPMENT/Static_toll_plans/Static_toll_P1b_V05/Reruns_toll_inputs.xlsx"
toll_plan_option="Static"

# specify HOV discounts
# a DiscountFactor of 0.5 means half price; a DiscountFactor of 0 means free; and a DiscountFactor of 1 means no discount.
DiscountFactor_HOV2 = 1.0
DiscountFactor_HOV3 = 0.5

# specify the relationship between arterial tolls and freeway tolls
# For example, if ArterialFactor = 0.2, this script will set arterial tolls to be 20% of freeway tolls
ArterialFactor = 0.2

# specify working directories and run ids
project_dir ="L:/Application/Model_One/NextGenFwys/" 
prev_run_with_tollcsv = "L:/Application/Model_One/NextGenFwys/INPUT_DEVELOPMENT/Networks/NGF_Networks_P2a_AllLaneTollingPlusArterials_ImproveTransit_03/NGF_P2a_AllLaneTollingPlusArterials_ImproveTransit_network_2035/hwy/tolls.csv"
output_dir = "INPUT_DEVELOPMENT/Static_toll_plans/Static_toll_P2b_V01"

#-------------------
# read inputs
#-------------------

#input 1 - a toll plan of 18 tollclass groupings and three time periods
workbook = load_workbook(filename=toll_input_file)
# select the active tab
sheet = workbook[toll_plan_option]
# Read the toll inputs
TollInputs = sheet.values
# Set the first row as the headers for the DataFrame
cols = next(TollInputs)
TollInputs = list(TollInputs)
TollInputs_df = pd.DataFrame(TollInputs, columns=cols)
# drop irrelevant columns
TollInputs_df = TollInputs_df[['Grouping major','Grouping minor', 'tollam_da', 'tollmd_da','tollpm_da']]

# input 2
# the toll plan of a previous run (this is where we'll get the values for the bridge an express lane tolls)
tollcsv_prev_df = pd.read_csv(prev_run_with_tollcsv) 

# input 3
# the file with the tollclass groupings
workbook = load_workbook(filename="X:/travel-model-one-master/utilities/NextGenFwys/TOLLCLASS_Designations.xlsx")
# make "By "Inputs_for_tollcalib" the active tab
sheet = workbook["Inputs_for_tollcalib"]
# Read the results in the "Inputs_for_tollcalib" tab
TollclassGrouping = sheet.values
# Set the first row as the headers for the DataFrame
cols = next(TollclassGrouping)
TollclassGrouping = list(TollclassGrouping)
TollclassGrouping_df = pd.DataFrame(TollclassGrouping, columns=cols)
# drop irrelevant columns
TollclassGrouping_df = TollclassGrouping_df[['tollclass', 'Grouping major','Grouping minor']]

#-------------------
# merge files
#-------------------
# merge the tollclass grouping to the tollcsv_prev_df
new_tollscsv_df = pd.merge(tollcsv_prev_df,
                       TollclassGrouping_df,
                       how='left',
                       left_on=['tollclass'], 
                       right_on = ['tollclass'],
                       indicator=False)

# merge in the new toll values 
new_tollscsv_df = pd.merge(new_tollscsv_df,
                       TollInputs_df,
                       how='left',
                       left_on=['Grouping minor'], 
                       right_on = ['Grouping minor'],
                       suffixes=('', '_new'),
                       indicator=False)

#-------------------
# replace the values
#-------------------
# only replace the values for AM, MD and PM

#freeways
#am
#-----
new_tollscsv_df["tollam_da"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_da"])
new_tollscsv_df["tollam_s2"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollam_da_new']).astype(float)*DiscountFactor_HOV2, new_tollscsv_df["tollam_s2"])
new_tollscsv_df["tollam_s3"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollam_da_new']).astype(float)*DiscountFactor_HOV3, new_tollscsv_df["tollam_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the same toll as da
new_tollscsv_df["tollam_vsm"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_vsm"])
new_tollscsv_df["tollam_sml"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_sml"])
new_tollscsv_df["tollam_med"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_med"])
new_tollscsv_df["tollam_lrg"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollam_da_new']                                    , new_tollscsv_df["tollam_lrg"])

#md
#-----
new_tollscsv_df["tollmd_da"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_da"])
new_tollscsv_df["tollmd_s2"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollmd_da_new']).astype(float)*DiscountFactor_HOV2, new_tollscsv_df["tollmd_s2"])
new_tollscsv_df["tollmd_s3"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollmd_da_new']).astype(float)*DiscountFactor_HOV3, new_tollscsv_df["tollmd_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the smde toll as da
new_tollscsv_df["tollmd_vsm"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_vsm"])
new_tollscsv_df["tollmd_sml"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_sml"])
new_tollscsv_df["tollmd_med"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_med"])
new_tollscsv_df["tollmd_lrg"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollmd_da_new']                                    , new_tollscsv_df["tollmd_lrg"])

#pm
#-----
new_tollscsv_df["tollpm_da"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_da"])
new_tollscsv_df["tollpm_s2"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollpm_da_new']).astype(float)*DiscountFactor_HOV2, new_tollscsv_df["tollpm_s2"])
new_tollscsv_df["tollpm_s3"]  = np.where(new_tollscsv_df["tollclass"] >= 900000, (new_tollscsv_df['tollpm_da_new']).astype(float)*DiscountFactor_HOV3, new_tollscsv_df["tollpm_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the spme toll as da
new_tollscsv_df["tollpm_vsm"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_vsm"])
new_tollscsv_df["tollpm_sml"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_sml"])
new_tollscsv_df["tollpm_med"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_med"])
new_tollscsv_df["tollpm_lrg"] = np.where(new_tollscsv_df["tollclass"] >= 900000, new_tollscsv_df['tollpm_da_new']                                    , new_tollscsv_df["tollpm_lrg"])

           
#arterials
#am
#-----
new_tollscsv_df["tollam_da"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_da"])
new_tollscsv_df["tollam_s2"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV2, new_tollscsv_df["tollam_s2"])
new_tollscsv_df["tollam_s3"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV3, new_tollscsv_df["tollam_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the same toll as da
new_tollscsv_df["tollam_vsm"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_vsm"])
new_tollscsv_df["tollam_sml"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_sml"])
new_tollscsv_df["tollam_med"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_med"])
new_tollscsv_df["tollam_lrg"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollam_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollam_lrg"])

#md
#-----
new_tollscsv_df["tollmd_da"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_da"])
new_tollscsv_df["tollmd_s2"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV2, new_tollscsv_df["tollmd_s2"])
new_tollscsv_df["tollmd_s3"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV3, new_tollscsv_df["tollmd_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the smde toll as da
new_tollscsv_df["tollmd_vsm"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_vsm"])
new_tollscsv_df["tollmd_sml"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_sml"])
new_tollscsv_df["tollmd_med"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_med"])
new_tollscsv_df["tollmd_lrg"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollmd_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollmd_lrg"])

#pm
#-----
new_tollscsv_df["tollpm_da"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_da"])
new_tollscsv_df["tollpm_s2"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV2, new_tollscsv_df["tollpm_s2"])
new_tollscsv_df["tollpm_s3"]  = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor*DiscountFactor_HOV3, new_tollscsv_df["tollpm_s3"])
# trucks (vsm, sml, med and lrg) are assumed to pay the spme toll as da
new_tollscsv_df["tollpm_vsm"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_vsm"])
new_tollscsv_df["tollpm_sml"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_sml"])
new_tollscsv_df["tollpm_med"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_med"])
new_tollscsv_df["tollpm_lrg"] = np.where(((new_tollscsv_df["tollclass"] >= 700000) & (new_tollscsv_df["tollclass"]<900000)), (new_tollscsv_df['tollpm_da_new']).astype(float)*ArterialFactor                    , new_tollscsv_df["tollpm_lrg"])




# write out the new tolls.csv
output_filename = os.path.join(project_dir, output_dir, "tolls_" + toll_plan_option + ".csv")
new_tollscsv_df.to_csv(output_filename, header=True, index=False)